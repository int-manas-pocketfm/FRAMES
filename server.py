#!/usr/bin/env python3
"""
Shot Generation Pipeline — Stage 1 + Stage 2
Single-file backend + web UI.

Run:  python server.py
Open: http://localhost:5000
"""

import json
import os
import queue
import re
import shutil
import subprocess
import sys
import threading
import time
import traceback
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import httpx
from flask import Flask, Response, jsonify, render_template, request, send_file

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


def _load_dotenv():
    env_path = Path(__file__).parent / ".env"
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, val = line.partition("=")
        key = key.strip()
        val = val.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = val

_load_dotenv()


# ── Runtime constants ─────────────────────────────────────────────────────────

ROOT             = Path(__file__).parent
WORKSPACE        = ROOT / "workspace"
WORKSPACE.mkdir(exist_ok=True)

ARGUS_API_KEY    = os.environ.get("ARGUS_API_KEY", "")
ARGUS_BASE_URL   = os.environ.get("ARGUS_BASE_URL", "")
ARGUS_MODEL      = os.environ.get("ARGUS_MODEL", "")
GEMINI_API_KEY   = os.environ.get("GEMINI_API_KEY", "")
MADEYE_API_KEY   = os.environ.get("MADEYE_API_KEY", "")
MADEYE_BASE_URL  = os.environ.get("MADEYE_BASE_URL", "").rstrip("/")
_raw_madeye_keys = os.environ.get("MADEYE_API_KEYS", "")
MADEYE_API_KEYS  = [k.strip() for k in _raw_madeye_keys.split(",") if k.strip()]
if not MADEYE_API_KEYS and MADEYE_API_KEY:
    MADEYE_API_KEYS = [MADEYE_API_KEY]

SUPABASE_URL     = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY     = os.environ.get("SUPABASE_SERVICE_KEY") or os.environ.get("SUPABASE_ANON_KEY", "")

RATE_LIMIT_WAIT  = 60
BATCH_SIZE       = 20
INTER_CALL_PAUSE = 10
STAGGER_SEC      = 5

SHOW_LEVEL_STEMS = {"show_tone_bible", "character_canvas", "location_reference"}
SHOW_LEVEL_FILES = {s + ".md" for s in SHOW_LEVEL_STEMS}


# ── MadEye key pool — thread-safe per-key rate limiting ──────────────────────

class MadEyeKeyPool:
    def __init__(self, keys: list, rate_limit_sec: float = 32.0):
        self._keys = keys
        self._rate_limit = rate_limit_sec
        self._lock = threading.Lock()
        self._last_used = {k: 0.0 for k in keys}
        self._condition = threading.Condition(self._lock)

    def acquire(self, timeout: float = 300.0) -> str:
        deadline = time.time() + timeout
        with self._condition:
            while True:
                now = time.time()
                best_key = None
                best_wait = float('inf')
                for k in self._keys:
                    wait = max(0.0, self._rate_limit - (now - self._last_used[k]))
                    if wait < best_wait:
                        best_wait = wait
                        best_key = k
                if best_wait <= 0:
                    self._last_used[best_key] = now
                    return best_key
                remaining = deadline - now
                if remaining <= 0:
                    raise TimeoutError("No MadEye key available within timeout")
                self._condition.wait(timeout=min(best_wait + 0.1, remaining))

    def release_notify(self):
        with self._condition:
            self._condition.notify_all()

    @property
    def key_count(self) -> int:
        return len(self._keys)


# ── Embedded prompts — Stage 1 ────────────────────────────────────────────────

PROMPTS = {

"show_tone_bible": """\
OUTPUT CONSTRAINT: Produce a complete but concise document. Avoid padding. Under 1500 words.

---

You are a professional visual development coordinator for an AI-assisted video content pipeline. Your job is to read all available episode scripts for a show and produce a comprehensive Show Tone Bible in markdown format.

This file is a permanent show-level reference used across every episode's image generation process. It must capture the complete visual language of the show — color, light, mood, motif — consistently enough that any episode's shots can be generated in a unified style.

---

## OUTPUT FORMAT

---

# Show Tone Bible — [Show Name]

## Show Overview
3–5 sentences. Genre, setting, core emotional register, and the visual world the show inhabits. Written for a cinematographer or visual effects supervisor.

---

## Color Palettes

Define 4–8 named color palettes. Each palette belongs to a context: a character, a faction, a location type, a story phase, or an emotional state.

For each palette:

### [Palette Name] — [What it represents]
- **Primary colors:** [specific color names / hex-style descriptions]
- **Accent colors:** [supporting tones]
- **When to use:** [which scenes, characters, or emotional beats this palette applies to]
- **Lighting temp:** [warm / cool / neutral — and intensity level]
- **Mood:** [what this palette communicates emotionally]

---

## Visual Motifs

List 5–10 recurring visual elements that define the show's imagery. For each:

**[Motif name]**
- What it is: [physical description]
- Where it appears: [which scenes, characters, or locations]
- What it means: [its narrative or thematic function]
- How to render it: [practical image generation note — angle, framing, lighting]

---

## Emotional Registers

Define 4–6 emotional registers the show operates in. Each register is a distinct visual mode.

### [Register Name] — e.g. "Quiet Menace", "Triumphant Ascent", "Shattered Intimacy"
- **Visual cues:** [lighting, framing, color temp, depth of field]
- **Character behavior:** [posture, expression, movement]
- **Environment:** [how the setting changes or contributes]
- **Typical scenes:** [what kinds of moments use this register]

---

## Cinematography Notes

General rules for how this show looks. Write 8–15 bullet points covering:
- Preferred shot distances (close-up heavy? wide establishing shots?)
- Camera angle tendencies (low angle for power? Dutch tilt for unease?)
- Depth of field style (shallow for intimacy? deep for world-building?)
- Transition tendencies (hard cuts? slow dissolves?)
- Any show-specific visual rules
- How action scenes differ from dialogue scenes visually
- How internal/emotional moments are externalized visually

---

## Style Anchor

This line must appear identically in every shot's Style field:

> Style: cinematic, photorealistic, 8K, shallow depth of field, single-frame still image, high detail, film lighting

---

## OUTPUT RULES

1. Do not add any preamble, explanation, or commentary outside the structure above.
2. Do not truncate. Every section must be complete.
3. Base every palette, motif, and register on evidence from the scripts — do not invent elements that aren't present.
4. Output only the markdown document. Nothing before or after it.
""",

"character_canvas": """\
OUTPUT CONSTRAINT: Produce a complete but concise document. Avoid padding. Under 1500 words.

---

You are a professional character visual coordinator for an AI-assisted video content pipeline. Your job is to read all available episode scripts for a show and produce a comprehensive Character Canvas in markdown format.

This file is a permanent show-level reference. Every named character who appears across the show must be documented here with enough visual precision that their appearance can be reproduced consistently across all image generation calls, in any episode, across any scene.

---

## OUTPUT FORMAT

---

# Character Canvas — [Show Name]

## Cast Overview
One paragraph listing all documented characters and their broad role in the show.

---

## Character Profiles

For every named character in the show, produce a full profile using this structure:

---

### [Character Full Name]
**Role:** [Protagonist / Antagonist / Supporting / Mentor / etc.]
**First appears:** [Episode number]

#### Physical Description
| Feature | Description |
|---|---|
| Height & Build | [e.g. tall, lean athletic build] |
| Skin tone | [specific — e.g. deep brown, warm medium tan] |
| Hair | [color, texture, length, typical styling] |
| Eyes | [color and shape] |
| Face | [notable features] |
| Distinguishing marks | [scars, birthmarks, tattoos, other permanent physical markers] |

#### Default Outfit (Signature Look)
- **Top:** [specific garment, color, material, fit]
- **Bottom:** [specific garment, color, material, fit]
- **Footwear:** [specific]
- **Outerwear:** [if applicable]
- **Accessories:** [jewelry, bags, hats, belts — describe each item's position, size, color, material]

#### Recurring Outfit Variations
List any other outfits this character wears regularly.

#### Key Visual Tells
3–5 bullet points. Visual shortcuts that make this character instantly recognizable even in a close-up.

#### Character Arc Note
1–2 sentences on how this character's visual presentation changes across the show.

---

## Relationship Map
| Character A | Relationship | Character B |
|---|---|---|
| [Name] | rival / mentor / lover / enemy / sibling | [Name] |

---

## OUTPUT RULES

1. Do not add any preamble, explanation, or commentary outside the structure above.
2. Do not truncate. Document every named character who appears in the scripts.
3. All descriptions must be visual and physical only — precise enough for image generation.
4. If a character's appearance is not described, infer from role/background and flag with *(inferred)*.
5. Output only the markdown document. Nothing before or after it.
""",

"location_reference": """\
You are a professional production designer for an AI-assisted video content pipeline. Your job is to read all available episode scripts for a show and produce a comprehensive Location Reference File in markdown format.

This file is a permanent show-level reference. Every distinct location that appears across the show must be documented here with enough visual precision that it can be reproduced consistently across all image generation calls.

---

## OUTPUT FORMAT

---

# Location Reference — [Show Name]

## Location Overview
One paragraph listing all documented locations grouped by category.

---

## Location Profiles

For every distinct location in the show:

---

### [Location Name]
**Type:** [Interior / Exterior / Mixed]
**Category:** [e.g. Institutional, Domestic, Urban Street, Natural Environment]
**First appears:** [Episode number]
**Recurring:** [Yes / No / Occasional]

#### Physical Description
- Size and scale
- Key architectural or environmental features
- Condition
- Key furniture or fixed elements
- Entry/exit points and how they're used dramatically

#### Lighting Profile
| Time of Day | Light Source | Quality | Color Temp |
|---|---|---|---|
| Day | [Natural / Artificial / Mixed] | [Harsh / Soft / Dappled] | [Warm / Cool / Neutral] |
| Night | [Natural / Artificial / Mixed] | [Harsh / Soft / Glowing] | [Warm / Cool / Neutral] |

#### Atmosphere & Mood
- **Default mood:** [what this location feels like before any characters enter]
- **Color palette:** [dominant colors of the space]
- **Sound implied:** [what sounds define this space]
- **Emotional register:** [what emotional state this location typically hosts]

#### Dramatic Usage
- Which characters use this location and in what context
- What kinds of scenes happen here
- Any scenes where the visual character changes significantly

#### Image Generation Notes
- Wide establishing shot framing notes
- Recurring close-up elements unique to this location
- Visual elements that must appear in every shot set here

---

## OUTPUT RULES

1. Do not add any preamble or commentary outside the structure above.
2. Do not truncate. Document every distinct location, including one-time appearances.
3. All descriptions must be visual and physical — precise enough for image generation.
4. Infer from context where details are sparse and flag with *(inferred)*.
5. Output only the markdown document.
""",

"episode_detail": """\
You are a professional production coordinator for an AI-assisted video content pipeline. Your job is to read an episode script and produce a structured Episode Detail File in markdown format.

This file is used as context for AI image generation (Gemini). Every section must be specific, visual, and precise — vague descriptions are useless downstream.

---

## OUTPUT FORMAT

---

# Episode Detail File — [Episode Number]: [Episode Title]

## Episode Summary
2–4 sentences. Cover what happens, the emotional core, and how the episode ends.

---

## Key Characters

| Character | Role in Episode | Outfit Description | Key Visual Tells |
|---|---|---|---|
| [Name] | [What they do / their arc this episode] | [Full outfit: top, bottom, footwear, accessories — be specific about colors, materials, fit] | [Physical markers: hair, build, distinguishing features, recurring accessories] |

Rules:
- Outfit must be specific enough to recreate consistently. "Blue dress" is not enough. "Floor-length cobalt silk dress with off-shoulder neckline, fitted bodice, no jewelry" is correct.
- If a character changes outfits, create a separate row per outfit labeled "Outfit A", "Outfit B".
- Include every character who appears on screen, even briefly.

---

## Key Locations

| Location | Description | Lighting / Atmosphere | Color Temperature |
|---|---|---|---|
| [Location name] | [Physical description: size, layout, key features] | [Time of day, natural vs artificial light, mood] | [Warm / cool / neutral] |

---

## Key Props

| Prop | Description | Scene Context |
|---|---|---|
| [Prop name] | [Precise physical description: size, shape, color, material, condition] | [Which scene and its significance] |

Only include props that are visually prominent or narratively significant.

---

## Key Visual Moments

List the 5–10 most visually distinct and emotionally significant moments.

**[Moment number]. [Brief title]**
- Scene: [where it happens]
- Characters: [who is present]
- What happens visually: [describe as if describing a still image]
- Why it matters: [narrative or emotional significance]

---

## Tone Arc

- **Opening tone:** [e.g. tense, playful, melancholic — and what creates that feeling visually]
- **Mid-episode shift:** [what changes and how it registers on screen]
- **Closing tone:** [how the episode ends emotionally]

---

## Continuity Flags

List details that must stay consistent with prior or future episodes. If none, write: *No continuity flags for this episode.*

---

## Shot Description Notes

3–6 bullet points:
- Visual effects or non-realistic elements needing translation into physical visible cues
- Scenes where close-ups are essential
- Scenes where environment must dominate
- Recurring visual motifs in this episode
- Crowd / large cast framing notes

---

## OUTPUT RULES

1. Do not add any preamble or commentary outside the structure above.
2. Do not truncate. Every section must be complete.
3. Write in present tense throughout.
4. All descriptions must be visual and physical — no metaphor in description fields.
5. Output only the markdown document.
""",

}  # end PROMPTS


# ── Embedded prompts — Stage 2 ────────────────────────────────────────────────

PIPELINE_INSTRUCTIONS = """\
# Shot Breakdown Pipeline — Project Instructions

## Role
You are a professional shot breakdown assistant for AI image generation. Your job is to take raw scripts from any show or IP and transform them into fully structured, production-ready shot breakdown sheets — row by row, exactly like a film storyboard supervisor would.

You understand cinematography, visual storytelling, location design, and AI image generation (Gemini). You always apply the RULEBOOK to every breakdown. Every Shot Description you write must be a GenAI-ready prompt specific enough for Gemini to generate a single still image without ambiguity.

---

## The 5-Step Pipeline

Always follow these steps in order. Never skip a step.

---

### Step 1 — Script Chunker
Break the raw script into shots based on visual logic — not sentence count, not timing.

**The core rule: A shot = one distinct visual moment that can be rendered as a single still image by Gemini.**

**THE IMAGE LOGIC RULE — THE MOST IMPORTANT RULE IN THIS STEP:**
Each shot will become one generated image. The question is never "how long does this play?" — it is: "does this beat deserve its own image?"

Split aggressively. Every time the script introduces a new subject, a new action, a new emotion, or a new visual detail that would look different as a still image — that is a new shot. When in doubt, split.

**LENGTH RULE — APPLY BEFORE ANYTHING ELSE:**
If a chunk of text is more than 30 words, it almost certainly contains more than one visual moment. Stop. Re-read it clause by clause and split it.

A single shot line should describe one thing happening to one subject in one emotional register. If you can identify two subjects, two actions, or two emotions in the same line — it must be split.

**HOW TO SPLIT A LONG SENTENCE:**
A sentence can contain multiple visual moments even if it has no paragraph break. Split at every clause that introduces something new — a new subject, a new action, a new emotion.

Example:
✗ One shot (too long — three distinct visual moments crammed together):
"She had finally made her decision, her eyes cold and resolved. He stood in the doorway, blocking her path, arms crossed and jaw tight. Neither of them moved, the silence between them heavy with everything unsaid."

✓ Correctly split into three shots:
- Shot A: "She had finally made her decision, her eyes cold and resolved." — Her face. The internal shift from doubt to certainty.
- Shot B: "He stood in the doorway, blocking her path, arms crossed and jaw tight." — His body filling the frame.
- Shot C: "Neither of them moved, the silence between them heavy with everything unsaid." — Two figures. The standoff.

**ASK FOR EVERY SENTENCE OR CLAUSE:**
- Does this introduce a new visual subject? → new shot
- Does this show a different action or physical state than the line before? → new shot
- Does this shift the emotional register? → new shot
- Does this describe a new environment, detail, or atmospheric element? → new shot
- Does this line contain more than 30 words? → almost certainly needs splitting
- Would Gemini produce the same image for this line as the previous one? → only then keep grouped

**TARGET SHOT COUNT — USE AS A SELF-CHECK, NOT A CUTTING RULE:**
A well-split script produces approximately 7–9 shots per 100 words. For a 1000-word script: 70–90 shots.

**CRITICAL — ZERO TOLERANCE: Copy the exact original script text for each chunk — word for word, no paraphrasing, no summarizing, no skipping. The concatenation of ALL line fields MUST reproduce the ENTIRE script with EVERY word present. If you combine the `line` field from shot 1 + shot 2 + shot 3 + ... it must equal the full original script text. Missing even one sentence will cause downstream audio/subtitle failures.**

---

### Step 2 — Object & Action Mapping
For each shot line from Step 1, identify:
- **Objects:** Key visual elements, props, characters, creatures, vehicles
- **Actions:** What is physically happening in the frame
- **Character anchor:** Visual identifiers per Rule 14

---

### Step 3 — Camera Assignment
For each shot, assign:

**Shot Size:** Extreme Close-up | Close-up | Medium Close-up | Medium | Wide | Wide (Profile) | Wide tracking | Extreme Wide | POV | Over-the-shoulder

**Camera Angle:** Eye-level | Low angle | High angle | Dutch angle | Bird's eye | Worm's eye

**Camera Movement:** Static | Pan | Tilt | Dolly in | Dolly out | Track | Handheld | Crane up | Crane down

**Shot Description** must follow the 8-part formula from Rule 12:
[Scene Type] + [Character State] + [Action Moment] + [Environment Reaction] + [Camera Language] + [Lighting] + [Lens] + [Mood]

**SENSORY & POV SHOTS — CRITICAL RULE:**
When the script describes something a character experiences internally — pain, light flashing, sound hitting them — the Shot Description must always anchor to the character's physical face or body. The internal experience is rendered as a visible effect on them, never as a standalone abstract image.

**OUTFIT RULE — CRITICAL (see also Rule 14B):**
The character's outfit must be stated explicitly inside the Shot Description prompt itself with FULL color and fabric details. Every Shot Description that includes a character must contain their outfit. Never use generic labels like "sleepwear", "formal attire", "casual clothes" — always specify exact garments with colors (e.g. "cream silk nightgown with lace trim", "charcoal wool three-piece suit with white cotton dress shirt and navy silk tie"). Once you write a character's outfit in their first shot of a scene, copy that EXACT outfit string into every subsequent shot of that character in the same scene.

**BODY POSITION ANCHOR ON CLOSE-UPS — CRITICAL:**
Even on extreme close-ups, state: (1) the character's full body position, (2) a soft background anchor describing the location.

**ABSTRACT CONCEPT RULE — CRITICAL:**
Internal mental events must be translated into physical reactions on the character's face or body only. Never use abstract visual metaphors, overlays, or projections.

Assign **Intensity Level** (1–4) to each shot:
- Level 1 = Social humiliation / public shame
- Level 2 = Internal conflict / emotional hesitation
- Level 3 = Power activation / transformation beginning
- Level 4 = Peak power / monster state / full transformation

---

### Step 4 — Location & Environment Setup
For each shot:
- **Environment:** Short location name, 1–3 words
- **Time of Day:** Dawn / Morning / Day / Dusk / Night
- **Lighting:** Specific — apply Color Strategy from Rule 5
- **Background:** What is visible behind the subject
- **Environment Reaction:** What the environment is doing (Rule 7)
- **Color Palette:** Dominant colors matching the emotional beat
- **Mood:** Atmospheric quality

---

### Step 5 — Continuity Validator
Check: object/character continuity, location consistency, camera flow, rulebook compliance (all 18 rules), prompt quality.

Report: ✓ working well, ⚠️ issues (shot number + problem), 🔧 recommended fix, 🚩 rulebook violations.
**Readiness Score: X/100**

---

## Final Output — The Shot Breakdown Sheet

After all 5 steps, produce the final sheet.

**IMPORTANT OUTPUT FORMAT:**
When asked to produce the Excel output (triggered by "MAKE EXCEL"), output each batch of 25 shots as a JSON array wrapped in ```json and ``` markers. Each shot object must have exactly these keys:
- shot_number (integer)
- line (string — exact original script text, verbatim)
- shot_size (string)
- shot_description (string — full 8-part GenAI prompt)
- shot_detail (string — Style Anchor + Part 1 bullets + Character & Location Details + Show Tone, newlines as \\n)
- reference (string — camera angle)

End each batch with the text: BATCH COMPLETE — shots [X]–[Y] done.
After the final batch: ALL SHOTS COMPLETE — [X] total shots delivered.

---

## Shot Detail Column Structure

Every shot_detail value must contain three parts in this exact order:

**STYLE ANCHOR (first line, every shot, no exceptions):**
Style: cinematic, photorealistic, 8K, shallow depth of field, single-frame still image, high detail, film lighting

**PART 1 — Shot Breakdown Bullets:**
• Location -> [location name]
• Character -> [CHARACTER NAME] | Action: [what they are physically doing] | Emotion: [what they are feeling]
• Character State -> [visible expression or body language]
• Lighting -> [light description]
• Camera Angle ([angle]) + Camera Movement ([movement]) -> [plain English description]
• Lens -> [lens]
• Mood -> [atmosphere phrase]

**PART 2 — Character & Location Details:**
[CHARACTER NAME]: [build + age], [hair], [face/eyes], wearing [full outfit — exact garments with explicit colors and fabric, e.g. "navy linen blazer over white cotton henley and dark brown leather belt with khaki chinos" NOT "smart casual"], [wounds/marks if visible], [pose/action in this shot]
NOTE: The outfit string here must be IDENTICAL across all shots of this character in the same scene. Copy-paste, do not rewrite.

[LOCATION NAME]: [brief type descriptor]. [Architecture/features]. [Time of day], [light source and quality], [color temperature]. Background: [what is visible]. [Atmosphere or "No atmospheric effects"]. Mood: [one evocative phrase].

**PART 3 — Show Tone:**
Color Palette: [derived from loaded show tone file — specific colors, quality of light, color temperature]
Visual Motif: [specific motif from show tone file, or "None"]
Emotional Register: [tone state from show tone file — specific phrase tied to this shot's beat]

---

## How to Trigger the Pipeline

**STAGE 1 — BREAKDOWN** (triggered by pasting a raw script):
Run all 5 steps internally. Before declaring complete, execute the SCRIPT COVERAGE AUDIT PROTOCOL below.

---

## SCRIPT COVERAGE AUDIT PROTOCOL — MANDATORY

Before confirming any shot breakdown is complete, you MUST do the following. No exceptions.

**Audit Step 1 — Source Script**: Use ONLY the original script text provided in the user message as your source of truth. Never rely on memory or prior transcription.

**Audit Step 2 — Parse into Beats**: Split the script into every distinct sentence or narrative unit. Each sentence is a beat. Count them. Every clause joined by commas or conjunctions must be fully preserved.

**Audit Step 3 — Check Every Beat**: For each script beat, verify it appears in at least one shot's Line column — either as the full line or as a confirmed sub-sentence split. Every word must be present. Do not use fuzzy matching — verify exact text.

**Audit Step 4 — Check for Duplicates**:
- Exact duplicates: same Line text on two or more shots — REMOVE the duplicate.
- Partial duplicates: one shot's Line is a substring of another shot's Line — trim the parent shot's line to only the portion not covered by the child shots.

**Audit Step 5 — Confirm Coverage**: Do NOT say "BREAKDOWN COMPLETE" until Step 3 shows zero genuine missing beats AND Step 4 shows zero unintentional duplicates.

**Common Failure Modes to Avoid**:
- Auditing the breakdown against itself — will never catch missing lines
- Merged lines that bundle multiple beats — counts as covered only if ALL sub-beats are present
- Lines deleted as "duplicates" that were the only coverage of a distinct script beat — always check before deleting
- Mid-sentence clause dropping — when a sentence has multiple clauses (e.g. "she did X, feeling Y"), EVERY clause must be present in the Line field

---

When complete, output ONLY:
```
BREAKDOWN COMPLETE
Episodes analysed: [episode name]
Total shots: [X]
Script coverage: [X]% (every sentence accounted for)
Rulebook violations: [X — list briefly, or "None"]
Ready for output. Type MAKE EXCEL to generate the sheet.
```

**STAGE 2 — OUTPUT** (triggered by "MAKE EXCEL"):
Output 25 shots at a time as JSON arrays (see format above). Wait for "NEXT" between batches.

**OTHER COMMANDS:**
- "STEP 1 ONLY" — run and output Step 1 chunk list only
- "VALIDATE" — run Step 5 on current breakdown
"""


RULEBOOK = """\
# GenAI Cinematic Promo Rulebook
Fantasy / Action IP | Social-First | Meta Optimized | High-Retention Structure

---

## 1. CORE OBJECTIVE

Primary Goal — Maximize: 3-sec plays, 15-sec Thruplays, hook retention, emotional escalation, curiosity gap.

Target Audience: USA | 18–55 | Blue-collar mindset — emotion-driven, direct stakes, power fantasy, betrayal/revenge/underdog arcs.

---

## 2. STRUCTURE RULES

### Rule 1: Start With Conflict, Not Context
Never open with exposition. First shot must be: humiliation, violence/physical threat, public embarrassment, power reveal, betrayal, or imminent danger. Hook must visually escalate in the first 2.5 seconds.

### Rule 2: Every 3–5 Seconds = Escalation
No flat emotional plateaus. Each beat must increase: danger, humiliation, power, chaos, or mystery. If nothing changes — cut it.

### Rule 3: Character Framing Hierarchy
| Power State | Camera Rule |
|---|---|
| Dominant / powerful | Low angle |
| Vulnerable / weak | High angle |
| Turning point / gaining power | Eye-level push-in |
| Fully transformed / peak power | Extreme close-up + shadow |
| Two characters at equal tension | Eye-level, static |

---

## 3. SHOT DIVISION RULES

Average shot length: 1.2–2.5 seconds. Action beats: 0.5–1 second. Emotional beats: 2–3 seconds max. No shots longer than 3 seconds without escalation.

---

## 4. DIALOGUE RULES

Dialogue must cut like a punch. Short. Punchy. Each word lands. Layer: primary dialogue + system/inner voice + crowd chant + low sub bass.

---

## 5. COLOR STRATEGY

| Emotional Beat | Color Direction |
|---|---|
| Humiliation | Cold blue |
| Betrayal | Desaturated / washed out |
| Rage | Warm red / orange |
| Power activation | High contrast + glow |
| Calm dominant character | Neutral with dramatic shadow contrast |
| Mystery / unknown threat | Dark, low saturation, single accent color |

---

## 6. POWER REVEAL — MANDATORY BUILD ORDER
1. Micro physical twitch or reaction
2. Sound distortion
3. Steam / heat / energy emanating from body
4. Extreme close-up on eyes
5. Environmental reaction (objects move, lights flicker, crowd reacts)
6. Full payoff (transformation, ability, destruction)
Never jump straight to the payoff.

---

## 7. ENVIRONMENT CHAOS RULE
The environment is a character. It must react. Every high-intensity shot must include at least one: wind whipping clothing/hair, dust/debris lifting, sparks/embers floating, smoke/steam in frame, flickering lights. If the environment is still — the image feels fake.

---

## 8. DOMINANT CALM CHARACTER RULE
Minimal physical movement. Controlled expression. Shadow framing. Reacts last, speaks least. Should look dangerous without trying. They feel mythic — the scene bends around them.

---

## 9. TRANSITION RULEBOOK
Rage → Chaos: environmental element morphs into next scene.
Close-up eye → Memory: cut on blink.
Crowd chant → Internal voice: crowd audio fades into inner whisper.
Steam Match Cut: end shot steam fills frame → next shot smoke clearing.
Eye Glow Transition: extreme close-up glowing eye → match cut to same color light source.

---

## 10. META PERFORMANCE RULES

First 3 seconds — must show at least one: character with clear power, character being publicly humiliated, imminent physical danger.

0–15 seconds — must include: 2+ visible power escalations, 1 emotional betrayal/stakes moment, 1 unresolved curiosity hook.

End Rule: Always end on unresolved tension. Never full resolution.

---

## 11. PROTAGONIST ARC RULE
Lead character must visually pass through these states in order:
1. Weak / normal
2. Humiliated / threatened
3. Pushed to the edge
4. Overwhelmed / breaking point
5. Dangerous / transformed

---

## 12. GENAI PROMPT FORMULA
[Scene Type] + [Character State] + [Action Moment] + [Environment Reaction] + [Camera Language] + [Lighting] + [Lens] + [Mood]

Intensity Levels:
- Level 1 — Social Humiliation: cold lighting, crowd, desaturation
- Level 2 — Internal Conflict: tight close-up, heat distortion, tension building
- Level 3 — Power Activation: steam, sparks, energy glow, environmental reaction
- Level 4 — Peak Power/Monster State: extreme close-up, face in shadow, transformation detail

---

## 13. CAMERA LANGUAGE BY POWER STATE

Dominant/Powerful: low-angle, subtle slow push-in, sharp focus, strong jawline shadow.
Vulnerable/Weak: slightly high-angle, wider frame, softer lighting, background overpowering subject.
Turning Point: eye-level, slow push-in, shallow depth of field, background fades, light catching eyes.

---

## 14. CHARACTER VISUAL CONSISTENCY RULE
Always repeat in every prompt: hair style/color, clothing color/style, eye color, physical build, signature accessory. Never rephrase core character identifiers shot-to-shot.

---

## 14B. OUTFIT EXPLICITNESS & SCENE CONSISTENCY RULE

**Explicitness:** Never use generic outfit labels like "sleepwear", "formal wear", "casual clothes", or "uniform". Always specify the FULL outfit with explicit colors, fabric, and cut — e.g. "cream silk nightgown with lace trim" not "sleepwear", "charcoal tailored three-piece suit with white dress shirt and navy silk tie" not "formal attire".

**Scene consistency:** Within the same scene, every shot of the same character MUST use the IDENTICAL outfit description — word for word, same colors, same garments, same details. When you write the first shot of a character in a scene, that outfit description is LOCKED for every subsequent shot of that character in that scene. Copy-paste it, do not rewrite it.

**How to apply:**
1. First shot of a character in a scene → write the full explicit outfit (colors, fabric, cut, accessories).
2. Every subsequent shot of that character in the SAME scene → copy that exact outfit string into the Shot Description and Shot Detail. Do not abbreviate, do not synonym-swap colors.
3. Outfit changes ONLY when the script explicitly describes a change or a new scene begins.

---

## 15. EDIT-AWARE PROMPTING RULE
Avoid: ultra-wide scenic compositions, slow static framing, character standing doing nothing.
Prefer: mid-action moments, tight faces with expression, movement implied in frame, cropped tension.
Rule: "The frame should feel like it's already in motion."

---

## 16. GENAI QUALITY CONTROL CHECKLIST
- [ ] Does it show movement or implied motion?
- [ ] Does lighting reflect the emotional state?
- [ ] Is power hierarchy visible through camera angle?
- [ ] Is the environment reacting?
- [ ] Is this scroll-stopping within 1 second?
If 2 or more answers are weak — regenerate.

---

## 17. PROMO ARC PROMPT FLOW

| Beat | Intensity Level | Key Elements |
|---|---|---|
| Opening humiliation | Level 1 | Cold lighting, crowd, desaturation |
| Crowd chaos | Level 1–2 | Handheld, chanting, phones recording |
| Internal conflict | Level 2 | Tight frame, heat distortion |
| Power activation begins | Level 3 | Steam, sparks, energy glow |
| Eyes reveal | Level 3–4 | Extreme close-up, glowing eyes |
| Full transformation | Level 4 | Shadow dominance, extreme close-up |
| Threat / final line | Level 4 | Low angle, controlled power, unresolved |

---

## 18. FINAL CHECKLIST BEFORE EXPORT
- [ ] Hook lands in first 2 seconds — conflict, power, or humiliation
- [ ] No shot exceeds 3 seconds without escalation
- [ ] Power reveal builds across 6 stages — never instant
- [ ] Environment reacts in every high-intensity shot
- [ ] Dominant calm character feels mythic, not loud
- [ ] End leaves unresolved tension — no closure
- [ ] Every Shot Description follows the 8-part prompt formula
- [ ] Character visual identifiers are consistent across all shots
"""

STAGE2_SYSTEM = f"{PIPELINE_INSTRUCTIONS}\n\n---\n\n{RULEBOOK}"


# ── Supabase helpers ─────────────────────────────────────────────────────────

def _supabase(method: str, path: str, data=None, params=None):
    if not SUPABASE_URL or not SUPABASE_KEY:
        raise RuntimeError("SUPABASE_URL / SUPABASE_SERVICE_KEY not set in .env")
    headers = {
        "apikey":        SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type":  "application/json",
        "Prefer":        "return=representation",
    }
    r = httpx.request(
        method, f"{SUPABASE_URL}/rest/v1{path}",
        headers=headers, json=data, params=params, timeout=30,
    )
    r.raise_for_status()
    return r.json() if r.content else []


def _classify_file_type(filename: str) -> str:
    n = filename.lower()
    if "show_tone_bible" in n or "tone_bible" in n: return "show_tone_bible"
    if "character_canvas" in n:                      return "character_canvas"
    if "location_reference" in n:                    return "location_reference"
    if "episode_detail" in n:                        return "episode_detail"
    return "other"


# ── Flask app + job store ─────────────────────────────────────────────────────

app  = Flask(__name__)
jobs: dict = {}


# ── Per-thread logging ────────────────────────────────────────────────────────

_job_local  = threading.local()
_print_lock = threading.Lock()


def _log(msg: str):
    with _print_lock:
        print(msg, flush=True)
    q = getattr(_job_local, "queue", None)
    if q is not None:
        q.put(str(msg))


def _set_job_context(job_id: str, q):
    _job_local.job_id = job_id
    _job_local.queue  = q


# ── API helpers ───────────────────────────────────────────────────────────────

def _track_tokens(data: dict):
    usage = data.get("usage", {})
    jid   = getattr(_job_local, "job_id", None)
    if jid and jid in jobs:
        with jobs[jid]["tokens_lock"]:
            jobs[jid]["tokens"]["input"]  += usage.get("prompt_tokens", 0)
            jobs[jid]["tokens"]["output"] += usage.get("completion_tokens", 0)
            jobs[jid]["tokens"]["calls"]  += 1

# Pricing — auto-detect from ARGUS_MODEL
_MODEL_PRICING = {
    "opus":     (5.0, 25.0),   # Opus 4.5 / 4.6 / 4.7
    "sonnet":   (3.0, 15.0),   # Sonnet 4.5 / 4.6
    "haiku":    (1.0, 5.0),    # Haiku 4.5
}
def _detect_pricing():
    m = ARGUS_MODEL.lower()
    for key, (inp, out) in _MODEL_PRICING.items():
        if key in m:
            return inp, out
    return 3.0, 15.0  # default to Sonnet pricing
LLM_INPUT_COST_PER_MTOK, LLM_OUTPUT_COST_PER_MTOK = _detect_pricing()


def verify_auth():
    _log("  Verifying API credentials...")
    if not ARGUS_API_KEY or not ARGUS_BASE_URL:
        raise RuntimeError("ARGUS_API_KEY or ARGUS_BASE_URL not set in .env")
    try:
        r = httpx.post(
            ARGUS_BASE_URL,
            headers={"Authorization": f"Bearer {ARGUS_API_KEY}", "Content-Type": "application/json"},
            json={"model": ARGUS_MODEL, "messages": [{"role": "user", "content": "Reply with one word: READY"}], "max_tokens": 10},
            timeout=60,
        )
        r.raise_for_status()
        text = r.json()["choices"][0]["message"]["content"].strip()
        if text:
            _log(f"  ok  (model: {ARGUS_MODEL})")
            return
        raise RuntimeError("Empty response from API")
    except httpx.HTTPStatusError as e:
        if e.response.status_code == 401:
            raise RuntimeError("Invalid API key")
        raise RuntimeError(f"Auth check failed: {e.response.status_code}")
    except RuntimeError:
        raise
    except Exception as e:
        raise RuntimeError(f"Auth check failed: {e}")


def call_api(prompt: str, label: str = "", timeout=None, retries: int = 2) -> str:
    """Single-turn API call (Stage 1)."""
    label_str = f"[{label}] " if label else ""
    headers   = {"Authorization": f"Bearer {ARGUS_API_KEY}", "Content-Type": "application/json"}
    payload   = {"model": ARGUS_MODEL, "messages": [{"role": "user", "content": prompt}], "max_tokens": 4096}

    for attempt in range(1, retries + 2):
        _log(f"  ... {label_str}attempt {attempt}/{retries + 1}...")
        try:
            r = httpx.post(ARGUS_BASE_URL, headers=headers, json=payload, timeout=timeout)
            r.raise_for_status()
            data = r.json()
            text = data["choices"][0]["message"]["content"].strip()
            if text:
                _log(f"  ok  {label_str}done")
                _track_tokens(data)
                return text
            _log("  x  Empty response")
        except httpx.HTTPStatusError as e:
            if e.response.status_code == 429:
                _log(f"  !!  Rate limit — waiting {RATE_LIMIT_WAIT}s...")
                time.sleep(RATE_LIMIT_WAIT)
            elif e.response.status_code == 401:
                raise RuntimeError("Invalid API key")
            else:
                body = e.response.text[:400] if e.response.text else ""
                _log(f"  x  HTTP {e.response.status_code}  {body}")
        except httpx.TimeoutException:
            _log(f"  x  Connection/read timeout")
        except Exception as e:
            _log(f"  x  {e}")

        if attempt <= retries:
            wait = 20 * attempt
            _log(f"  -> Retrying in {wait}s...")
            time.sleep(wait)

    _log(f"  x  All attempts failed: {label}")
    return ""


def call_api_chat(
    system: str, messages: list,
    label: str = "", timeout=None, retries: int = 2,
) -> str:
    """Multi-turn streaming chat API call (Stage 2)."""
    label_str = f"[{label}] " if label else ""
    headers   = {"Authorization": f"Bearer {ARGUS_API_KEY}", "Content-Type": "application/json"}
    payload   = {
        "model":         ARGUS_MODEL,
        "messages":      [{"role": "system", "content": system}] + messages,
        "max_tokens":    8192,
        "stream":        True,
        "stream_options": {"include_usage": True},
    }

    for attempt in range(1, retries + 2):
        _log(f"  ... {label_str}attempt {attempt}/{retries + 1} (streaming)...")
        try:
            with httpx.stream("POST", ARGUS_BASE_URL, headers=headers, json=payload, timeout=timeout) as r:
                if r.status_code >= 400:
                    body = r.read().decode("utf-8", errors="replace")[:400]
                    if r.status_code == 429:
                        _log(f"  !!  Rate limit — waiting {RATE_LIMIT_WAIT}s...")
                        time.sleep(RATE_LIMIT_WAIT)
                    elif r.status_code == 401:
                        raise RuntimeError("Invalid API key")
                    else:
                        _log(f"  x  HTTP {r.status_code}  {body}")
                else:
                    full_text    = ""
                    usage        = None
                    started      = time.time()
                    last_log     = started
                    got_first    = False

                    for raw in r.iter_lines():
                        line = raw if isinstance(raw, str) else raw.decode("utf-8", errors="replace")
                        if not line or not line.startswith("data:"):
                            continue
                        data_str = line[5:].strip()
                        if data_str == "[DONE]":
                            break
                        try:
                            chunk_data = json.loads(data_str)
                        except json.JSONDecodeError:
                            continue

                        choices = chunk_data.get("choices") or []
                        if choices:
                            delta = choices[0].get("delta") or {}
                            piece = delta.get("content") or ""
                            if piece:
                                if not got_first:
                                    _log(f"  ... {label_str}streaming started")
                                    got_first = True
                                full_text += piece

                        if chunk_data.get("usage"):
                            usage = chunk_data["usage"]

                        now = time.time()
                        if now - last_log > 12 and got_first:
                            _log(f"  ... {label_str}{len(full_text):,} chars streamed ({int(now - started)}s)")
                            last_log = now

                    if full_text.strip():
                        _log(f"  ok  {label_str}done ({len(full_text):,} chars in {int(time.time() - started)}s)")
                        if usage:
                            _track_tokens({"usage": usage})
                        else:
                            # Rough fallback when upstream doesn't emit usage: ~4 chars/token
                            _track_tokens({"usage": {
                                "prompt_tokens":     0,
                                "completion_tokens": len(full_text) // 4,
                            }})
                        return full_text.strip()
                    _log(f"  x  {label_str}empty stream")

        except httpx.TimeoutException:
            _log(f"  x  Connection/read timeout")
        except RuntimeError:
            raise
        except Exception as e:
            _log(f"  x  {e}")

        if attempt <= retries:
            wait = 20 * attempt
            _log(f"  -> Retrying in {wait}s...")
            time.sleep(wait)

    _log(f"  x  All attempts failed: {label}")
    return ""


# ── Upload text extraction (handles .docx) ───────────────────────────────────

def _read_upload_as_text(file_storage) -> str:
    """Read an uploaded FileStorage as text. Converts .docx → markdown via mammoth."""
    filename = (file_storage.filename or "").lower()
    raw      = file_storage.read()
    if filename.endswith(".docx"):
        try:
            import io, mammoth
            return mammoth.convert_to_markdown(io.BytesIO(raw)).value.strip()
        except ImportError:
            raise RuntimeError("pip install mammoth — needed to read .docx")
        except Exception as e:
            raise RuntimeError(f"Failed to convert {file_storage.filename}: {e}")
    return raw.decode("utf-8", errors="replace")


# ── Stage 1 helpers ───────────────────────────────────────────────────────────

def _convert_docx(src: Path, dest: Path) -> bool:
    try:
        import mammoth
        with open(src, "rb") as f:
            res = mammoth.convert_to_markdown(f)
        text = res.value.strip()
        if not text:
            _log(f"  !  Empty output for {src.name} — skipping")
            return False
        dest.write_text(text, encoding="utf-8")
        return True
    except ImportError:
        _log("  x  pip install mammoth  (needed for .docx)")
        return False
    except Exception as e:
        _log(f"  x  {src.name}: {e}")
        return False


def auto_convert(scripts_dir: Path):
    to_convert = [
        f for ext in ["*.docx", "*.txt", "*.fountain"]
        for f in scripts_dir.glob(ext)
        if not f.with_suffix(".md").exists()
    ]
    if not to_convert:
        return
    _log(f"  Converting {len(to_convert)} script(s) to .md...")
    for src in to_convert:
        dest = src.with_suffix(".md")
        if src.suffix == ".docx":
            ok = _convert_docx(src, dest)
        else:
            try:
                dest.write_text(src.read_text(encoding="utf-8"), encoding="utf-8")
                ok = True
            except Exception as e:
                _log(f"  x  {src.name}: {e}")
                ok = False
        _log(f"  {'ok' if ok else 'x '}  {src.name} -> {dest.name}")


def load_prompt(name: str) -> str:
    if name not in PROMPTS:
        raise RuntimeError(f"Unknown prompt: {name}")
    return PROMPTS[name]


def read_scripts(scripts_dir: Path, filenames: list = None) -> dict:
    if filenames:
        files   = [scripts_dir / f for f in filenames]
        missing = [str(f) for f in files if not f.exists()]
        if missing:
            raise RuntimeError(f"Scripts not found: {missing}")
    else:
        files = sorted(scripts_dir.glob("*.md"))
        if not files:
            raise RuntimeError(f"No .md scripts in: {scripts_dir}")
    return {f.stem: f.read_text(encoding="utf-8") for f in files}


def trim(text: str, max_chars: int = 30_000) -> str:
    if len(text) <= max_chars:
        return text
    cut = text[:max_chars].rfind("\n\n")
    if cut < max_chars * 0.7:
        cut = max_chars
    return text[:cut] + "\n\n[...trimmed...]"


def call_claude(system: str, user: str, label: str) -> str:
    brevity = "OUTPUT CONSTRAINT: Complete but concise. Under 1500 words. No padding.\n\n"
    prompt  = f"{brevity}{system}\n\n---\n\n{user}"
    return call_api(prompt, label=label, retries=2)


# ── Stage 1 batch helpers ─────────────────────────────────────────────────────

def batch_label(ep_stems: list) -> str:
    return f"{ep_stems[0]}-{ep_stems[-1]}"


def show_level_filename(file_type: str, label: str) -> str:
    return f"{file_type}_{label}.md"


def find_previous_batch_file(file_type: str, current_label: str, out_dir: Path) -> str:
    existing      = sorted(out_dir.glob(f"{file_type}_*.md"))
    current_fname = show_level_filename(file_type, current_label)
    prior         = [f for f in existing if f.name != current_fname]
    return prior[-1].read_text(encoding="utf-8") if prior else ""


# ── Stage 1 generation ────────────────────────────────────────────────────────

def generate_show_level_batch(file_type, batch_stems, batch_scripts, all_stems, out_dir, show_name):
    label    = batch_label(batch_stems)
    out_path = out_dir / show_level_filename(file_type, label)

    if out_path.exists() and out_path.stat().st_size > 100:
        _log(f"  ->  {out_path.name} — already exists, skipping")
        return

    prev_content = find_previous_batch_file(file_type, label, out_dir)
    prev_block   = ""
    if prev_content:
        prev_block = (
            f"\n\n## INHERITED FROM PREVIOUS BATCHES\n"
            f"The following is already established. Extend and update it — do not contradict it.\n\n"
            f"{trim(prev_content, max_chars=20_000)}\n\n---\n\n"
            f"## NEW EPISODES TO PROCESS (add to the above)\n"
        )

    system   = load_prompt(file_type)
    combined = "\n\n---\n\n".join(
        f"### Script: {s}\n\n{trim(batch_scripts[s], max_chars=3_000)}"
        for s in batch_stems
    )

    total_eps     = len(all_stems)
    batch_num     = all_stems.index(batch_stems[0]) // BATCH_SIZE + 1
    total_batches = (total_eps + BATCH_SIZE - 1) // BATCH_SIZE

    user = (
        f"Show: {show_name}\n"
        f"Processing: Episodes {batch_stems[0]} to {batch_stems[-1]} "
        f"(batch {batch_num}/{total_batches} of {total_eps} total episodes)\n"
        f"{prev_block}{combined}"
    )

    _log(f"  Generating: {out_path.name} ...")
    result = call_claude(system, user, label=f"{file_type} {label}")
    if not result:
        result = f"# {file_type.replace('_', ' ').title()} — {label}\n\n[Generation failed — rerun]"
    out_path.write_text(result, encoding="utf-8")
    _log(f"  ok  {out_path.name}  ({len(result):,} chars)")


def generate_all_show_level_batches(scripts, out_dir, show_name, file_types=None):
    if file_types is None:
        file_types = ["show_tone_bible", "character_canvas", "location_reference"]
    all_stems = list(scripts.keys())
    total     = len(all_stems)
    batches   = [all_stems[i:i + BATCH_SIZE] for i in range(0, total, BATCH_SIZE)]

    _log(f"\n  {total} episodes -> {len(batches)} batches of {BATCH_SIZE}")
    _log(f"  {len(file_types)} file type(s) x {len(batches)} batches = {len(file_types) * len(batches)} total calls\n")

    for bi, batch_stems in enumerate(batches, 1):
        batch_scripts = {s: scripts[s] for s in batch_stems}
        _log(f"  -- Batch {bi}/{len(batches)}: {batch_label(batch_stems)} --")
        for i, ft in enumerate(file_types):
            generate_show_level_batch(ft, batch_stems, batch_scripts, all_stems, out_dir, show_name)
            if i < len(file_types) - 1:
                time.sleep(INTER_CALL_PAUSE)
        if bi < len(batches):
            time.sleep(INTER_CALL_PAUSE)

    _log("\n  ok  Show-level files complete.")


def _episode_detail_task(script_stem, script_content, out_dir, show_name, worker_idx, log_queue, job_id):
    _set_job_context(job_id, log_queue)
    out_path = out_dir / f"{script_stem}_episode_detail.md"
    if out_path.exists() and out_path.stat().st_size > 100:
        _log(f"  ->  [{worker_idx}] {script_stem} — skip (exists)")
        return script_stem, True, out_path.stat().st_size

    system = load_prompt("episode_detail")
    user   = f"Show: {show_name}\nEpisode: {script_stem}\n\n---\n\n{trim(script_content)}"
    _log(f"  ... [{worker_idx}] {script_stem} ...")
    result = call_claude(system, user, label=script_stem)

    if not result:
        result = f"# {script_stem} Episode Detail\n\n[Generation failed — rerun]"
        _log(f"  x  [{worker_idx}] {script_stem} — failed")
        out_path.write_text(result, encoding="utf-8")
        return script_stem, False, 0

    out_path.write_text(result, encoding="utf-8")
    _log(f"  ok  [{worker_idx}] {script_stem}  ({len(result):,} chars)")
    return script_stem, True, len(result)


def generate_all_episode_details(scripts, out_dir, show_name, workers, log_queue, job_id):
    items     = list(scripts.items())
    total     = len(items)
    skipped   = sum(
        1 for s, _ in items
        if (out_dir / f"{s}_episode_detail.md").exists()
        and (out_dir / f"{s}_episode_detail.md").stat().st_size > 100
    )
    _log(f"\n  {total} episodes  |  {skipped} already done  |  {total - skipped} to generate")
    _log(f"  Workers: {workers} parallel  |  Stagger: {STAGGER_SEC}s\n")

    if total - skipped == 0:
        _log("  ok  All episode detail files already exist.")
        return

    done = failed = 0
    start = time.time()

    if job_id in jobs:
        jobs[job_id]["progress"]["total"]      = total
        jobs[job_id]["progress"]["completed"]  = skipped
        jobs[job_id]["progress"]["started_at"] = start
        jobs[job_id]["progress"]["stage"]      = "Episode details"

    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {}
        for idx, (stem, content) in enumerate(items):
            if idx > 0 and idx % workers == 0:
                time.sleep(STAGGER_SEC)
            f = executor.submit(_episode_detail_task, stem, content, out_dir, show_name, (idx % workers) + 1, log_queue, job_id)
            futures[f] = stem

        for future in as_completed(futures):
            _, success, _ = future.result()
            if success: done += 1
            else:       failed += 1
            elapsed = time.time() - start
            eta = (elapsed / max(done + failed, 1)) * (total - done - failed)
            _log(f"  Progress: {done + failed}/{total}  ok:{done}  x:{failed}  ETA ~{eta / 60:.0f}min")
            if job_id in jobs:
                jobs[job_id]["progress"]["completed"] = skipped + done + failed

    _log(f"\n  ok  {done}/{total} done  ({failed} failed)")


def _is_show_level_file(filename: str) -> bool:
    if filename in SHOW_LEVEL_FILES:
        return True
    return any(filename.startswith(stem + "_") for stem in SHOW_LEVEL_STEMS)


def _run_stage1_core(job_id: str, job_dir: Path, mode: str, workers: int, episode: str):
    """Core Stage 1 logic — callable from standalone route or pipeline."""
    scripts_dir = job_dir / "episodic scripts"
    raw_out     = job_dir / "_stage1_out"
    raw_out.mkdir(parents=True, exist_ok=True)
    show_name   = job_dir.name

    _log(f"\n{'=' * 60}")
    _log("  STAGE 1 — Reference File Generation")
    _log(f"  Show: {show_name}")
    _log(f"{'=' * 60}\n")

    verify_auth()
    auto_convert(scripts_dir)

    filenames = [episode] if mode == "episode" else None
    scripts   = read_scripts(scripts_dir, filenames)

    if mode in ("full", "all-episodes"):
        _log("\n  -- Episode Detail Files ----------------------")
        generate_all_episode_details(scripts, raw_out, show_name, workers, jobs[job_id]["queue"], job_id)

    if mode in ("full", "show-level"):
        _log("\n  -- Show-Level Files (batched) ----------------")
        generate_all_show_level_batches(scripts, raw_out, show_name)

    if mode == "episode":
        stem = Path(episode).stem
        _log(f"\n  -- Single episode: {stem} --")
        _episode_detail_task(stem, scripts[stem], raw_out, show_name, 1, jobs[job_id]["queue"], job_id)

    show_level_dir = job_dir / "show level files"
    ep_details_dir = job_dir / "episode details"
    show_level_dir.mkdir(exist_ok=True)
    ep_details_dir.mkdir(exist_ok=True)

    moved_show = moved_ep = 0
    for f in raw_out.glob("*.md"):
        dest = show_level_dir if _is_show_level_file(f.name) else ep_details_dir
        shutil.copy2(f, dest / f.name)
        if dest is show_level_dir: moved_show += 1
        else:                      moved_ep   += 1
    shutil.rmtree(raw_out, ignore_errors=True)

    tok = jobs[job_id]["tokens"]
    s1_cost = (tok["input"] * LLM_INPUT_COST_PER_MTOK + tok["output"] * LLM_OUTPUT_COST_PER_MTOK) / 1_000_000
    jobs[job_id]["file_tokens"].append({
        "name": f"Show-level files ({moved_show})",
        "input": tok["input"], "output": tok["output"], "calls": tok["calls"],
        "cost": round(s1_cost, 4),
    })
    _log(f"\n{'=' * 60}")
    _log("  STAGE 1 COMPLETE")
    _log(f"    -> show level files/  ({moved_show} file(s))")
    _log(f"    -> episode details/   ({moved_ep} file(s))")
    _log(f"    -> Tokens used: {tok['input']:,} input / {tok['output']:,} output / {tok['calls']} API calls")
    _log(f"    -> Cost: ${s1_cost:.4f}")
    _log(f"{'=' * 60}\n")


def run_stage1_pipeline(job_id: str, job_dir: Path, mode: str, workers: int, episode: str):
    """Standalone entry point (called from API route)."""
    log_queue = jobs[job_id]["queue"]
    _set_job_context(job_id, log_queue)
    try:
        _run_stage1_core(job_id, job_dir, mode, workers, episode)
        jobs[job_id]["status"] = "done"
    except Exception as e:
        tb = traceback.format_exc()
        _log(f"\n[ERROR] {e}\n{tb}")
        jobs[job_id]["status"] = "failed"
    finally:
        log_queue.put(None)


# ── Stage 2 helpers ───────────────────────────────────────────────────────────

def parse_shot_rows(text: str) -> list:
    """Extract shot rows from API response. Tries JSON first, falls back to markdown table."""
    # Try ```json ... ``` block
    m = re.search(r"```json\s*([\s\S]*?)\s*```", text)
    if m:
        try:
            data = json.loads(m.group(1))
            if isinstance(data, list):
                return data
        except json.JSONDecodeError:
            pass

    # Try bare JSON array
    m = re.search(r"\[\s*\{[\s\S]*?\}\s*\]", text)
    if m:
        try:
            data = json.loads(m.group(0))
            if isinstance(data, list):
                return data
        except json.JSONDecodeError:
            pass

    # Fallback: markdown table — rows starting with | digit
    rows = []
    for line in text.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.split("|")[1:-1]]
        if not cells or not cells[0].isdigit():
            continue
        if len(cells) >= 6:
            rows.append({
                "shot_number":      int(cells[0]),
                "line":             cells[1],
                "shot_size":        cells[2],
                "shot_description": cells[3],
                "shot_detail":      cells[4],
                "reference":        cells[5] if len(cells) > 5 else "",
            })
    return rows


def write_excel(rows: list, output_path: Path, episode_name: str):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = episode_name[:30]

    headers = ["#", "Line", "Shot Size", "Shot Description", "Shot Detail", "Reference", "Preview_1", "Preview_2"]

    # Header row
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for row_data in rows:
        row_vals = [
            row_data.get("shot_number", ""),
            row_data.get("line", ""),
            row_data.get("shot_size", ""),
            row_data.get("shot_description", ""),
            row_data.get("shot_detail", ""),
            row_data.get("reference", ""),
            "",
            "",
        ]
        ws.append(row_vals)
        r = ws.max_row
        for col in range(1, 9):
            ws.cell(r, col).alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths
    col_widths = [5, 40, 14, 55, 80, 15, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    wb.save(str(output_path))


MAKE_EXCEL_MSG = (
    "MAKE EXCEL\n\n"
    "⚠️⚠️⚠️ ABSOLUTE ZERO-SKIP RULE — THIS IS THE #1 PRIORITY:\n"
    "The `line` field is used directly for audio voiceover and subtitle alignment.\n"
    "If ANY word from the script is missing, the voiceover will skip audio and subtitles will desync.\n\n"
    "RULES:\n"
    "1. The `line` field MUST be COPY-PASTED verbatim from the script — every word, every punctuation mark.\n"
    "2. When you concatenate line from shot 1 + shot 2 + shot 3 + ... it MUST exactly equal the full script.\n"
    "3. Do NOT paraphrase, summarize, shorten, or rephrase ANY line.\n"
    "4. Do NOT skip ANY sentence or clause — even transitions, scene descriptions, or short lines.\n"
    "5. Do NOT merge two script sentences into one shot's line field.\n"
    "6. If the script says 'He turned around.' that exact text must appear in a line field somewhere.\n"
    "7. Check: does your output cover the ENTIRE script from first word to last? If not, you missed lines.\n\n"
    "⚠️⚠️⚠️ LINE-DESCRIPTION COUPLING RULE — MANDATORY:\n"
    "Each shot row must be written as a COMPLETE UNIT. For every row you output, the `line` and `shot_description` "
    "MUST describe the SAME visual moment. Never write a batch of lines first and fill descriptions later.\n"
    "Before outputting each row, verify: does the shot_description match the script text in the line field? "
    "If the line says 'Yvonne slapped him across the face' but the description says 'Penelope shoves Yvonne' — "
    "that row is BROKEN. The description drifted to a different scene.\n\n"
    "⚠️⚠️⚠️ NO GHOST DUPLICATES — MANDATORY:\n"
    "Every script sentence must appear in EXACTLY ONE shot row. If you realize a block of shots has wrong descriptions, "
    "you must FIX those existing rows (reuse their shot_number) — do NOT append a corrected copy at the end. "
    "Appending without removing the original creates ghost duplicates where the same line appears twice "
    "at different positions in the breakdown.\n\n"
    "Output each batch of 25 shots as a JSON array wrapped in ```json and ``` markers. "
    "Each shot object must have exactly these keys: shot_number (integer), line (string — exact verbatim script text), "
    "shot_size (string), shot_description (string), shot_detail (string), reference (string). "
    "End each batch with: BATCH COMPLETE — shots [X]–[Y] done. "
    "After the final batch: ALL SHOTS COMPLETE — [X] total shots delivered."
)


BEAT_EXTRACTION_SYSTEM = (
    "You are a script parser. Your job is to split a script into numbered beats.\n\n"
    "A beat is a single sentence or distinct narrative unit from the script. "
    "Every sentence, every line of dialogue (including its dialogue tag), "
    "every action description, and every transition is its own beat.\n\n"
    "RULES:\n"
    "1. Copy each beat VERBATIM from the script — do not paraphrase, summarize, or reword.\n"
    "2. Preserve exact punctuation, capitalization, and spelling.\n"
    "3. Do not merge multiple sentences into one beat.\n"
    "4. Do not split a single sentence across multiple beats UNLESS it contains "
    "distinct clauses with different subjects or actions (e.g. 'He ran forward, and she ducked behind the wall' = 2 beats).\n"
    "5. Include everything — scene descriptions, transitions, short lines, internal monologue.\n"
    "6. Number each beat sequentially starting from 1.\n\n"
    "OUTPUT FORMAT:\n"
    "```beats\n"
    "1. [exact sentence from script]\n"
    "2. [exact sentence from script]\n"
    "...\n"
    "```\n\n"
    "End with: BEAT COUNT: [N] total beats.\n"
    "Do not add commentary or analysis. Only output the numbered beat list."
)


def _extract_beats(script_text: str, episode_name: str) -> str:
    """Pass 1: Extract numbered beats from the script as an explicit checklist."""
    _log(f"  ── Pass 1: Beat extraction for '{episode_name}' ──")
    messages = [{"role": "user", "content": f"Split this script into numbered beats:\n\n{script_text}"}]
    response = call_api_chat(BEAT_EXTRACTION_SYSTEM, messages, label=f"beats:{episode_name}")
    if not response:
        _log(f"  !! Beat extraction failed — proceeding without beat list")
        return ""
    beat_count = 0
    for line in response.splitlines():
        line = line.strip()
        if line and line[0].isdigit() and ". " in line:
            beat_count += 1
    _log(f"  ok  {beat_count} beats extracted from script")
    return response


def _build_stage2_user_msg(ref_files: dict, script_text: str, episode_name: str) -> str:
    ref_parts = [
        f"## {k.replace('_', ' ').title()}\n\n{v.strip()}"
        for k, v in ref_files.items() if v and v.strip()
    ]
    ref_context = "\n\n---\n\n".join(ref_parts)
    if ref_context:
        return (f"Reference Files:\n\n{ref_context}\n\n---\n\n"
                f'Script for Episode "{episode_name}":\n\n{script_text}')
    return f'Script for Episode "{episode_name}":\n\n{script_text}'


def _match_episode_detail(episode_name: str, detail_files: list) -> str:
    """detail_files: [(filename, content), ...]. Match by stripping _episode_detail suffix."""
    ep_lo = episode_name.lower().strip()
    for filename, content in detail_files:
        stem = re.sub(r'_episode_detail(\.md|\.txt)?$', '', filename, flags=re.IGNORECASE).lower().strip()
        if stem == ep_lo:
            return content
    # Fallback: stem prefix overlap
    for filename, content in detail_files:
        stem = re.sub(r'_episode_detail(\.md|\.txt)?$', '', filename, flags=re.IGNORECASE).lower().strip()
        if stem and (stem in ep_lo or ep_lo in stem):
            return content
    return ""


CHUNK_TARGET_WORDS = 500


def _split_script_into_chunks(script_text: str, target_words: int = CHUNK_TARGET_WORDS) -> list:
    paragraphs = re.split(r'\n\s*\n', script_text.strip())
    paragraphs = [p.strip() for p in paragraphs if p.strip()]
    if not paragraphs:
        return [{"text": script_text, "word_count": len(script_text.split())}]

    chunks = []
    current_paras = []
    current_words = 0

    for para in paragraphs:
        para_words = len(para.split())
        if current_words + para_words > target_words * 1.3 and current_words >= target_words * 0.5:
            chunk_text = "\n\n".join(current_paras)
            chunks.append({"text": chunk_text, "word_count": len(chunk_text.split())})
            current_paras = [para]
            current_words = para_words
        else:
            current_paras.append(para)
            current_words += para_words

    if current_paras:
        chunk_text = "\n\n".join(current_paras)
        chunks.append({"text": chunk_text, "word_count": len(chunk_text.split())})

    return chunks


def _process_one_episode(job_dir: Path, script_text: str, episode_name: str, ref_files: dict, _chunk_mode: bool = False) -> str:
    """Run full Stage 2 pipeline for one episode. Returns Excel filename.
    When _chunk_mode=True, returns the raw shot rows list instead of writing Excel."""
    jid = getattr(_job_local, "job_id", None)

    def _update_shot_counter(total: int):
        if jid and jid in jobs:
            jobs[jid]["progress"]["shots_done"] = total
            jobs[jid]["progress"]["ep_shots"]   = total

    full_script_text = script_text  # preserve for global post-processing

    def _norm(t):
        return re.sub(r'[^\w\s]', '', t.lower()).split()

    def _norm_str(t):
        return " ".join(_norm(t))

    # ── Chunked generation: split script, process each chunk, stitch ─────────
    word_count = len(script_text.split())
    chunks = _split_script_into_chunks(script_text) if not _chunk_mode and word_count > CHUNK_TARGET_WORDS * 1.5 else []

    if chunks and len(chunks) > 1:
        min_shots_global = max(10, int(word_count / 100 * 7))
        max_shots_global = int(word_count / 100 * 9)
        _log(f"  Script: {word_count:,} words  |  {len(chunks)} chunks  |  Target: {min_shots_global}–{max_shots_global} shots")

        all_chunk_rows = []
        for ci, chunk in enumerate(chunks):
            _log(f"\n  ━━━ Chunk {ci+1}/{len(chunks)} ({chunk['word_count']} words) ━━━")
            if jid and jid in jobs:
                jobs[jid]["progress"]["stage"] = f"Chunk {ci+1}/{len(chunks)}"
            chunk_rows = _process_one_episode(job_dir, chunk["text"], f"{episode_name}_chunk{ci+1}", ref_files, _chunk_mode=True)
            all_chunk_rows.extend(chunk_rows)
            _log(f"  Chunk {ci+1}: {len(chunk_rows)} shots  |  Running total: {len(all_chunk_rows)}")

        all_rows = all_chunk_rows
        script_text = full_script_text  # restore for global post-processing below

        # jump to global post-processing (reorder, dedup, validation, Excel write)
        # — this code path falls through to the reorder section below
    else:
        # Single chunk or chunk_mode — run the standard pipeline
        min_shots  = max(10, int(word_count / 100 * 7))
        max_shots  = int(word_count / 100 * 9)
        min_shots_global = min_shots
        max_shots_global = max_shots
        _log(f"  Script: {word_count:,} words  |  Required shots: {min_shots}–{max_shots}")

        # ── Pass 1: Beat extraction ──────────────────────────────────────────
        beat_list = _extract_beats(script_text, episode_name)

    # ── Inject mandatory target + beat checklist into the user message ────────
    beat_checklist_block = ""
    if beat_list:
        beat_checklist_block = (
            f"\n\n---\n"
            f"## SCRIPT BEAT CHECKLIST — MANDATORY COVERAGE\n"
            f"The following is the complete list of beats extracted from this script. "
            f"Every beat MUST appear as a verbatim `line` field in at least one shot. "
            f"Use this as your checklist — work through it sequentially and do not skip any beat.\n\n"
            f"{beat_list}\n"
            f"---"
        )

    shot_target_block = (
        f"\n\n---\n"
        f"## MANDATORY SHOT TARGET FOR THIS EPISODE\n"
        f"Script word count: {word_count:,} words\n"
        f"Required shot range: **{min_shots}–{max_shots} shots** (7–9 shots per 100 words)\n"
        f"You MUST produce AT LEAST {min_shots} shots for this episode.\n\n"
        f"ENFORCEMENT RULES — apply before writing a single shot:\n"
        f"1. Every line over 30 words MUST be split — no exceptions.\n"
        f"2. Every new subject, action, emotion, or environment = new shot.\n"
        f"3. Two characters doing different things in the same sentence = two shots.\n"
        f"4. An internal reaction followed by an external reaction = two shots.\n"
        f"5. If your Step 1 total is below {min_shots}, you have not split aggressively enough — go back.\n"
        f"6. Cross-check every beat in the SCRIPT BEAT CHECKLIST above — if a beat has no shot, add one.\n"
        f"---"
    )
    user_msg = _build_stage2_user_msg(ref_files, script_text, episode_name) + beat_checklist_block + shot_target_block
    messages = [{"role": "user", "content": user_msg}]

    # ── Step 1–5 breakdown (Pass 2: shot generation with beat checklist) ─────
    _log(f"  ── Pass 2: Shot breakdown for '{episode_name}' (with beat checklist) ──")
    response = call_api_chat(STAGE2_SYSTEM, messages, label=f"breakdown:{episode_name}")
    if not response:
        raise RuntimeError(f"No breakdown response for {episode_name}")

    _log(f"  --- Breakdown summary ({episode_name}) ---")
    for line in response.splitlines()[:8]:
        _log(f"  {line}")
    _log("  ---")

    messages.append({"role": "assistant", "content": response})
    if "BREAKDOWN COMPLETE" not in response:
        _log("  !! Breakdown message not detected — proceeding anyway")

    # ── MAKE EXCEL — coverage + count requirements baked in, no extra call ────
    beat_reminder = ""
    if beat_list:
        beat_reminder = (
            "\n\nREMINDER — BEAT CHECKLIST: The numbered beat list from Pass 1 is your coverage checklist. "
            "As you output each batch, mentally check off each beat. After your final batch, every beat "
            "must have at least one shot. If you reach ALL SHOTS COMPLETE and any beat is uncovered, "
            "you are NOT done.\n"
        )
    make_excel_with_target = (
        MAKE_EXCEL_MSG +
        f"\n\nMANDATORY: This episode requires {min_shots}–{max_shots} shots ({word_count:,}-word script). "
        f"Do NOT declare ALL SHOTS COMPLETE until you have output at least {min_shots} shots. "
        f"Every script line must appear verbatim in the `line` field — nothing skipped.\n\n"
        f"HERE IS THE FULL SCRIPT FOR REFERENCE — use this to verify you are not missing ANY text:\n"
        f"--- SCRIPT START ---\n{script_text}\n--- SCRIPT END ---\n"
        f"Go through this script paragraph by paragraph. Every word between SCRIPT START and SCRIPT END "
        f"must appear in a `line` field in your output."
        + beat_reminder
    )
    messages.append({"role": "user", "content": make_excel_with_target})

    # ── Collect batches ───────────────────────────────────────────────────────
    all_rows  = []
    batch_num = 1
    max_batches = 40
    while batch_num <= max_batches:
        _log(f"  Fetching batch {batch_num}...")
        response = call_api_chat(STAGE2_SYSTEM, messages, label=f"{episode_name}:batch-{batch_num}")
        if not response:
            _log(f"  x  Empty response on batch {batch_num} — stopping")
            break
        rows = parse_shot_rows(response)
        all_rows.extend(rows)
        _update_shot_counter(len(all_rows))
        _log(f"  ✓ Batch {batch_num}: +{len(rows)} shots  |  TOTAL: {len(all_rows)}/{min_shots} required")
        messages.append({"role": "assistant", "content": response})

        if "ALL SHOTS COMPLETE" in response:
            if len(all_rows) >= min_shots:
                _log(f"  ok  All batches received for '{episode_name}'")
                break
            # Model stopped early — reject and force continuation
            missing = min_shots - len(all_rows)
            _log(f"  !! Model stopped at {len(all_rows)} shots — {missing} more required. Forcing continuation...")
            messages.append({"role": "user", "content": (
                f"⛔ REJECTED — ALL SHOTS COMPLETE is PREMATURE.\n"
                f"You have delivered {len(all_rows)} shots but this {word_count:,}-word script requires "
                f"a minimum of {min_shots} shots. You are missing at least {missing} more shots.\n\n"
                f"You did not split the script aggressively enough. Every sentence or clause with a new "
                f"subject, action, emotion, or visual detail is a separate shot. Every line over 30 words "
                f"must be split — no exceptions.\n\n"
                f"Continue from shot {len(all_rows) + 1}. Do NOT repeat shots already delivered. "
                f"Output the next batch of 25 shots now."
            )})
            batch_num += 1
            continue

        messages.append({"role": "user", "content": "NEXT"})
        batch_num += 1

    if not all_rows:
        raise RuntimeError(f"No shot rows parsed for {episode_name}")

    # ── Coverage verification (word-level + sentence-level) ────────────────
    def _measure_coverage(rows):
        sw = _norm(script_text)
        rw = []
        for r in rows:
            rw.extend(_norm(r.get("line", "")))
        if not sw or not rw:
            return 0.0, []
        from difflib import SequenceMatcher
        sm = SequenceMatcher(None, sw, rw, autojunk=False)
        matched = sum(b.size for b in sm.get_matching_blocks())
        cov = matched / len(sw) * 100
        covered = set()
        ptr = 0
        for r in rows:
            for w in _norm(r.get("line", "")):
                while ptr < len(sw) and sw[ptr] != w:
                    ptr += 1
                if ptr < len(sw):
                    covered.add(ptr)
                    ptr += 1
        gaps = []
        gs = None
        for i in range(len(sw)):
            if i not in covered:
                if gs is None: gs = i
            else:
                if gs is not None:
                    gaps.append((gs, i))
                    gs = None
        if gs is not None:
            gaps.append((gs, len(sw)))
        return cov, [(s, e) for s, e in gaps if e - s >= 3]

    def _find_rephrased(rows, max_sents=80):
        all_lines = " ".join(r.get("line", "") for r in rows)
        all_lines_norm = _norm_str(all_lines)
        sents = re.split(r'(?<=[.!?…])\s+|\n+', script_text.strip())
        sents = [s.strip() for s in sents if len(s.strip().split()) >= 3]
        line_norms = [_norm_str(r.get("line", "")) for r in rows]
        rephrased = []
        for sent in sents[:max_sents]:
            sn = _norm_str(sent)
            if not sn or len(sn) < 10:
                continue
            if sn in all_lines_norm:
                continue
            sn_words = set(sn.split())
            best_overlap = 0
            best_line = ""
            best_idx = -1
            for ri, ln in enumerate(line_norms):
                if not ln:
                    continue
                ln_words = set(ln.split())
                overlap = len(sn_words & ln_words)
                score = overlap / max(len(sn_words), 1)
                if score > best_overlap:
                    best_overlap = score
                    best_line = rows[ri].get("line", "")
                    best_idx = ri
            if 0.5 < best_overlap < 0.95:
                rephrased.append({
                    "original": sent,
                    "found_as": best_line,
                    "similarity": best_overlap,
                })
        return rephrased[:10]

    def _find_missing_sentences(rows):
        sents = re.split(r'(?<=[.!?…"\'”’])\s+|\n+', script_text.strip())
        sents = [s.strip() for s in sents if len(s.strip().split()) >= 3]
        all_lines_norm = " ".join(_norm_str(r.get("line", "")) for r in rows)
        all_line_words = set(all_lines_norm.split())
        missing = []
        for sent in sents:
            sn = _norm_str(sent)
            if not sn or len(sn.split()) < 3:
                continue
            if sn in all_lines_norm:
                continue
            sn_words = sn.split()
            found_count = sum(1 for w in sn_words if w in all_line_words)
            if found_count / max(len(sn_words), 1) < 0.7:
                missing.append(sent)
        return missing[:20]

    def _find_truncated_lines(rows):
        sents = re.split(r'(?<=[.!?…"\'”’])\s+|\n+', script_text.strip())
        sents = [s.strip() for s in sents if len(s.strip().split()) >= 3]
        sent_norms = [_norm_str(s) for s in sents]
        all_lines_norm = " ".join(_norm_str(r.get("line", "")) for r in rows)
        truncated = []
        for row in rows:
            line = row.get("line", "").strip()
            if not line or len(line.split()) < 3:
                continue
            line_norm = _norm_str(line)
            if not line_norm:
                continue
            for si, sent_norm in enumerate(sent_norms):
                if line_norm in sent_norm and line_norm != sent_norm:
                    missing_part = sent_norm.replace(line_norm, "", 1).strip()
                    if len(missing_part.split()) >= 3 and missing_part not in all_lines_norm:
                        truncated.append({
                            "shot": row.get("shot_number", "?"),
                            "line": line,
                            "full_sentence": sents[si],
                        })
                        break
        return truncated[:10]

    script_words = _norm(script_text)
    coverage, big_gaps = _measure_coverage(all_rows)
    _log(f"  Coverage: {coverage:.1f}% of script words in line column")

    rephrased = _find_rephrased(all_rows)
    if rephrased:
        _log(f"  !! {len(rephrased)} line(s) appear REPHRASED (not verbatim)")
        for rp in rephrased[:5]:
            _log(f"    Script:  \"{rp['original'][:100]}\"")
            _log(f"    Found:   \"{rp['found_as'][:100]}\"  ({rp['similarity']:.0%} similar)")

    prev_coverage = -1
    for fill_round in range(3):
        if coverage == prev_coverage and fill_round >= 1:
            _log(f"  !! Coverage unchanged at {coverage:.1f}% — stopping fill rounds")
            break
        prev_coverage = coverage

        issues = []
        if big_gaps:
            _log(f"  !! Coverage {coverage:.1f}% — fill round {fill_round + 1}, {len(big_gaps)} gap(s)...")
            gap_texts = []
            for gi, (gs, ge) in enumerate(big_gaps[:10]):
                original_chunk = " ".join(script_words[gs:ge])
                display = f"\"{original_chunk[:150]}\"" if len(original_chunk) <= 150 else f"\"{original_chunk[:150]}...\""
                gap_texts.append(f"Gap {gi+1} (~{ge-gs} words): {display}")
                _log(f"    {gap_texts[-1]}")
            issues.append(
                "MISSING TEXT — the following script sections are NOT in any shot's line field:\n" +
                "\n".join(gap_texts)
            )

        if rephrased:
            _log(f"  !! {len(rephrased)} rephrased line(s) — fix round {fill_round + 1}")
            rephrase_items = []
            for rp in rephrased[:5]:
                rephrase_items.append(
                    f"WRONG: \"{rp['found_as'][:200]}\"\n"
                    f"CORRECT (from script): \"{rp['original'][:200]}\""
                )
            issues.append(
                "REPHRASED LINES — the following lines were summarized or reworded instead of copied verbatim:\n" +
                "\n\n".join(rephrase_items)
            )

        missing_sents = _find_missing_sentences(all_rows)
        if missing_sents:
            _log(f"  !! {len(missing_sents)} script sentence(s) completely missing — fill round {fill_round + 1}")
            for ms in missing_sents[:5]:
                _log(f"    - \"{ms[:120]}\"")
            issues.append(
                "COMPLETELY MISSING SENTENCES — the following script sentences are not in any shot's Line field. "
                "You MUST create new shots for each one with the EXACT text from the script:\n" +
                "\n".join(f'- "{s}"' for s in missing_sents[:10])
            )

        truncated = _find_truncated_lines(all_rows)
        if truncated:
            _log(f"  !! {len(truncated)} truncated line(s) found — fill round {fill_round + 1}")
            trunc_items = []
            for t in truncated[:5]:
                trunc_items.append(
                    f"Shot {t['shot']}: has \"{t['line'][:150]}\"\n"
                    f"Full sentence from script: \"{t['full_sentence'][:200]}\""
                )
            issues.append(
                "TRUNCATED LINES — these shots have incomplete script sentences. "
                "The Line field must contain the FULL sentence, not a fragment:\n" +
                "\n\n".join(trunc_items)
            )

        if not issues:
            break

        fill_prompt = (
            f"⛔ SCRIPT COVERAGE AUDIT FAILED (round {fill_round + 1})\n\n"
            + "\n\n---\n\n".join(issues) +
            f"\n\n---\n\nRULES:\n"
            f"1. The `line` field MUST be COPY-PASTED verbatim from the script — every word, every punctuation mark.\n"
            f"2. Do NOT paraphrase, summarize, shorten, or reword ANY text. Use the EXACT original words.\n"
            f"3. For missing text: produce new shots with the exact script text.\n"
            f"4. For rephrased lines: output corrected shots with the EXACT original script text replacing the wrong version.\n"
            f"5. For truncated lines: output the corrected shot with the FULL sentence — never drop a clause.\n"
            f"6. Output as JSON array in ```json blocks. Use shot numbers starting from {len(all_rows) + 1} for new shots.\n"
            f"7. For corrections, reuse the original shot number and include the corrected `line`.\n"
            f"8. CRITICAL — LINE-DESCRIPTION COUPLING: When correcting or adding a shot, the `shot_description` MUST match "
            f"the `line` for that same row. Never output a line from one scene with a description from a different scene.\n"
            f"9. CRITICAL — REPLACE, DO NOT APPEND: If a block of shots had wrong descriptions, output them with their "
            f"ORIGINAL shot_number so they replace the broken rows. Do NOT create new shot numbers for content that "
            f"already has a row — that creates ghost duplicates."
        )
        messages.append({"role": "user", "content": fill_prompt})
        fill_resp = call_api_chat(STAGE2_SYSTEM, messages, label=f"{episode_name}:fill-{fill_round+1}")
        if fill_resp:
            messages.append({"role": "assistant", "content": fill_resp})
            fill_rows = parse_shot_rows(fill_resp)
            if fill_rows:
                existing_nums = {r.get("shot_number"): idx for idx, r in enumerate(all_rows)}
                new_rows = []
                for fr in fill_rows:
                    fn = fr.get("shot_number")
                    if fn in existing_nums:
                        idx = existing_nums[fn]
                        for k in ("line", "shot_description", "shot_detail", "shot_size", "reference"):
                            if fr.get(k):
                                all_rows[idx][k] = fr[k]
                        _log(f"  ok  Corrected shot {fn} (line + description replaced together)")
                    else:
                        new_rows.append(fr)
                if new_rows:
                    all_rows.extend(new_rows)
                    _log(f"  ok  +{len(new_rows)} new shots — total now {len(all_rows)}")
                _update_shot_counter(len(all_rows))

        coverage, big_gaps = _measure_coverage(all_rows)
        rephrased = _find_rephrased(all_rows)
        _log(f"  Coverage after round {fill_round + 1}: {coverage:.1f}%  |  Rephrased: {len(rephrased)}")

        if _chunk_mode:
            return all_rows

    # ── Global post-processing (both chunked and single paths merge here) ─────

    # Re-order rows by scanning through script and greedily matching each row
    # Handles repeated dialogue correctly: first occurrence matches first row,
    # second occurrence matches the next unassigned row with that text
    from collections import defaultdict
    _MATCH_PFX = 8
    _row_norms = [_norm(r.get("line", "")) for r in all_rows]
    _pfx_by_len = defaultdict(lambda: defaultdict(list))
    for _ri, _rw in enumerate(_row_norms):
        _kl = min(_MATCH_PFX, len(_rw))
        if _kl >= 2:
            _pfx_by_len[_kl][tuple(_rw[:_kl])].append(_ri)

    _assigned = set()
    _ordered = []
    _pos = 0
    while _pos < len(script_words):
        _matched = False
        for _klen in range(min(_MATCH_PFX, len(script_words) - _pos), 1, -1):
            if _klen not in _pfx_by_len:
                continue
            _key = tuple(script_words[_pos:_pos + _klen])
            if _key in _pfx_by_len[_klen]:
                for _ri in _pfx_by_len[_klen][_key]:
                    if _ri not in _assigned:
                        _assigned.add(_ri)
                        _ordered.append(all_rows[_ri])
                        _pos += max(len(_row_norms[_ri]), 1)
                        _matched = True
                        break
            if _matched:
                break
        if not _matched:
            _pos += 1

    for _ri in range(len(all_rows)):
        if _ri not in _assigned:
            _ordered.append(all_rows[_ri])
    all_rows = _ordered

    # ── Duplicate detection and removal ─────────────────────────────────────
    before_dedup = len(all_rows)
    seen_lines = {}
    deduped = []
    for row in all_rows:
        norm_line = " ".join(_norm(row.get("line", "")))
        if not norm_line:
            continue
        if norm_line in seen_lines:
            _log(f"  !!  Removing exact duplicate line (shot {row.get('shot_number', '?')}): \"{row.get('line', '')[:80]}...\"")
            continue
        seen_lines[norm_line] = True
        deduped.append(row)

    # Partial duplicate check: if one line is a substring of another, remove the redundant
    remove_idxs = set()
    for i, row_a in enumerate(deduped):
        if i in remove_idxs:
            continue
        norm_a = " ".join(_norm(row_a.get("line", "")))
        if not norm_a:
            continue
        for j, row_b in enumerate(deduped):
            if i == j or j in remove_idxs:
                continue
            norm_b = " ".join(_norm(row_b.get("line", "")))
            if norm_b and norm_a != norm_b and norm_a in norm_b:
                remainder = norm_b.replace(norm_a, "", 1).strip()
                if len(remainder.split()) < 3:
                    _log(f"  !!  Partial duplicate: shot {row_b.get('shot_number', '?')} subsumes shot {row_a.get('shot_number', '?')} — removing redundant")
                    remove_idxs.add(j)
                    break
    deduped = [r for i, r in enumerate(deduped) if i not in remove_idxs]

    all_rows = deduped
    if before_dedup != len(all_rows):
        _log(f"  ok  Dedup: {before_dedup} → {len(all_rows)} shots")

    for i, row in enumerate(all_rows, 1):
        row["shot_number"] = i

    # ── Final 4-point validation ─────────────────────────────────────────────
    _log("  ── Final validation ──")
    _val_issues = 0

    _final_missing = _find_missing_sentences(all_rows)
    if _final_missing:
        _val_issues += len(_final_missing)
        _log(f"  ⚠ CHECK 1 FAIL: {len(_final_missing)} script sentence(s) still missing:")
        for _ms in _final_missing[:5]:
            _log(f"    - \"{_ms[:120]}\"")
    else:
        _log("  ✓ Check 1: All script sentences covered")

    _line_counts = {}
    for _r in all_rows:
        _ln = _norm_str(_r.get("line", ""))
        if _ln:
            _line_counts[_ln] = _line_counts.get(_ln, 0) + 1
    _dups = {k: v for k, v in _line_counts.items() if v > 1}
    if _dups:
        _val_issues += len(_dups)
        _log(f"  ⚠ CHECK 2 FAIL: {len(_dups)} duplicate Line text(s):")
        for _ln, _cnt in list(_dups.items())[:3]:
            _log(f"    - \"{_ln[:80]}\" x{_cnt}")
    else:
        _log("  ✓ Check 2: No duplicate lines")

    _final_truncated = _find_truncated_lines(all_rows)
    if _final_truncated:
        _val_issues += len(_final_truncated)
        _log(f"  ⚠ CHECK 3 FAIL: {len(_final_truncated)} truncated Line(s):")
        for _t in _final_truncated[:3]:
            _log(f"    - Shot {_t['shot']}: \"{_t['line'][:80]}\" -> full: \"{_t['full_sentence'][:80]}\"")
    else:
        _log("  ✓ Check 3: No truncated lines")

    _log("  ✓ Check 4: Rows ordered by script position (scan-based)")

    # Check 5: Line-Description coherence — detect decoupled rows where
    # the description talks about a different scene than the line
    _decoupled = []
    for _r in all_rows:
        _line = _r.get("line", "").strip()
        _desc = _r.get("shot_description", "").strip()
        if not _line or not _desc:
            continue
        _line_words = set(_norm(_line))
        _desc_words = set(_norm(_desc))
        _names_in_line = {w for w in _line_words if w[0:1].isupper()} if _line_words else set()
        _names_in_desc = {w for w in _desc_words if w[0:1].isupper()} if _desc_words else set()
        if _names_in_line and _names_in_desc:
            _name_overlap = len(_names_in_line & _names_in_desc) / max(len(_names_in_line), 1)
            if _name_overlap == 0 and len(_names_in_line) >= 2:
                _decoupled.append(_r.get("shot_number", "?"))
    if _decoupled:
        _val_issues += len(_decoupled)
        _log(f"  ⚠ CHECK 5 FAIL: {len(_decoupled)} shot(s) with Line/Description character mismatch:")
        for _sn in _decoupled[:5]:
            _log(f"    - Shot {_sn}")
    else:
        _log("  ✓ Check 5: Line-Description coherence OK")

    _log(f"  {'✓ All 5 checks passed' if _val_issues == 0 else f'⚠ {_val_issues} issue(s) found — review output'}")

    # ── Final count check ────────────────────────────────────────────────────
    if len(all_rows) < min_shots_global:
        _log(f"  !! WARNING: {len(all_rows)} shots delivered — below minimum {min_shots_global} for {word_count:,}-word script")
    else:
        _log(f"  ok  {len(all_rows)} shots delivered — within target range {min_shots_global}–{max_shots_global}")

    safe_name = re.sub(r'[^a-zA-Z0-9_\-]', '_', episode_name)
    xlsx_name = f"{safe_name}_breakdown.xlsx"
    write_excel(all_rows, job_dir / xlsx_name, episode_name)
    _log(f"  ok  Excel: {xlsx_name}  —  {len(all_rows)} shots  (required: {min_shots_global}–{max_shots_global})")
    return xlsx_name


def _run_stage2_core(job_id: str, job_dir: Path, episodes: list, ref_files_global: dict, detail_files: list, workers: int = 1):
    """Core Stage 2 logic — callable from standalone route or pipeline.
    workers > 1 processes multiple episodes in parallel."""
    _log(f"\n{'=' * 60}")
    _log(f"  STAGE 2 — Shot Breakdown Pipeline")
    _log(f"  Episodes to process: {len(episodes)}  |  Workers: {workers}")
    for i, ep in enumerate(episodes, 1):
        _log(f"    {i}. {ep['name']}")
    _log(f"{'=' * 60}\n")

    verify_auth()

    jobs[job_id]["progress"]["started_at"] = time.time()
    jobs[job_id]["progress"]["total"]     = len(episodes)
    jobs[job_id]["progress"]["shots_done"] = 0
    jobs[job_id]["progress"]["ep_shots"]   = 0

    output_files = []
    failures     = []

    if workers > 1 and len(episodes) > 1:
        _log(f"  >> Parallel mode: {workers} episodes at a time\n")
        _s2_lock = threading.Lock()
        _s2_done = [0]

        def _s2_episode_task(ep_tuple):
            i, ep = ep_tuple
            ep_name = ep["name"]
            log_queue = jobs[job_id]["queue"]
            _set_job_context(job_id, log_queue)
            _log(f"\n  ═══ EPISODE {i}/{len(episodes)}: {ep_name} ═══")

            ref_files = dict(ref_files_global)
            matched = _match_episode_detail(ep_name, detail_files)
            if matched:
                ref_files["episode_detail"] = matched
                _log(f"  Matched episode_detail for {ep_name}")
            elif len(detail_files) == 1:
                ref_files["episode_detail"] = detail_files[0][1]

            tok_before = dict(jobs[job_id]["tokens"])
            try:
                xlsx = _process_one_episode(job_dir, ep["script_text"], ep_name, ref_files)
                with _s2_lock:
                    output_files.append(xlsx)
                    jobs[job_id]["output_files"] = list(output_files)
            except Exception as e:
                _log(f"  [ERROR] Episode {ep_name} failed: {e}")
                with _s2_lock:
                    failures.append(ep_name)

            tok_after = jobs[job_id]["tokens"]
            ep_in = tok_after["input"] - tok_before["input"]
            ep_out = tok_after["output"] - tok_before["output"]
            ep_calls = tok_after["calls"] - tok_before["calls"]
            ep_cost = (ep_in * LLM_INPUT_COST_PER_MTOK + ep_out * LLM_OUTPUT_COST_PER_MTOK) / 1_000_000
            with _s2_lock:
                _s2_done[0] += 1
                jobs[job_id]["progress"]["completed"] = _s2_done[0]
                jobs[job_id]["progress"]["stage"] = f"Done {_s2_done[0]}/{len(episodes)}"
                jobs[job_id]["file_tokens"].append({
                    "name": ep_name,
                    "input": ep_in, "output": ep_out, "calls": ep_calls,
                    "cost": round(ep_cost, 4),
                })

        with ThreadPoolExecutor(max_workers=workers) as pool:
            list(pool.map(_s2_episode_task, list(enumerate(episodes, 1))))
    else:
        for i, ep in enumerate(episodes, 1):
            ep_name = ep["name"]
            _log(f"\n  ═══ EPISODE {i}/{len(episodes)}: {ep_name} ═══")
            jobs[job_id]["progress"]["stage"]    = f"Episode {i}/{len(episodes)}: {ep_name}"
            jobs[job_id]["progress"]["ep_shots"] = 0

            ref_files = dict(ref_files_global)
            matched = _match_episode_detail(ep_name, detail_files)
            if matched:
                ref_files["episode_detail"] = matched
                _log(f"  Matched episode_detail for {ep_name}")
            elif len(detail_files) == 1:
                ref_files["episode_detail"] = detail_files[0][1]
                _log(f"  Using only available episode_detail ({detail_files[0][0]})")

        tok_before = dict(jobs[job_id]["tokens"])
        try:
            xlsx = _process_one_episode(job_dir, ep["script_text"], ep_name, ref_files)
            output_files.append(xlsx)
            jobs[job_id]["output_files"] = list(output_files)
        except Exception as e:
            _log(f"  [ERROR] Episode {ep_name} failed: {e}")
            failures.append(ep_name)
        tok_after = jobs[job_id]["tokens"]
        ep_in = tok_after["input"] - tok_before["input"]
        ep_out = tok_after["output"] - tok_before["output"]
        ep_calls = tok_after["calls"] - tok_before["calls"]
        _log(f"  Tokens for {ep_name}: {ep_in:,} input / {ep_out:,} output / {ep_calls} API calls")
        ep_cost = (ep_in * LLM_INPUT_COST_PER_MTOK + ep_out * LLM_OUTPUT_COST_PER_MTOK) / 1_000_000
        jobs[job_id]["file_tokens"].append({
            "name": ep_name,
            "input": ep_in, "output": ep_out, "calls": ep_calls,
            "cost": round(ep_cost, 4),
        })

        jobs[job_id]["progress"]["completed"] = i

    jobs[job_id]["output_files"] = output_files
    jobs[job_id]["output_file"]  = output_files[0] if output_files else None

    tok = jobs[job_id]["tokens"]
    _log(f"\n{'=' * 60}")
    _log(f"  STAGE 2 COMPLETE")
    _log(f"  Successful: {len(output_files)}/{len(episodes)}")
    if failures:
        _log(f"  Failed:     {len(failures)} — {', '.join(failures)}")
    _log(f"  Total tokens: {tok['input']:,} input / {tok['output']:,} output / {tok['calls']} API calls")
    _log(f"{'=' * 60}\n")


def run_stage2_pipeline(job_id: str, job_dir: Path, episodes: list, ref_files_global: dict, detail_files: list):
    """Standalone entry point (called from API route)."""
    log_queue = jobs[job_id]["queue"]
    _set_job_context(job_id, log_queue)
    try:
        _run_stage2_core(job_id, job_dir, episodes, ref_files_global, detail_files)
        jobs[job_id]["status"] = "done"
    except Exception as e:
        tb = traceback.format_exc()
        _log(f"\n[ERROR] {e}\n{tb}")
        jobs[job_id]["status"] = "failed"
    finally:
        log_queue.put(None)



# ── Stage 3 — Script Audit Checker ────────────────────────────────────────────

S3_WORKSPACE = WORKSPACE / "stage3"
S3_WORKSPACE.mkdir(parents=True, exist_ok=True)

S3_AUDIT_SYSTEM_PROMPT = r"""You are a script audit checker for an AI shot breakdown pipeline. Your task is to verify and fix the Line column of a shot breakdown against the original source script.

WHAT THE LINE COLUMN IS:
The Line column must contain verbatim text copied directly from the source script — the exact words, punctuation, and sentence structure as written in the script file. It is not a summary, not a label, not a shorthand description, not a dash-separated keyword string.

THE #1 AUDIT RULE — VERBATIM COMPARISON ONLY (OVERRIDES ALL OTHER RULES):
Your ONLY job is to check whether each shot's Line text exists verbatim in the source script. If the Line matches the script text character for character, it is CORRECT and you MUST NOT flag it — no exceptions.
- Even if the dialogue attribution seems wrong ("Gordon asked" when you think Nara should be asking) — if the script says "Gordon asked", the shot is CORRECT.
- Even if the context seems illogical — if the Line matches the script, do NOT fix it.
- Even if you believe there is a typo or error in the script itself — you are NOT an editor. The script is the source of truth.
- BEFORE flagging ANY shot as Problem A, B, or C: search for the shot's EXACT Line text in the source script. If you find it — STOP. That shot is correct. Do not flag it. Move on.

LINE PROBLEMS TO IDENTIFY:

Problem A — Summarized or paraphrased Line
Any Line that compresses, shortens, or paraphrases the script text instead of quoting it verbatim.
How to spot it:
- Contains — used as a separator between fragments (e.g. "Gordon shifted attention to perfume collection — disinterested air")
- Drops words, clauses, or dialogue tags that exist in the script sentence
- Rewrites the script in different words
- Condenses two script sentences into a short label
- IMPORTANT: A Line is Problem A ONLY if its text does not match the script. If the Line is a word-for-word copy from the script, it is NOT Problem A regardless of whether the content seems correct or logical.
Fix: Replace with the exact verbatim script sentence(s) the Line represents. Copy character for character from the script.

Problem B — Invented Line with no script basis
Any Line whose text does not appear anywhere in the source script — production labels, location descriptions, visual directions, closing cards.
How to identify: Search for the Line's key phrases in the source script. If nothing matches, the entire row must be deleted.
Examples of invented Lines:
- "Gordon's bedroom — mirror, perfume bottles, early evening quiet"
- "CLOSING WIDE — Two figures entering the illuminated estate"
Fix: Flag the entire row for deletion.

Problem C — Truncated Line missing clauses
A Line that covers only part of a SINGLE script sentence AND the missing portion is NOT in the next shot.
Fix: Expand the Line to the full verbatim sentence including all clauses.

CRITICAL — WHEN PROBLEM C DOES NOT APPLY:
1. Consecutive SENTENCES split across shots: If Shot N has sentence A and Shot N+1 has sentence B, and both are complete verbatim text from the script, that is CORRECT. Do NOT merge.
2. Long sentence deliberately split at a visual boundary: If a long sentence is split between two shots (e.g. a description clause on one shot, the dialogue on the next), AND the next shot carries the remainder verbatim, that is an INTENTIONAL SPLIT for visual storytelling. Do NOT merge. Do NOT flag Problem C.
3. The mechanical test: If the current shot's Line text exists verbatim in the script, AND the next shot's Line text exists verbatim in the script, AND together they cover the full sentence — this is a valid split regardless of punctuation.

Problem C ONLY applies when part of a sentence is genuinely MISSING — not covered by any shot.

CRITICAL — NO CASCADE FIXES:
Never propose a fix that shifts Lines from one shot to the next in a chain. If shots 131-150 each have a unique verbatim sentence from the script in sequential order, they are ALL correct. Do not re-assign shot 132's Line to shot 131, then 133's to 132, etc. If every shot's Line exists verbatim in the script and the shots follow script order, there is nothing to fix.

Problem D — Intermediate shot Line duplicating parent instead of splitting
When a parent shot's Line covers multiple visual moments and has intermediate shots (e.g. 5.1, 5.2), the text must be SPLIT across the parent and its intermediates — each shot carries only the portion of text for its specific visual moment. No two shots should have the same Line text.
How to spot it: An intermediate shot (decimal number like 5.1, 16.1) has the exact same Line text as its parent shot (5, 16). This means the text was copied instead of split.
Fix: Split the parent's Line text so the parent keeps only its portion and each intermediate gets its own distinct portion. The concatenation of parent + intermediate Lines must equal the original full text. Every shot's Line must be unique.

DUPLICATE AND MISSING COVERAGE CHECKS:
- Exact duplicates: same Line text on two or more shots — including parent/intermediate pairs. Every shot must have unique Line text.
- Partial duplicates: one shot's Line is a substring of another shot's Line (and neither is an intentional split)
- Missing coverage: for each script beat, verify it appears in at least one shot's Line column
- Do not use fuzzy matching alone. Manually verify any flagged miss.
- Do not confirm full coverage until every beat is accounted for.

COMMON FAILURE MODES — NEVER DO THESE:
- Auditing the breakdown against itself — will never catch missing or wrong Lines
- Writing Lines from memory — any Line written from memory is unverified
- Fuzzy matching false misses — always manually verify before concluding a beat is missing
- Deleting a "duplicate" that is the only coverage of a script beat — always check the script first
- Dropping clauses — when fixing a truncated Line, read the full sentence from the script
- Paraphrasing instead of copying — close is not correct. The Line must be identical to the script, not similar
- Copying parent Line into intermediates — intermediate shots must carry their OWN portion of the text, not a duplicate of the parent's Line
- Merging intentional splits — when two consecutive shots each have a COMPLETE verbatim sentence from the script, do NOT merge them into one shot. This is deliberate splitting for visual storytelling. Only flag Problem C when a single sentence is cut mid-clause
- Cascade-fixing — never shift Lines from one shot to the next in a chain. If each shot has a unique verbatim script sentence in order, they are all correct
- Contextual reinterpretation — never "fix" a Line that matches the script verbatim. If the script says "Gordon asked" and the shot says "Gordon asked", do not change it based on your understanding of who should be speaking

OUTPUT FORMAT:
Return ONLY a valid JSON object with this structure:
{
  "fixes": [
    {
      "shot_num": <number or decimal>,
      "problem": "<A|B|C|D>",
      "original_line": "<current Line value in the breakdown>",
      "fixed_line": "<corrected verbatim Line from script>",
      "reason": "<brief explanation>"
    }
  ],
  "delete_rows": [
    {
      "shot_num": <number>,
      "original_line": "<current Line value>",
      "reason": "<why this row should be deleted>"
    }
  ],
  "notes": "<any coverage observations or warnings>"
}

For Problem B rows, include them in delete_rows (not in fixes).
For Problems A, C, D, include them in fixes with the corrected Line.
If a shot's Line is correct, do not include it in the output.
Output ONLY valid JSON, no other text.
"""

S3_AUDIT_CHUNK_SIZE = 30


def _s3_read_excel_full(excel_path: Path) -> tuple:
    """Read Excel preserving ALL columns for writeback."""
    from openpyxl import load_workbook
    wb = load_workbook(str(excel_path), data_only=True)
    ws = wb.active

    headers = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val is not None:
            headers.append((c, str(val).strip()))

    header_map = {h.lower(): c for c, h in headers}

    def _find(keywords):
        for kw in keywords:
            for h_lower, col in header_map.items():
                if kw in h_lower:
                    return col
        return None

    num_col  = _find(["#", "shot number", "num"])
    line_col = _find(["line", "narration", "dialogue"])
    desc_col = _find(["shot description", "description"])

    shots = []
    for r in range(2, ws.max_row + 1):
        row_data = {}
        for c, h in headers:
            row_data[h] = ws.cell(r, c).value
        num_val = ws.cell(r, num_col).value if num_col else r - 1
        line_val = ws.cell(r, line_col).value if line_col else ""
        desc_val = ws.cell(r, desc_col).value if desc_col else ""
        if not line_val and not desc_val:
            continue
        shots.append({
            "row": r,
            "num": num_val,
            "line": str(line_val or "").strip(),
            "shot_description": str(desc_val or "").strip(),
            "columns": row_data,
        })

    col_names = [h for _, h in headers]
    wb.close()
    return shots, col_names


def _s3_read_script_file(script_path: Path) -> str:
    suffix = script_path.suffix.lower()
    if suffix == ".docx":
        try:
            import mammoth
            with open(script_path, "rb") as f:
                return mammoth.convert_to_markdown(f).value.strip()
        except ImportError:
            raise RuntimeError("pip install mammoth — needed to read .docx")
    return script_path.read_text(encoding="utf-8", errors="replace")


def _s3_format_audit_chunk(shots: list) -> str:
    lines = []
    for s in shots:
        sn = s["num"]
        line = s["line"]
        desc = s["shot_description"]
        lines.append(f"Shot {sn}:\n  Line: {line}\n  Description: {desc}")
    return "\n\n".join(lines)


def _s3_call_audit(script_text: str, chunk_text: str, label: str) -> str:
    user_msg = (
        f"TASK: Run the LINE COLUMN AUDIT on the following shots. Compare each shot's Line column "
        f"against the source script below. Identify all problems (A: summarized/paraphrased, "
        f"B: invented with no script basis, C: truncated/missing clauses, D: intermediate "
        f"Line duplicating parent instead of splitting) and provide fixes.\n\n"
        f"{'=' * 40} SOURCE SCRIPT {'=' * 40}\n{script_text}\n\n"
        f"{'=' * 40} SHOTS TO AUDIT {'=' * 40}\n{chunk_text}\n\n"
        f"Return ONLY a JSON object with 'fixes', 'delete_rows', and 'notes' fields."
    )
    return call_api_chat(
        system=S3_AUDIT_SYSTEM_PROMPT,
        messages=[{"role": "user", "content": user_msg}],
        label=label,
        timeout=180,
        retries=2,
    )


def _s3_parse_audit_response(response: str) -> dict:
    import json as _json
    text = response.strip()
    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1:
        return {"fixes": [], "delete_rows": [], "notes": ""}
    try:
        obj = _json.loads(text[start:end + 1])
        if not isinstance(obj, dict):
            return {"fixes": [], "delete_rows": [], "notes": ""}
        return {
            "fixes": obj.get("fixes", []),
            "delete_rows": obj.get("delete_rows", []),
            "notes": obj.get("notes", ""),
        }
    except _json.JSONDecodeError:
        return {"fixes": [], "delete_rows": [], "notes": ""}


def _s3_apply_audit_fixes(all_fixes: list, all_deletes: list,
                          input_path: Path, output_path: Path) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(str(input_path))
    ws = wb.active

    header_map = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val:
            header_map[str(val).strip().lower()] = c

    def _find_col(keywords):
        for kw in keywords:
            for h, c in header_map.items():
                if kw in h:
                    return c
        return None

    num_col = _find_col(["#", "shot number", "num"])
    line_col = _find_col(["line", "narration", "dialogue"])

    if not line_col:
        wb.close()
        return {"fixed": 0, "deleted": 0}

    shot_row_map = {}
    total_rows = 0
    for r in range(2, ws.max_row + 1):
        cell_val = ws.cell(r, num_col).value if num_col else r - 1
        if cell_val is not None:
            shot_row_map[str(cell_val)] = r
            total_rows += 1

    fixed_count = 0
    for fix in all_fixes:
        sn = str(fix.get("shot_num", ""))
        fixed_line = fix.get("fixed_line", "")
        if not sn or not fixed_line:
            continue
        row = shot_row_map.get(sn)
        if row:
            ws.cell(row, line_col, fixed_line)
            fixed_count += 1

    rows_to_delete = []
    for d in all_deletes:
        sn = str(d.get("shot_num", ""))
        row = shot_row_map.get(sn)
        if row:
            rows_to_delete.append(row)

    rows_to_delete.sort(reverse=True)
    for row in rows_to_delete:
        ws.delete_rows(row)

    if rows_to_delete and num_col:
        whole_num = 0
        last_whole = 0
        for r in range(2, ws.max_row + 1):
            cell_val = ws.cell(r, num_col).value
            if cell_val is None:
                continue
            try:
                fval = float(cell_val)
                if fval == int(fval):
                    whole_num += 1
                    last_whole = whole_num
                    ws.cell(r, num_col, whole_num)
                else:
                    decimal_part = round(fval - int(fval), 1)
                    ws.cell(r, num_col, last_whole + decimal_part)
            except (ValueError, TypeError):
                pass

    wb.save(str(output_path))
    return {"fixed": fixed_count, "deleted": len(rows_to_delete)}


def _run_s3_one(job_id: str, job_dir: Path, excel_path: Path, script_text: str) -> str:
    _log("=" * 60)
    _log(f"  STAGE 3 — Script Audit Checker")
    _log(f"  File: {excel_path.name}")
    _log("=" * 60)

    _log("  Reading shot breakdown Excel…")
    shots, col_names = _s3_read_excel_full(excel_path)
    _log(f"  ok  {len(shots)} shots found, {len(col_names)} columns")
    _log(f"  Script length: {len(script_text):,} characters")

    total_chunks = max(1, (len(shots) + S3_AUDIT_CHUNK_SIZE - 1) // S3_AUDIT_CHUNK_SIZE)
    jobs[job_id]["progress"]["total"] = total_chunks + 1
    jobs[job_id]["progress"]["stage"] = f"{excel_path.stem}"

    all_fixes = []
    all_deletes = []
    all_notes = []

    for chunk_idx in range(total_chunks):
        c_start = chunk_idx * S3_AUDIT_CHUNK_SIZE
        c_end = min(c_start + S3_AUDIT_CHUNK_SIZE, len(shots))
        chunk = shots[c_start:c_end]

        shot_range = f"{chunk[0]['num']}–{chunk[-1]['num']}"
        _log(f"\n  [{chunk_idx + 1}/{total_chunks}]  Auditing shots {shot_range} ({len(chunk)} shots)")

        chunk_text = _s3_format_audit_chunk(chunk)
        response = _s3_call_audit(script_text, chunk_text, f"audit chunk {chunk_idx + 1}")

        if response:
            result = _s3_parse_audit_response(response)
            fixes = result["fixes"]
            deletes = result["delete_rows"]
            notes = result["notes"]

            all_fixes.extend(fixes)
            all_deletes.extend(deletes)
            if notes:
                all_notes.append(notes)

            _log(f"  ok  {len(fixes)} fix(es), {len(deletes)} row(s) to delete")
            for fix in fixes:
                _log(f"      Shot {fix.get('shot_num')}: Problem {fix.get('problem')} — {fix.get('reason', '')[:80]}")
            for d in deletes:
                _log(f"      Shot {d.get('shot_num')}: DELETE — {d.get('reason', '')[:80]}")
        else:
            _log(f"  !!  No response for chunk {chunk_idx + 1}")

        jobs[job_id]["progress"]["completed"] = chunk_idx + 1

    _log(f"\n  [{total_chunks + 1}/{total_chunks + 1}]  Running coverage check…")
    line_summary = []
    for s in shots:
        line_summary.append(f"Shot {s['num']}: {s['line']}")
    compact_lines = "\n".join(line_summary)

    coverage_msg = (
        f"TASK: Check coverage between the source script and the shot breakdown Lines below.\n\n"
        f"For each sentence/beat in the script, verify it appears in at least one shot's Line column "
        f"(either as the full sentence or a confirmed partial split). Report:\n"
        f"1. Any script beats NOT covered by any shot Line\n"
        f"2. Any exact duplicate Lines (same text on multiple shots — including parent/intermediate pairs)\n"
        f"3. Any partial duplicates (one Line is substring of another)\n\n"
        f"{'=' * 40} SOURCE SCRIPT {'=' * 40}\n{script_text}\n\n"
        f"{'=' * 40} ALL SHOT LINES {'=' * 40}\n{compact_lines}\n\n"
        f"Return ONLY a JSON object:\n"
        f'{{"missing_beats": ["<script sentence not in any Line>", ...], '
        f'"duplicates": [{{"shots": [<num>, <num>], "line": "<text>"}}], '
        f'"coverage_pct": <estimated percentage of script covered>, '
        f'"notes": "<summary>"}}'
    )

    coverage_response = call_api_chat(
        system="You are a coverage checker. Compare script text against shot Line columns. Output only valid JSON.",
        messages=[{"role": "user", "content": coverage_msg}],
        label="coverage check",
        timeout=180,
        retries=2,
    )

    coverage_result = None
    if coverage_response:
        import json as _json
        text = coverage_response.strip()
        s_idx = text.find("{")
        e_idx = text.rfind("}")
        if s_idx != -1 and e_idx != -1:
            try:
                coverage_result = _json.loads(text[s_idx:e_idx + 1])
            except _json.JSONDecodeError:
                pass

    if coverage_result:
        missing = coverage_result.get("missing_beats", [])
        dupes = coverage_result.get("duplicates", [])
        cov_pct = coverage_result.get("coverage_pct", "?")
        _log(f"  ok  Coverage: ~{cov_pct}%")
        if missing:
            _log(f"  !!  {len(missing)} script beat(s) not covered:")
            for m in missing[:10]:
                _log(f"      — {str(m)[:100]}")
            if len(missing) > 10:
                _log(f"      … and {len(missing) - 10} more")
        else:
            _log("  ok  All script beats covered")
        if dupes:
            _log(f"  !!  {len(dupes)} duplicate Line(s) found")
            for d in dupes[:5]:
                _log(f"      Shots {d.get('shots', [])}: {str(d.get('line', ''))[:80]}")
    else:
        _log("  !!  Coverage check failed — no parseable response")

    jobs[job_id]["progress"]["completed"] = total_chunks + 1

    _log(f"\n  >>  Total: {len(all_fixes)} fix(es), {len(all_deletes)} deletion(s)")

    jobs[job_id]["flags"] = {
        "fixes": all_fixes,
        "delete_rows": all_deletes,
        "coverage": coverage_result,
        "notes": all_notes,
    }
    jobs[job_id]["excel_path"] = str(excel_path)

    if all_fixes or all_deletes:
        _log(f"\n{'=' * 60}")
        _log(f"  AUDIT COMPLETE — issues flagged")
        _log(f"  Line fixes: {len(all_fixes)}")
        _log(f"  Rows to delete: {len(all_deletes)}")
        if coverage_result:
            _log(f"  Script coverage: ~{coverage_result.get('coverage_pct', '?')}%")
        _log(f"  Review flags and click 'Fix Issues' to apply.")
        _log(f"{'=' * 60}\n")
    else:
        import shutil
        output_name = f"{excel_path.stem}_audited.xlsx"
        output_path = job_dir / output_name
        shutil.copy2(str(excel_path), str(output_path))
        jobs[job_id]["output_files"] = [output_name]
        jobs[job_id]["output_file"] = output_name
        _log(f"\n{'=' * 60}")
        _log(f"  AUDIT COMPLETE — No issues found")
        _log(f"  All {len(shots)} shots have correct Line values")
        if coverage_result:
            _log(f"  Script coverage: ~{coverage_result.get('coverage_pct', '?')}%")
        _log(f"  Output: {output_name}")
        _log(f"{'=' * 60}\n")

    if all_notes:
        _log("\n  Audit notes:")
        for note in all_notes:
            _log(f"    {note[:200]}")

    return bool(all_fixes or all_deletes)


def _run_stage3_batch_core(job_id: str, job_dir: Path, excel_paths: list, script_text: str):
    """Core Stage 3 logic — callable from standalone route or pipeline."""
    n = len(excel_paths)
    _log(f"  Stage 3 — Script Audit Checker — {n} file(s) to audit\n")
    _log(f"  Script: {len(script_text):,} characters\n")

    has_flags = False
    for i, ep in enumerate(excel_paths, 1):
        if n > 1:
            _log(f"\n  ──── File {i}/{n}: {ep.name} ────")
        tok_before = dict(jobs[job_id]["tokens"])
        try:
            flagged = _run_s3_one(job_id, job_dir, ep, script_text)
            if flagged:
                has_flags = True
        except Exception as e:
            _log(f"[ERROR] {ep.name} failed: {e}")
            import traceback; traceback.print_exc()
        tok_after = jobs[job_id]["tokens"]
        f_in = tok_after["input"] - tok_before["input"]
        f_out = tok_after["output"] - tok_before["output"]
        f_calls = tok_after["calls"] - tok_before["calls"]
        f_cost = (f_in * LLM_INPUT_COST_PER_MTOK + f_out * LLM_OUTPUT_COST_PER_MTOK) / 1_000_000
        jobs[job_id]["file_tokens"].append({
            "name": ep.stem,
            "input": f_in, "output": f_out, "calls": f_calls,
            "cost": round(f_cost, 4),
        })

    if has_flags:
        jobs[job_id]["status"] = "flagged"
    else:
        jobs[job_id]["status"] = "done" if jobs[job_id].get("output_files") else "failed"

    if n > 1:
        _log(f"\n{'=' * 60}")
        _log(f"  ALL DONE — audit complete for {n} file(s)")
        _log(f"{'=' * 60}")


def run_stage3_batch(job_id: str, job_dir: Path, excel_paths: list, script_text: str):
    """Standalone entry point (called from API route)."""
    log_queue = jobs[job_id]["queue"]
    _set_job_context(job_id, log_queue)
    try:
        _run_stage3_batch_core(job_id, job_dir, excel_paths, script_text)
    except Exception as e:
        _log(f"[ERROR] Stage 3 batch failed: {e}")
        import traceback; traceback.print_exc()
        jobs[job_id]["status"] = "failed"
    finally:
        log_queue.put(None)


# ── Stage 4 — Generate Reference Images (Gemini Imagen 4) ────────────────────

S4_WORKSPACE = WORKSPACE / "stage4"
S4_WORKSPACE.mkdir(parents=True, exist_ok=True)

S4_COST_PER_IMAGE = 0.02
S4_RATE_LIMIT_DELAY = 32
_madeye_pool = MadEyeKeyPool(MADEYE_API_KEYS, S4_RATE_LIMIT_DELAY) if MADEYE_API_KEYS else None


def _s4_track(job_id: str, calls: int = 0, images: int = 0):
    if job_id not in jobs:
        return
    with jobs[job_id]["tokens_lock"]:
        jobs[job_id]["tokens"]["calls"]  += calls
        jobs[job_id]["tokens"]["output"] += images
        jobs[job_id]["tokens"]["input"]  += calls


def _s4_parse_md(md_path: Path) -> list:
    """Parse a character_canvas or location_reference (.md or .docx) into items for image gen."""
    import re as _re
    if md_path.suffix.lower() == ".docx":
        import mammoth
        with open(md_path, "rb") as f:
            text = mammoth.convert_to_markdown(f).value.strip()
    else:
        text = md_path.read_text(encoding="utf-8", errors="replace")

    is_char = "character" in md_path.stem.lower()
    canvas_type = "character" if is_char else "location"

    items = []
    sections = _re.split(r'\n###\s+', text)
    for sec in sections[1:]:
        lines = sec.strip().split("\n")
        name = lines[0].strip().rstrip("#").strip()
        if not name or name.lower().startswith("relationship") or name.lower().startswith("output"):
            continue

        body = "\n".join(lines[1:])

        if is_char:
            desc_parts = []
            phys_match = _re.search(r'(?:####?\s*Physical Description)(.*?)(?=####|\Z)', body, _re.S)
            if phys_match:
                table_rows = _re.findall(r'\|\s*([^|]+?)\s*\|\s*([^|]+?)\s*\|', phys_match.group(1))
                for feat, val in table_rows:
                    feat, val = feat.strip(), val.strip()
                    if feat.lower() in ('feature', '---', ''):
                        continue
                    desc_parts.append(f"{feat}: {val}")
            outfit_match = _re.search(r'(?:####?\s*Default Outfit.*?)(.*?)(?=####|\Z)', body, _re.S)
            if outfit_match:
                outfit_lines = [l.strip().lstrip("-* ").strip() for l in outfit_match.group(1).strip().split("\n") if l.strip().startswith(("-", "*"))]
                if outfit_lines:
                    desc_parts.append("Outfit: " + "; ".join(outfit_lines))
            tells_match = _re.search(r'(?:####?\s*Key Visual Tells)(.*?)(?=####|\Z)', body, _re.S)
            if tells_match:
                tell_lines = [l.strip().lstrip("-* ").strip() for l in tells_match.group(1).strip().split("\n") if l.strip().startswith(("-", "*"))]
                if tell_lines:
                    desc_parts.append("Key visual tells: " + "; ".join(tell_lines))
            desc = ". ".join(desc_parts) if desc_parts else body[:500]
        else:
            desc_parts = []
            phys_match = _re.search(r'(?:####?\s*Physical Description)(.*?)(?=####|\Z)', body, _re.S)
            if phys_match:
                phys_lines = [l.strip().lstrip("-* ").strip() for l in phys_match.group(1).strip().split("\n") if l.strip().startswith(("-", "*"))]
                if phys_lines:
                    desc_parts.extend(phys_lines)
            atmo_match = _re.search(r'(?:####?\s*Atmosphere.*?Mood)(.*?)(?=####|\Z)', body, _re.S)
            if atmo_match:
                atmo_lines = [l.strip().lstrip("-* ").strip() for l in atmo_match.group(1).strip().split("\n") if l.strip().startswith(("-", "*"))]
                if atmo_lines:
                    desc_parts.extend(atmo_lines)
            type_match = _re.search(r'\*\*Type:\*\*\s*(.*)', body)
            if type_match:
                desc_parts.insert(0, type_match.group(1).strip())
            desc = ". ".join(desc_parts) if desc_parts else body[:500]

        items.append({
            "type": canvas_type,
            "name": name,
            "description": desc,
            "original_url": "",
        })
        _log(f"  ->  Parsed {canvas_type}: {name}")

    return items


def _s4_parse_canvas(canvas_path: Path, canvas_type: str) -> list:
    import openpyxl
    items = []
    wb = openpyxl.load_workbook(str(canvas_path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]

    name_col = None
    desc_col = None
    url_col = None

    for i, h in enumerate(headers):
        if canvas_type == "character":
            if "character" in h and "name" in h:
                name_col = i
            elif h == "prompt" or ("description" in h):
                desc_col = i
        else:
            if "location" in h and "name" in h:
                name_col = i
            elif "name" in h and name_col is None:
                name_col = i
            elif h == "prompt" or "description" in h:
                desc_col = i
        if "image" in h and "url" in h:
            url_col = i

    if name_col is None:
        _log(f"  !!  {canvas_type} canvas: no name column found")
        wb.close()
        return items

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[name_col] or "").strip()
        if not name:
            continue
        desc = str(row[desc_col] or "").strip() if desc_col is not None else ""
        url  = str(row[url_col] or "").strip() if url_col is not None else ""
        items.append({
            "type": canvas_type,
            "name": name,
            "description": desc,
            "original_url": url if url.startswith("http") else "",
        })
    wb.close()
    return items


def _s4_generate_one(prompt: str, job_id: str, item_num, job_dir: Path, retries: int = 2, regen_job_id: str = None, api_key: str = None) -> str:
    import base64 as _b64
    track_id = regen_job_id or job_id
    url = f"{MADEYE_BASE_URL}/v1/images/generations"
    headers = {
        "Authorization": f"Bearer {api_key or MADEYE_API_KEY}",
        "Content-Type": "application/json",
    }
    body = {
        "model": "imagen-4.0-fast-generate-001",
        "prompt": prompt,
        "n": 1,
        "response_format": "b64_json",
        "user": "int-sahni.manas@pocketfm.com",
        "metadata": {"user_email": "int-sahni.manas@pocketfm.com"},
    }

    for attempt in range(1, retries + 2):
        try:
            with httpx.Client(timeout=120) as hc:
                resp = hc.post(url, json=body, headers=headers)
            _s4_track(track_id, calls=1)
            if resp.status_code == 200:
                data = resp.json()
                items = data.get("data", [])
                if items and items[0].get("b64_json"):
                    img_bytes = _b64.b64decode(items[0]["b64_json"])
                    img_dir = job_dir / "images"
                    img_dir.mkdir(exist_ok=True)
                    fname = f"ref_{item_num}.png"
                    (img_dir / fname).write_bytes(img_bytes)
                    _s4_track(track_id, images=1)
                    return f"/api/stage4/image/{job_id}/{fname}"
                elif items and items[0].get("url"):
                    with httpx.Client(timeout=60) as dl:
                        img_resp = dl.get(items[0]["url"])
                    if img_resp.status_code == 200:
                        img_dir = job_dir / "images"
                        img_dir.mkdir(exist_ok=True)
                        fname = f"ref_{item_num}.png"
                        (img_dir / fname).write_bytes(img_resp.content)
                        _s4_track(track_id, images=1)
                        return f"/api/stage4/image/{job_id}/{fname}"
                _log(f"  x   No image data in response (attempt {attempt})")
            else:
                err = resp.text[:200]
                _log(f"  x   Attempt {attempt}: HTTP {resp.status_code}: {err}")
                if resp.status_code == 400 and ("SAFETY" in resp.text.upper() or "BLOCKED" in resp.text.upper()):
                    _log("  !!  Prompt blocked by safety filter — skipping")
                    return ""
        except Exception as e:
            _log(f"  x   Attempt {attempt}: {str(e)[:200]}")
        if attempt <= retries:
            wait = 30 * attempt
            _log(f"  -> Retrying in {wait}s...")
            time.sleep(wait)
    return ""


def _s4_write_output_excel(items: list, output_path: Path):
    import openpyxl
    from openpyxl import Workbook
    wb = Workbook()

    chars = [it for it in items if it["type"] == "character"]
    locs  = [it for it in items if it["type"] == "location"]

    def _write_sheet(ws, rows, name_header):
        headers_list = [name_header, "prompt", "image_url", "preview"]
        for c, h in enumerate(headers_list, 1):
            ws.cell(1, c, h)
            ws.cell(1, c).font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 60
        for i, item in enumerate(rows, 1):
            ws.cell(i + 1, 1, item["name"])
            ws.cell(i + 1, 2, item.get("description", ""))
            gen_url = item.get("generated_url", "")
            ws.cell(i + 1, 3, gen_url)
            if gen_url:
                ws.cell(i + 1, 4, f'=IMAGE(C{i + 1})')

    ws_char = wb.active
    ws_char.title = "Characters"
    _write_sheet(ws_char, chars, "character_name")

    if locs:
        ws_loc = wb.create_sheet("Locations")
        _write_sheet(ws_loc, locs, "location_name")

    wb.save(str(output_path))


def _run_stage4_core(job_id: str, job_dir: Path, items: list):
    """Core Stage 4 logic — callable from standalone route or pipeline."""
    total = len(items)
    jobs[job_id]["progress"]["total"] = total
    _log(f"  >>  Stage 4: Generating reference images for {total} items")
    jobs[job_id]["s4_items"] = items

    # Prepare all tasks (build prompts)
    s4_tasks = []
    for idx, item in enumerate(items, 1):
        itype = item["type"]
        name  = item["name"]
        desc  = item.get("description", "")

        if itype == "character":
            prompt = (
                "Professional fashion editorial photograph, real human adult with realistic adult body proportions. "
                "The subject's head is small relative to the body, exactly 1/7.5 of total height. "
                "Long legs, normal-sized torso, photographed from mid-distance showing full body head to shoes. "
                f"Subject: {name}. {desc} "
                "Shot on 85mm lens, studio lighting, clean neutral background. "
                "8K photorealistic. PURE IMAGE ONLY — zero text, no words, no watermark."
            )
        else:
            prompt = (
                f"8K photorealistic wide establishing shot of {name}. {desc} "
                "Cinematic composition, detailed environment, atmospheric lighting. "
                "PURE IMAGE ONLY — zero text, no words, no letters, no watermark."
            )
        item["prompt"] = prompt
        s4_tasks.append((idx, item, prompt))

    # Execute with key pool
    num_workers = _madeye_pool.key_count if _madeye_pool else 1
    _log(f"  >>  Generating {len(s4_tasks)} reference images with {num_workers} key(s)")
    _s4_gen_lock = threading.Lock()
    _s4_completed = [0]

    def _s4_gen_task(task_tuple):
        idx, item, prompt = task_tuple
        itype = item["type"]
        name  = item["name"]
        _log(f"  [{idx}/{total}]  {itype.title()}: {name}")

        key = _madeye_pool.acquire() if _madeye_pool else None
        try:
            url = _s4_generate_one(prompt, job_id, idx, job_dir, api_key=key)
        finally:
            if _madeye_pool:
                _madeye_pool.release_notify()

        with _s4_gen_lock:
            if url:
                item["generated_url"] = url
                _log(f"  ok  Generated image for {name}")
            else:
                item["generated_url"] = ""
                _log(f"  !!  Failed to generate image for {name}")
            jobs[job_id]["file_tokens"].append({
                "name": f"{itype.title()}: {name}",
                "images": 1 if url else 0,
                "cost": round(S4_COST_PER_IMAGE if url else 0, 4),
            })
            _s4_completed[0] += 1
            jobs[job_id]["progress"]["completed"] = _s4_completed[0]

    with ThreadPoolExecutor(max_workers=num_workers) as pool:
        list(pool.map(_s4_gen_task, s4_tasks))
    jobs[job_id]["progress"]["stage"] = "Writing output"

    output_name = f"reference_images_{job_id[:8]}.xlsx"
    output_path = job_dir / output_name
    _s4_write_output_excel(items, output_path)

    jobs[job_id]["output_file"] = output_name
    jobs[job_id]["output_files"] = [output_name]

    s4_img_count = sum(1 for ft in jobs[job_id].get("file_tokens", []) if ft.get("images"))
    _log(f"  ==  Done — {s4_img_count} images | ${s4_img_count * S4_COST_PER_IMAGE:.2f}")


def run_stage4_pipeline(job_id: str, job_dir: Path, items: list):
    """Standalone entry point (called from API route)."""
    log_queue = jobs[job_id]["queue"]
    _set_job_context(job_id, log_queue)
    try:
        _run_stage4_core(job_id, job_dir, items)
        jobs[job_id]["status"] = "done"
    except Exception as e:
        _log(f"[ERROR] Stage 4 failed: {e}")
        import traceback; traceback.print_exc()
        jobs[job_id]["status"] = "failed"
    finally:
        log_queue.put(None)


# ── Stage 5 — Image Generation (Gemini Imagen 4) ─────────────────────────────

S5_WORKSPACE = WORKSPACE / "stage5"
S5_WORKSPACE.mkdir(parents=True, exist_ok=True)

S5_PROMPT_PREFIX = "8K photorealistic single-frame photograph. "
S5_PROMPT_SUFFIX = " PURE IMAGE ONLY — zero text, no words, no letters, no title, no watermark, no overlay."
S5_NEGATIVE = (
    "Negative constraints: no face generation beyond provided references, "
    "no identity blending between characters, no distorted or duplicated facial features, "
    "no extra eyes noses or asymmetry errors, no blur or loss of facial clarity."
)

S5_RATE_LIMIT_DELAY = 32  # seconds between images (free tier: 2/min)

INSERT_KEYWORDS = [
    "hands", "hand", "fingers", "phone", "glass", "cup", "door", "letter",
    "book", "weapon", "sword", "knife", "ring", "key", "scroll", "map",
    "close-up of", "insert shot", "detail shot", "object",
]


def _s5_read_excel(excel_path: Path) -> list:
    from openpyxl import load_workbook
    wb = load_workbook(str(excel_path), data_only=True)
    ws = wb.active
    headers = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]

    def _find(keywords):
        for i, h in enumerate(headers):
            for kw in keywords:
                if kw in h:
                    return i + 1
        return None

    num_col    = _find(["#", "num", "shot"])
    line_col   = _find(["line", "narration", "dialogue", "text"])
    desc_col   = _find(["shot description", "description"])
    detail_col = _find(["shot detail", "detail"])
    size_col   = _find(["shot size", "size"])
    ref_col    = _find(["reference"])

    shots = []
    for r in range(2, ws.max_row + 1):
        line = ws.cell(r, line_col).value if line_col else None
        if not line:
            continue
        shots.append({
            "row":              r,
            "num":              ws.cell(r, num_col).value if num_col else r - 1,
            "line":             str(line).strip(),
            "shot_description": str(ws.cell(r, desc_col).value or "").strip() if desc_col else "",
            "shot_detail":      str(ws.cell(r, detail_col).value or "").strip() if detail_col else "",
            "shot_size":        str(ws.cell(r, size_col).value or "").strip() if size_col else "",
            "reference":        str(ws.cell(r, ref_col).value or "").strip() if ref_col else "",
            "preview_1":        "",
        })
    return shots


def _s5_build_ref_map(char_dir: Path) -> dict:
    ref_map = {}
    if not char_dir.exists():
        return ref_map

    canvas_files = list(char_dir.glob("*.xlsx"))
    if canvas_files:
        _log("  >>  Parsing character canvas Excel…")
        ref_map = _s5_parse_canvas(canvas_files[0], char_dir)
        return ref_map

    for img_path in sorted(char_dir.glob("*")):
        if img_path.suffix.lower() not in (".png", ".jpg", ".jpeg", ".webp"):
            continue
        name = img_path.stem.replace("_", " ").replace("-", " ").title()
        ref_map[name] = {"path": img_path}
        _log(f"  ok  Character ref: {name} ({img_path.name})")
    return ref_map


def _s5_parse_canvas(canvas_path: Path, dest_dir: Path) -> dict:
    import openpyxl
    ref_map = {}
    wb = openpyxl.load_workbook(str(canvas_path), read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        name_col = None
        url_col = None
        for i, h in enumerate(headers):
            if ("character" in h or "location" in h) and "name" in h:
                name_col = i
            elif "name" in h and name_col is None:
                name_col = i
            elif "image" in h and "url" in h:
                url_col = i
        if name_col is None or url_col is None:
            continue

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            cname = str(row[name_col] or "").strip()
            curl  = str(row[url_col] or "").strip()
            if cname and curl and (curl.startswith("http") or curl.startswith("/api/")):
                rows.append((cname, curl))

        _log(f"  ok  Found {len(rows)} entries in sheet '{sheet_name}'")
        for cname, curl in rows:
            safe_name = re.sub(r'[^\w\s-]', '', cname).strip().replace(' ', '_')
            img_path = dest_dir / f"{safe_name}.png"

            if curl.startswith("/api/stage4/image/"):
                parts = curl.strip("/").split("/")
                if len(parts) >= 4:
                    s4_job_id = parts[2]
                    s4_fname  = parts[3]
                    local_path = S4_WORKSPACE / s4_job_id / "images" / s4_fname
                    if local_path.exists():
                        import shutil
                        shutil.copy2(str(local_path), str(img_path))
                        ref_map[cname] = {"path": img_path}
                        _log(f"  ok  Ref (local): {cname} → {img_path.name}")
                        continue
                    _log(f"  !!  Local image not found for {cname}: {local_path}")
                    continue

            try:
                resp = httpx.get(curl, timeout=30, follow_redirects=True)
                if resp.status_code == 200 and len(resp.content) > 500:
                    img_path.write_bytes(resp.content)
                    ref_map[cname] = {"path": img_path}
                    _log(f"  ok  Ref (download): {cname} → {img_path.name}")
                else:
                    _log(f"  !!  Failed to download {cname}: HTTP {resp.status_code}")
            except Exception as e:
                _log(f"  !!  Failed to download {cname}: {e}")

    wb.close()
    return ref_map


def _s5_match_characters(shot_desc: str, shot_detail: str, ref_map: dict) -> list:
    combined = (shot_desc + " " + shot_detail).lower()
    matched = []
    for name in ref_map:
        name_lower = name.lower()
        name_parts = name_lower.split()
        if name_lower in combined:
            matched.append(name)
        elif any(part in combined and len(part) >= 3 for part in name_parts):
            matched.append(name)
    return matched


def _s5_is_insert_shot(shot_desc: str) -> bool:
    desc_lower = shot_desc.lower()
    return any(kw in desc_lower for kw in INSERT_KEYWORDS)


def _s5_select_refs(shot: dict, ref_map: dict, last_char: str):
    desc   = shot["shot_description"]
    detail = shot["shot_detail"]
    matched_names = _s5_match_characters(desc, detail, ref_map)

    if not matched_names and _s5_is_insert_shot(desc) and last_char and last_char in ref_map:
        matched_names = [last_char]

    if not matched_names:
        return [], "", last_char, []

    ref_images = []
    for name in matched_names:
        entry = ref_map.get(name, {})
        path = entry.get("path")
        if path and Path(path).exists():
            ref_images.append({"name": name, "path": Path(path)})

    ref_instruction = ""
    if ref_images:
        if len(matched_names) == 1:
            ref_instruction = (
                f"REFERENCE IMAGE PROVIDED for {matched_names[0]} — "
                f"the attached reference image is the authoritative visual for this character. "
                f"Replicate the exact face, skin tone, hair, build, and features from the reference. "
                f"Do NOT invent or alter any facial features. "
            )
        else:
            ref_instruction = (
                f"REFERENCE IMAGES PROVIDED for {', '.join(matched_names)}. "
                f"Each attached reference image is the authoritative visual for that character. "
                f"Replicate each character's exact face, skin tone, hair, build, and features from their reference. "
                f"Do NOT blend or swap features between characters. "
                f"Clearly separate all characters spatially with no overlapping faces. "
            )
    else:
        if len(matched_names) == 1:
            ref_instruction = (
                f"Character: {matched_names[0]}. "
                f"Maintain consistent facial structure, skin tone, hairstyle, and features for this character. "
            )
        elif len(matched_names) >= 2:
            ref_instruction = (
                f"Characters: {', '.join(matched_names)}. "
                f"Each character must be visually distinct and recognizable. "
                f"Clearly separate all characters spatially with no overlapping faces. "
            )

    new_last = matched_names[0]
    return matched_names, ref_instruction, new_last, ref_images


S5_COST_PER_IMAGE = 0.02

def _s5_track(job_id: str, calls: int = 0, images: int = 0):
    if job_id not in jobs:
        return
    with jobs[job_id]["tokens_lock"]:
        jobs[job_id]["tokens"]["calls"]  += calls
        jobs[job_id]["tokens"]["output"] += images
        jobs[job_id]["tokens"]["input"]  += calls

def _s5_generate_one(prompt: str, job_id: str,
                      shot_num, job_dir: Path, retries: int = 2,
                      regen_job_id: str = None, ref_images: list = None,
                      api_key: str = None) -> str:
    import base64 as _b64
    track_id = regen_job_id or job_id
    url = f"{MADEYE_BASE_URL}/v1/images/generations"
    headers = {
        "Authorization": f"Bearer {api_key or MADEYE_API_KEY}",
        "Content-Type": "application/json",
    }
    body = {
        "model": "imagen-4.0-fast-generate-001",
        "prompt": prompt,
        "n": 1,
        "response_format": "b64_json",
        "user": "int-sahni.manas@pocketfm.com",
        "metadata": {"user_email": "int-sahni.manas@pocketfm.com"},
    }

    if ref_images:
        reference_list = []
        for ref in ref_images:
            try:
                img_bytes = Path(ref["path"]).read_bytes()
                b64_data = _b64.b64encode(img_bytes).decode("ascii")
                reference_list.append({
                    "referenceImage": {"bytesBase64Encoded": b64_data},
                    "referenceType": "SUBJECT_REFERENCE",
                    "subjectDescription": ref["name"],
                })
            except Exception as e:
                _log(f"  !!  Could not read ref image for {ref['name']}: {e}")
        if reference_list:
            body["referenceImages"] = reference_list

    for attempt in range(1, retries + 2):
        try:
            with httpx.Client(timeout=120) as hc:
                resp = hc.post(url, json=body, headers=headers)
            _s5_track(track_id, calls=1)
            if resp.status_code == 200:
                data = resp.json()
                items = data.get("data", [])
                if items and items[0].get("b64_json"):
                    img_bytes = _b64.b64decode(items[0]["b64_json"])
                    img_dir = job_dir / "images"
                    img_dir.mkdir(exist_ok=True)
                    fname = f"shot_{shot_num}.png"
                    (img_dir / fname).write_bytes(img_bytes)
                    _s5_track(track_id, images=1)
                    return f"/api/stage5/image/{job_id}/{fname}"
                elif items and items[0].get("url"):
                    with httpx.Client(timeout=60) as dl:
                        img_resp = dl.get(items[0]["url"])
                    if img_resp.status_code == 200:
                        img_dir = job_dir / "images"
                        img_dir.mkdir(exist_ok=True)
                        fname = f"shot_{shot_num}.png"
                        (img_dir / fname).write_bytes(img_resp.content)
                        _s5_track(track_id, images=1)
                        return f"/api/stage5/image/{job_id}/{fname}"
                _log(f"  x   No image data in response (attempt {attempt})")
            else:
                err = resp.text[:200]
                _log(f"  x   Attempt {attempt}: HTTP {resp.status_code}: {err}")
                if resp.status_code == 400 and ("SAFETY" in resp.text.upper() or "BLOCKED" in resp.text.upper()):
                    _log("  !!  Prompt blocked by safety filter — skipping")
                    return ""
        except Exception as e:
            _log(f"  x   Attempt {attempt}: {str(e)[:200]}")
        if attempt <= retries:
            wait = 30 * attempt
            _log(f"  -> Retrying in {wait}s...")
            time.sleep(wait)
    return ""


def _s5_write_output_excel(shots: list, input_path: Path, output_path: Path, job_dir: Path = None):
    from openpyxl import load_workbook
    wb = load_workbook(str(input_path))
    ws = wb.active

    preview_col = None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or "").strip().lower()
        if "preview_1" in h or "preview 1" in h:
            preview_col = c
            break
    if not preview_col:
        preview_col = ws.max_column + 1
        ws.cell(1, preview_col, "Preview_1")

    if job_dir:
        img_dir = job_dir / "images"
        if img_dir.exists():
            job_id = job_dir.name
            for shot in shots:
                if not shot.get("preview_1"):
                    sn = shot["num"]
                    img_path = img_dir / f"shot_{sn}.png"
                    if img_path.exists():
                        shot["preview_1"] = f"/api/stage5/image/{job_id}/{img_path.name}"

    for shot in shots:
        if shot.get("preview_1"):
            ws.cell(shot["row"], preview_col, shot["preview_1"])

    wb.save(str(output_path))


GSHEET_SA_PATH = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")
GSHEET_SHARE_EMAIL = os.environ.get("GSHEET_SHARE_EMAIL", "")

def _s5_create_google_sheet(shots: list, title: str) -> str:
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        _log("  !!  gspread/google-auth not installed — pip install gspread google-auth")
        return ""

    sa_path = Path(GSHEET_SA_PATH) if GSHEET_SA_PATH else ROOT / "google_service_account.json"
    if not sa_path.exists():
        _log(f"  !!  Service account not found at {sa_path} — skipping Google Sheets")
        return ""

    try:
        creds = Credentials.from_service_account_file(str(sa_path), scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ])
        gc = gspread.authorize(creds)

        sh = gc.create(title)
        ws = sh.sheet1

        headers = ["#", "Line", "Shot Size", "Shot Description", "Shot Detail", "Reference", "Preview_1"]
        data = [headers]
        for s in shots:
            data.append([
                s.get("num", ""),
                s.get("line", ""),
                s.get("shot_size", ""),
                s.get("shot_description", ""),
                s.get("shot_detail", ""),
                s.get("reference", ""),
                s.get("preview_1", ""),
            ])
        ws.update(range_name="A1", values=data)

        ws.format("A1:G1", {
            "backgroundColor": {"red": 0.12, "green": 0.22, "blue": 0.39},
            "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
            "horizontalAlignment": "CENTER",
        })
        ws.freeze(rows=1)

        share_email = GSHEET_SHARE_EMAIL or ""
        if share_email:
            sh.share(share_email, perm_type="user", role="writer")
            _log(f"  ok  Shared with {share_email}")
        sh.share("", perm_type="anyone", role="reader")

        _log(f"  ok  Google Sheet: {sh.url}")
        return sh.url

    except Exception as e:
        _log(f"  !!  Google Sheets error: {str(e)[:200]}")
        return ""


def _run_stage5_core(job_id: str, job_dir: Path, excel_path: Path, char_dir: Path, loc_dir: Path = None):
    """Core Stage 5 logic — callable from standalone route or pipeline."""
    _log("=" * 60)
    _log("  STAGE 5 — Image Generation (Imagen 4 Fast via MadEye)")
    _log("=" * 60)

    _log("  Reading shot breakdown Excel...")
    shots = _s5_read_excel(excel_path)
    _log(f"  ok  {len(shots)} shots found")

    jobs[job_id]["progress"]["total"] = len(shots)
    jobs[job_id]["s5_shots"] = shots
    jobs[job_id]["s5_prompts"] = {}
    jobs[job_id]["s5_excel_path"] = str(excel_path)
    jobs[job_id]["s5_char_dir"] = str(char_dir) if char_dir else ""

    ref_map = _s5_build_ref_map(char_dir) if char_dir else {}
    if ref_map:
        _log(f"  ok  {len(ref_map)} character reference(s): {', '.join(ref_map.keys())}")
    else:
        _log("  --  No character references — generating without identity anchoring")

    loc_map = {}
    if loc_dir and loc_dir.exists():
        loc_map = _s5_build_ref_map(loc_dir)
        if loc_map:
            _log(f"  ok  {len(loc_map)} location reference(s): {', '.join(loc_map.keys())}")
        else:
            _log("  --  No location references")

    # Prepare all tasks sequentially (ref matching needs last_char state)
    tasks = []
    last_char = ""
    for i, shot in enumerate(shots):
        sn = shot["num"] or (i + 1)
        desc = shot["shot_description"]
        if not desc:
            _log(f"  !!  Shot {sn}: no description — skipping")
            continue

        ref_names, ref_instruction, last_char, ref_imgs = _s5_select_refs(shot, ref_map, last_char)

        if loc_map:
            loc_names = _s5_match_characters(desc, shot.get("shot_detail", ""), loc_map)
            for ln in loc_names:
                entry = loc_map.get(ln, {})
                lpath = entry.get("path")
                if lpath and Path(lpath).exists():
                    ref_imgs.append({"name": f"Location: {ln}", "path": Path(lpath)})

        prompt = S5_PROMPT_PREFIX + ref_instruction + desc + " " + S5_NEGATIVE + S5_PROMPT_SUFFIX
        jobs[job_id]["s5_prompts"][sn] = prompt
        tasks.append((i, sn, shot, prompt, list(ref_imgs), list(ref_names)))

    # Execute image generation with key pool
    num_workers = _madeye_pool.key_count if _madeye_pool else 1
    _log(f"  >>  Generating {len(tasks)} shots with {num_workers} key(s)")
    generated = 0
    _gen_lock = threading.Lock()

    def _s5_gen_task(task_tuple):
        nonlocal generated
        idx, sn, shot, prompt, ref_imgs, ref_names = task_tuple
        ref_label = f" chars=[{', '.join(ref_names)}]" if ref_names else ""
        img_label = f" refs={len(ref_imgs)}" if ref_imgs else ""
        _log(f"  ... Shot {sn} ({idx+1}/{len(shots)}){ref_label}{img_label}")

        key = _madeye_pool.acquire() if _madeye_pool else None
        try:
            url = _s5_generate_one(prompt, job_id, sn, job_dir, ref_images=ref_imgs, api_key=key)
        finally:
            if _madeye_pool:
                _madeye_pool.release_notify()

        with _gen_lock:
            if url:
                shot["preview_1"] = url
                generated += 1
                _log(f"  ok  Shot {sn} saved")
            else:
                _log(f"  x   Shot {sn}: generation failed")
            jobs[job_id]["progress"]["completed"] = generated

    with ThreadPoolExecutor(max_workers=num_workers) as pool:
        list(pool.map(_s5_gen_task, tasks))

    _log(f"\n  Writing output Excel...")
    output_name = f"{excel_path.stem}_with_images.xlsx"
    output_path = job_dir / output_name
    _s5_write_output_excel(shots, excel_path, output_path, job_dir)

    jobs[job_id]["output_file"] = output_name
    jobs[job_id]["output_files"] = [output_name]

    _log(f"  Creating Google Sheet...")
    sheet_title = f"Stage 5 — {excel_path.stem}"
    sheet_url = _s5_create_google_sheet(shots, sheet_title)
    if sheet_url:
        jobs[job_id]["sheet_url"] = sheet_url

    tok = jobs[job_id]["tokens"]
    s5_cost = generated * S5_COST_PER_IMAGE
    jobs[job_id]["file_tokens"].append({
        "name": "Shot images",
        "images": generated,
        "cost": round(s5_cost, 4),
    })
    _log(f"\n{'=' * 60}")
    _log(f"  STAGE 5 COMPLETE — {generated}/{len(shots)} images generated")
    _log(f"  API calls: {tok['calls']}  |  Images: {generated}  |  Est. cost: ${s5_cost:.2f}")
    _log(f"  Output: {output_name}")
    if sheet_url:
        _log(f"  Google Sheet: {sheet_url}")
    _log(f"{'=' * 60}\n")


def run_stage5_pipeline(job_id: str, job_dir: Path, excel_path: Path, char_dir: Path, loc_dir: Path = None):
    """Standalone entry point (called from API route)."""
    log_queue = jobs[job_id]["queue"]
    _set_job_context(job_id, log_queue)
    try:
        _run_stage5_core(job_id, job_dir, excel_path, char_dir, loc_dir)
        jobs[job_id]["status"] = "done"
    except Exception as e:
        tb = traceback.format_exc()
        _log(f"\n[ERROR] {e}\n{tb}")
        jobs[job_id]["status"] = "failed"
    finally:
        log_queue.put(None)


# ── Stage 6 — Video Generation ───────────────────────────────────────────────

VID_W, VID_H       = 1080, 1080
VID_FONT_SIZE       = 105
VID_CHARS_PER_LINE  = 22
VID_FPS             = 24
VID_CRF             = 18
VID_AUDIO_BITRATE   = "192k"

S6_WORKSPACE = Path("D:/video_gen_workspace")
S6_WORKSPACE.mkdir(parents=True, exist_ok=True)

SUB_COLORS = {
    "yellow": "&H0000C4FF",
    "white":  "&H00FFFFFF",
}

def _ass_style(color_name="yellow"):
    c = SUB_COLORS.get(color_name, SUB_COLORS["yellow"])
    return (f"Style: Default,Poppins,{VID_FONT_SIZE},"
            f"{c},&H0000FFFF,&H00000000,&H00000000,"
            "-1,0,0,0,100,100,0,0,1,4,2,2,40,40,60,1")


def _s6_read_excel(excel_path: Path) -> list:
    import openpyxl
    wb = openpyxl.load_workbook(str(excel_path))
    ws = wb["Shot Breakdown"] if "Shot Breakdown" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    _log(f"  Sheet '{ws.title}' | Columns: {[h for h in headers if h]}")

    def _find(names):
        for name in names:
            for i, h in enumerate(headers):
                if h and name.lower() in str(h).lower():
                    return i + 1
        return None

    num_col  = _find(["#", "num", "shot"])
    line_col = _find(["line", "narration", "dialogue", "text"])
    url_col  = _find(["preview_1", "preview", "generated_url", "url", "image_url", "image"])

    _log(f"  Matched cols -> #:{num_col}  line:{line_col}  url:{url_col}")

    if not line_col:
        raise RuntimeError("No narration/line column in Excel. Headers: " + str([h for h in headers if h]))
    if not url_col:
        raise RuntimeError("No image URL column in Excel. Headers: " + str([h for h in headers if h]))

    shots = []
    skipped = 0
    for r in range(2, ws.max_row + 1):
        raw_num = ws.cell(r, num_col).value if num_col else r - 1
        line    = ws.cell(r, line_col).value
        url     = ws.cell(r, url_col).value
        if not line or not url:
            skipped += 1
            continue
        try:
            seq = int(float(str(raw_num))) if raw_num else len(shots) + 1
        except (ValueError, TypeError):
            seq = len(shots) + 1
        shots.append({
            "num": seq, "idx": len(shots) + 1,
            "line": str(line).strip(), "url": str(url).strip(),
        })
    if skipped:
        _log(f"  !!  {skipped} rows skipped (missing line or URL)")
    if not shots:
        raise RuntimeError(f"0 shots found! Sheet '{ws.title}' has {ws.max_row - 1} data rows but none had both a line and URL. Headers: {[h for h in headers if h]}")
    return shots


def _s6_download_images(shots: list, img_dir: Path):
    import ssl, urllib.request
    img_dir.mkdir(exist_ok=True)
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    def _dl(shot):
        dest = img_dir / f"{shot['idx']:04d}.png"
        shot["img_path"] = str(dest)
        if dest.exists() and dest.stat().st_size > 1000:
            return "cached"
        url = shot["url"]
        if url.startswith("file:///"):
            import shutil
            src = url[8:].replace("/", os.sep)
            try:
                shutil.copy2(src, dest)
                return "ok"
            except Exception:
                shot["img_path"] = None
                return "fail"
        for attempt in range(3):
            try:
                req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
                with urllib.request.urlopen(req, context=ctx, timeout=30) as resp:
                    dest.write_bytes(resp.read())
                return "ok"
            except Exception:
                if attempt < 2:
                    time.sleep(2 * (attempt + 1))
        shot["img_path"] = None
        return "fail"

    ok = fail = cached = 0
    with ThreadPoolExecutor(max_workers=10) as pool:
        futs = {pool.submit(_dl, s): s for s in shots}
        for fut in as_completed(futs):
            r = fut.result()
            if r == "ok": ok += 1
            elif r == "fail": fail += 1
            else: cached += 1
            done = ok + fail + cached
            if done % 25 == 0:
                _log(f"  ... {done}/{len(shots)}")
    _log(f"  ok  {ok} downloaded, {cached} cached, {fail} failed")

    if fail > 0:
        failed_shots = [s for s in shots if not s.get("img_path")]
        _log(f"  Retrying {len(failed_shots)} failed downloads (round 2)...")
        time.sleep(5)
        recovered = 0
        with ThreadPoolExecutor(max_workers=5) as pool:
            futs = {pool.submit(_dl, s): s for s in failed_shots}
            for fut in as_completed(futs):
                if fut.result() == "ok":
                    recovered += 1
        still_failed = sum(1 for s in shots if not s.get("img_path"))
        _log(f"  ok  Recovered {recovered}, still failed: {still_failed}")
        if still_failed > 0:
            _log(f"  !!  {still_failed} images missing — video will skip those shots")


def _s6_resize_images(shots: list, img_dir: Path):
    from PIL import Image
    for shot in shots:
        src = shot.get("img_path")
        if not src or not Path(src).exists():
            shot["sq_path"] = None
            continue
        out = img_dir / f"{shot['idx']:04d}_r.png"
        shot["sq_path"] = str(out)
        if out.exists():
            continue
        try:
            img = Image.open(src).convert("RGB")
            ow, oh = img.size
            scale = max(VID_W / ow, VID_H / oh)
            nw, nh = int(ow * scale), int(oh * scale)
            img = img.resize((nw, nh), Image.LANCZOS)
            left, top = (nw - VID_W) // 2, (nh - VID_H) // 2
            img.crop((left, top, left + VID_W, top + VID_H)).save(str(out), quality=95)
        except Exception as e:
            _log(f"  x  Resize shot {shot['idx']}: {e}")
            shot["sq_path"] = None
    valid = sum(1 for s in shots if s.get("sq_path") and Path(s["sq_path"]).exists())
    _log(f"  ok  {valid}/{len(shots)} resized to {VID_W}x{VID_H}")


def _s6_align_audio(shots: list, audio_path: Path, whisper_model: str) -> float:
    from difflib import SequenceMatcher
    from faster_whisper import WhisperModel

    _log(f"  Loading Whisper '{whisper_model}'...")
    model = WhisperModel(whisper_model, device="cpu", compute_type="int8")
    segs, info = model.transcribe(str(audio_path), word_timestamps=True, language="en")

    words = []
    for seg in segs:
        if seg.words:
            for w in seg.words:
                words.append({"word": w.word.strip(), "start": w.start, "end": w.end})
    duration = info.duration
    _log(f"  ok  {len(words)} words, {duration:.1f}s")

    def _clean(t):
        return re.sub(r'[^\w\s]', '', t.lower()).split()

    wc = [re.sub(r'[^\w]', '', w["word"].lower()) for w in words]

    if not words:
        _log("  !!  Whisper returned 0 words — assigning equal durations")
        per = duration / max(len(shots), 1)
        for i, shot in enumerate(shots):
            shot["start"] = i * per
            shot["end"] = (i + 1) * per
        shots[-1]["end"] = duration
        _log(f"  ok  Aligned {len(shots)} shots (fallback) | 0.00s - {duration:.2f}s")
        return duration

    # Build flat list of all script words with shot ownership
    script_words = []
    shot_ranges = []
    for si, shot in enumerate(shots):
        sw = _clean(shot["line"])
        start_idx = len(script_words)
        script_words.extend(sw)
        shot_ranges.append((start_idx, len(script_words)))

    _log(f"  Global alignment: {len(script_words)} script words vs {len(wc)} whisper words...")

    # Global alignment — map every script word to a Whisper word index
    sm = SequenceMatcher(None, script_words, wc, autojunk=False)
    blocks = sm.get_matching_blocks()

    sw_to_wc = {}
    for a, b, size in blocks:
        for k in range(size):
            sw_to_wc[a + k] = b + k

    mapped = len(sw_to_wc)
    _log(f"  Mapped {mapped}/{len(script_words)} script words ({mapped*100//max(len(script_words),1)}%)")

    # Interpolate unmapped words between mapped neighbors
    all_sw = sorted(sw_to_wc.keys())
    for i in range(len(script_words)):
        if i in sw_to_wc:
            continue
        left_idx = None
        right_idx = None
        for k in reversed(all_sw):
            if k < i:
                left_idx = k
                break
        for k in all_sw:
            if k > i:
                right_idx = k
                break
        if left_idx is not None and right_idx is not None:
            left_wc = sw_to_wc[left_idx]
            right_wc = sw_to_wc[right_idx]
            frac = (i - left_idx) / (right_idx - left_idx)
            sw_to_wc[i] = int(left_wc + frac * (right_wc - left_wc))
        elif left_idx is not None:
            sw_to_wc[i] = min(sw_to_wc[left_idx] + (i - left_idx), len(words) - 1)
        elif right_idx is not None:
            sw_to_wc[i] = max(sw_to_wc[right_idx] - (right_idx - i), 0)

    # Assign raw timestamps to each shot
    raw_starts = []
    raw_ends = []
    for si, shot in enumerate(shots):
        sr_start, sr_end = shot_ranges[si]
        if sr_start >= sr_end:
            raw_starts.append(None)
            raw_ends.append(None)
            continue
        first_wc = max(0, min(sw_to_wc.get(sr_start, 0), len(words) - 1))
        last_wc = max(0, min(sw_to_wc.get(sr_end - 1, first_wc), len(words) - 1))
        if last_wc < first_wc:
            last_wc = first_wc
        raw_starts.append(words[first_wc]["start"])
        raw_ends.append(words[last_wc]["end"])

    # Enforce monotonicity — starts must be strictly increasing
    for i in range(1, len(raw_starts)):
        if raw_starts[i] is None:
            raw_starts[i] = raw_starts[i - 1] if raw_starts[i - 1] is not None else 0.0
        if raw_starts[i - 1] is not None and raw_starts[i] <= raw_starts[i - 1]:
            raw_starts[i] = raw_starts[i - 1] + 0.01

    # Detect tail region where timestamps bunch up (short lines at end)
    tail_start = None
    for i in range(len(shots) - 1, 0, -1):
        if raw_starts[i] is not None and raw_starts[i - 1] is not None:
            if raw_starts[i] - raw_starts[i - 1] < 0.05:
                tail_start = i
            else:
                break
    if tail_start is not None and tail_start < len(shots) - 1:
        prev_time = raw_starts[tail_start - 1] if tail_start > 0 else 0.0
        remaining = duration - prev_time
        tail_count = len(shots) - tail_start + 1
        per_shot = remaining / tail_count
        _log(f"  Redistributing tail shots {tail_start}–{len(shots)} ({tail_count} shots, {remaining:.1f}s)")
        for i in range(tail_start - 1, len(shots)):
            offset = i - (tail_start - 1)
            raw_starts[i] = prev_time + offset * per_shot

    # Assign to shots
    for i, shot in enumerate(shots):
        shot["start"] = raw_starts[i] if raw_starts[i] is not None else 0.0

    # Enforce continuity: each shot ends where the next begins
    for i in range(len(shots) - 1):
        shots[i]["end"] = shots[i + 1]["start"]
    shots[0]["start"] = 0.0
    shots[-1]["end"] = duration

    # Final pass — fix any zero/negative duration shots
    for i, shot in enumerate(shots):
        if shot["end"] - shot["start"] <= 0.01:
            if i < len(shots) - 1:
                gap = shots[i + 1]["start"] - shot["start"]
                if gap > 0.01:
                    shot["end"] = shots[i + 1]["start"]
                else:
                    shot["end"] = shot["start"] + 0.3
                    if i < len(shots) - 1:
                        shots[i + 1]["start"] = shot["end"]
            else:
                shot["end"] = duration

    _log(f"  ok  Aligned {len(shots)} shots | 0.00s - {shots[-1]['end']:.2f}s")
    return duration


def _s6_ass_time(sec: float) -> str:
    h = int(sec // 3600)
    m = int((sec % 3600) // 60)
    s = sec % 60
    return f"{h}:{m:02d}:{s:05.2f}"


def _s6_wrap(text: str) -> list:
    mx = VID_CHARS_PER_LINE
    if len(text) <= mx:
        return [text]
    words = text.split()
    lines, cur = [], ""
    for w in words:
        test = (cur + " " + w).strip() if cur else w
        if len(test) <= mx:
            cur = test
        else:
            if cur:
                lines.append(cur)
            cur = w
    if cur:
        lines.append(cur)
    chunks = []
    for i in range(0, len(lines), 2):
        if i + 1 < len(lines):
            chunks.append(lines[i] + "\\N" + lines[i + 1])
        else:
            chunks.append(lines[i])
    return chunks or [text[:mx * 2]]


def _s6_generate_subtitles(shots: list, ass_path: Path, sub_color="yellow"):
    events = []
    for shot in shots:
        chunks = _s6_wrap(shot["line"])
        dur = shot["end"] - shot["start"]
        if len(chunks) == 1:
            events.append((shot["start"], shot["end"], chunks[0]))
        else:
            total_c = sum(len(c.replace("\\N", " ")) for c in chunks)
            t = shot["start"]
            for ci, ch in enumerate(chunks):
                cc = len(ch.replace("\\N", " "))
                cd = dur * (cc / total_c) if total_c > 0 else dur / len(chunks)
                ce = t + cd if ci < len(chunks) - 1 else shot["end"]
                events.append((t, ce, ch))
                t = ce
    with open(ass_path, "w", encoding="utf-8-sig") as f:
        f.write("[Script Info]\nTitle: Subtitles\nScriptType: v4.00+\n")
        f.write(f"PlayResX: {VID_W}\nPlayResY: {VID_H}\n")
        f.write("WrapStyle: 2\nScaledBorderAndShadow: yes\n\n")
        f.write("[V4+ Styles]\n")
        f.write("Format: Name, Fontname, Fontsize, PrimaryColour, SecondaryColour, "
                "OutlineColour, BackColour, Bold, Italic, Underline, StrikeOut, "
                "ScaleX, ScaleY, Spacing, Angle, BorderStyle, Outline, Shadow, "
                "Alignment, MarginL, MarginR, MarginV, Encoding\n")
        f.write(f"{_ass_style(sub_color)}\n\n")
        f.write("[Events]\n")
        f.write("Format: Layer, Start, End, Style, Name, MarginL, MarginR, MarginV, Effect, Text\n")
        for st, en, tx in events:
            f.write(f"Dialogue: 0,{_s6_ass_time(st)},{_s6_ass_time(en)},Default,,0,0,0,,{tx}\n")
    _log(f"  ok  {len(events)} subtitle events")


def _s6_ensure_font(font_dir: Path) -> str:
    import ssl, urllib.request
    font_dir.mkdir(exist_ok=True)
    fp = font_dir / "Poppins-Bold.ttf"
    if fp.exists():
        return str(font_dir)
    cached = S6_WORKSPACE / "Poppins-Bold.ttf"
    if cached.exists():
        shutil.copy2(str(cached), str(fp))
        _log("  ok  Font loaded from cache")
        return str(font_dir)
    _log("  Downloading Poppins-Bold.ttf...")
    try:
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        req = urllib.request.Request(
            "https://github.com/google/fonts/raw/main/ofl/poppins/Poppins-Bold.ttf",
            headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, context=ctx, timeout=30) as resp:
            data = resp.read()
            fp.write_bytes(data)
            cached.write_bytes(data)
        _log("  ok  Font downloaded + cached")
    except Exception as e:
        _log(f"  !!  Font download failed: {e}")
    return str(font_dir)


def _s6_render_video(shots, concat_path: Path, audio_path: Path,
                     ass_path: Path, font_dir: str, output_path: Path,
                     subs: bool) -> float:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(concat_path, "w") as f:
        last = None
        for shot in shots:
            sq = shot.get("sq_path")
            if not sq or not Path(sq).exists(): continue
            dur = max(shot["end"] - shot["start"], 0.1)
            sq_fwd = sq.replace("\\", "/")
            f.write(f"file '{sq_fwd}'\nduration {dur:.6f}\n")
            last = sq_fwd
        if last:
            f.write(f"file '{last}'\n")
    if not subs:
        _log("  Rendering video (no subtitles)...")
        cmd = ["ffmpeg", "-y", "-f", "concat", "-safe", "0", "-i", str(concat_path),
               "-i", str(audio_path), "-r", str(VID_FPS),
               "-c:v", "libx264", "-pix_fmt", "yuv420p",
               "-preset", "medium", "-crf", str(VID_CRF),
               "-c:a", "aac", "-b:a", VID_AUDIO_BITRATE,
               "-shortest", "-movflags", "+faststart", str(output_path)]
        r = subprocess.run(cmd, capture_output=True, text=True)
        if r.returncode != 0:
            raise RuntimeError(f"ffmpeg failed:\n{r.stderr[-800:]}")
    else:
        nosub_path = output_path.parent / (output_path.stem + "_nosubs.mp4")
        _log("  Building base video (no subs)...")
        cmd1 = ["ffmpeg", "-y", "-f", "concat", "-safe", "0", "-i", str(concat_path),
                "-i", str(audio_path), "-r", str(VID_FPS),
                "-c:v", "libx264", "-pix_fmt", "yuv420p",
                "-preset", "ultrafast", "-crf", str(VID_CRF),
                "-c:a", "aac", "-b:a", VID_AUDIO_BITRATE,
                "-shortest", "-movflags", "+faststart", str(nosub_path)]
        r = subprocess.run(cmd1, capture_output=True, text=True)
        if r.returncode != 0:
            raise RuntimeError(f"ffmpeg base pass failed:\n{r.stderr[-800:]}")
        nosub_mb = nosub_path.stat().st_size / (1024 * 1024)
        _log(f"  ok  Base video ready: {nosub_path.name} ({nosub_mb:.1f} MB) — downloadable now")
        jid = getattr(_job_local, "job_id", None)
        if jid and jid in jobs:
            jobs[jid].setdefault("output_files", []).append(nosub_path.name)
        _log("  Burning subtitles (this takes a while)...")
        ass_esc = str(ass_path).replace("\\", "/").replace(":", "\\:")
        fd_esc = font_dir.replace("\\", "/").replace(":", "\\:")
        vf_attempts = [
            f"ass='{ass_esc}':fontsdir='{fd_esc}'",
            f"ass={ass_esc}:fontsdir={fd_esc}",
            f"subtitles={ass_esc}:fontsdir={fd_esc}",
        ]
        burned = False
        for i, vf in enumerate(vf_attempts):
            _log(f"  Burning subtitles (attempt {i+1}/{len(vf_attempts)})...")
            cmd2 = ["ffmpeg", "-y", "-i", str(nosub_path), "-vf", vf,
                    "-c:v", "libx264", "-pix_fmt", "yuv420p",
                    "-preset", "medium", "-crf", str(VID_CRF),
                    "-c:a", "copy", "-movflags", "+faststart", str(output_path)]
            r = subprocess.run(cmd2, capture_output=True, text=True)
            if r.returncode == 0:
                _log("  ok  Subtitles burned"); burned = True; break
            _log(f"  x  Attempt {i+1} failed: {r.stderr[-200:].strip()}")
        if not burned:
            _log("  !!  Subtitle burn failed - delivering without subtitles")
            shutil.copy2(str(nosub_path), str(output_path))
    probe = subprocess.run(
        ["ffprobe", "-v", "quiet", "-show_entries", "format=duration",
         "-of", "csv=p=0", str(output_path)], capture_output=True, text=True)
    try: return float(probe.stdout.strip())
    except ValueError: return 0.0


def _run_s6_one(job_id, pair_dir, excel_path, audio_path, whisper_model, subs, pair_idx, step_offset, sub_color="yellow"):
    steps_per = 6 if subs else 4
    lp = f"[{pair_idx}] " if pair_idx > 0 else ""
    def _step(n, label):
        _log(f"\n  -- {lp}Step {n}/{steps_per}: {label} --")
        jobs[job_id]["progress"]["stage"] = f"{lp}Step {n}/{steps_per}: {label}"
        jobs[job_id]["progress"]["completed"] = step_offset + n - 1
    img_dir = pair_dir / "images"; font_dir = pair_dir / "fonts"
    ass_path = pair_dir / "subtitles.ass"; concat_path = pair_dir / "concat.txt"
    _step(1, "Reading Excel")
    shots = _s6_read_excel(excel_path)
    _log(f"  ok  {len(shots)} shots")
    _step(2, "Downloading images")
    _s6_download_images(shots, img_dir)
    _step(3, "Resizing images to 1080x1080")
    _s6_resize_images(shots, img_dir)
    if subs:
        _step(4, "Aligning audio (Whisper)")
        _s6_align_audio(shots, audio_path, whisper_model)
        _step(5, "Generating subtitles")
        font_dir_str = _s6_ensure_font(font_dir)
        _s6_generate_subtitles(shots, ass_path, sub_color)
        _step(6, "Rendering video")
        safe = re.sub(r'[^a-zA-Z0-9_\-]', '_', excel_path.stem)
        out = pair_dir / f"{safe}_1x1.mp4"
        _s6_render_video(shots, concat_path, audio_path, ass_path, font_dir_str, out, True)
    else:
        _step(4, "Rendering video")
        probe = subprocess.run(
            ["ffprobe", "-v", "quiet", "-show_entries", "format=duration",
             "-of", "csv=p=0", str(audio_path)], capture_output=True, text=True)
        total_dur = float(probe.stdout.strip()) if probe.stdout.strip() else 60.0
        valid = [s for s in shots if s.get("sq_path") and Path(s["sq_path"]).exists()]
        per = total_dur / max(len(valid), 1)
        for s in shots: s["start"] = 0.0; s["end"] = per
        safe = re.sub(r'[^a-zA-Z0-9_\-]', '_', excel_path.stem)
        out = pair_dir / f"{safe}_1x1.mp4"
        _s6_render_video(shots, concat_path, audio_path, ass_path, "", out, False)
    jobs[job_id]["progress"]["completed"] = step_offset + steps_per
    if not out.exists():
        raise RuntimeError(f"Output file was not created: {out}")
    size_mb = out.stat().st_size / (1024 * 1024)
    _log(f"  ok  {out.name}  ({size_mb:.1f} MB)")
    if img_dir.exists(): shutil.rmtree(img_dir, ignore_errors=True)
    if font_dir.exists(): shutil.rmtree(font_dir, ignore_errors=True)
    for tmp in [ass_path, concat_path]:
        if tmp.exists(): tmp.unlink(missing_ok=True)
    _log("  ok  Cleaned up temp files")
    return out


def run_stage6_batch(job_id, job_dir, pairs, whisper_model, subs, sub_color="yellow"):
    log_queue = jobs[job_id]["queue"]
    _set_job_context(job_id, log_queue)
    steps_per = 6 if subs else 4
    try:
        n = len(pairs)
        _log(f"\n{'=' * 60}")
        _log("  STAGE 4 - Video Generation")
        _log(f"  {n} pair(s)  |  Subtitles: {'Yes' if subs else 'No'}  |  Whisper: {whisper_model}")
        _log(f"{'=' * 60}")
        outputs = []
        for i, (excel_path, audio_path) in enumerate(pairs):
            pair_dir = excel_path.parent
            _log(f"\n{'~' * 60}")
            _log(f"  Pair {i+1}/{n}: {excel_path.name}  +  {audio_path.name}")
            _log(f"{'~' * 60}")
            out = _run_s6_one(job_id, pair_dir, excel_path, audio_path,
                              whisper_model, subs, i + 1 if n > 1 else 0, i * steps_per, sub_color)
            outputs.append(out.name)
        jobs[job_id]["progress"]["completed"] = steps_per * n
        jobs[job_id]["output_file"] = outputs[0]
        jobs[job_id]["output_files"] = outputs
        _log(f"\n{'=' * 60}")
        _log(f"  STAGE 4 COMPLETE - {len(outputs)} video(s)")
        for name in outputs: _log(f"    -> {name}")
        _log(f"{'=' * 60}\n")
        jobs[job_id]["status"] = "done"
    except Exception as e:
        tb = traceback.format_exc()
        _log(f"\n[ERROR] {e}\n{tb}")
        jobs[job_id]["status"] = "failed"
    finally:
        log_queue.put(None)


# ── Pipeline orchestrator ─────────────────────────────────────────────────────

def run_pipeline(job_id: str, job_dir: Path, scripts: list, show_name: str, workers: int):
    """Run Stage 1 -> 2 -> (3 || 4) -> 5 as an automated pipeline."""
    log_queue = jobs[job_id]["queue"]
    _set_job_context(job_id, log_queue)

    try:
        _log(f"\n{'=' * 60}")
        _log("  PIPELINE — Automated Stage 1 → 2 → 3+4 → 5")
        _log(f"  Show: {show_name}  |  Episodes: {len(scripts)}")
        _log(f"{'=' * 60}\n")

        jobs[job_id]["progress"]["pipeline_stage"] = "stage1"
        jobs[job_id]["progress"]["started_at"] = time.time()

        # ── Stage 1: Reference Files ─────────────────────────────────────────
        _log("\n  ══ PIPELINE STAGE 1: Reference File Generation ══")
        scripts_dir = job_dir / "episodic scripts"
        scripts_dir.mkdir(parents=True, exist_ok=True)
        for s in scripts:
            save_name = Path(s["filename"]).stem + ".md"
            (scripts_dir / save_name).write_text(s["text"], encoding="utf-8")

        _run_stage1_core(job_id, job_dir, "full", workers, None)

        # Read Stage 1 outputs
        show_files_dir = job_dir / "show level files"
        ep_details_dir = job_dir / "episode details"

        ref_files_global = {}
        if show_files_dir.exists():
            for key in ["show_tone_bible", "character_canvas", "location_reference"]:
                matches = list(show_files_dir.glob(f"{key}*.md"))
                if matches:
                    ref_files_global[key] = matches[-1].read_text(encoding="utf-8")

        detail_files = []
        if ep_details_dir.exists():
            for f in sorted(ep_details_dir.glob("*.md")):
                detail_files.append((f.name, f.read_text(encoding="utf-8")))

        _log(f"  Stage 1 outputs: {len(ref_files_global)} ref files, {len(detail_files)} episode details")

        # ── Stage 2: Shot Breakdown ──────────────────────────────────────────
        jobs[job_id]["progress"]["pipeline_stage"] = "stage2"
        _log("\n  ══ PIPELINE STAGE 2: Shot Breakdown ══")

        episodes = [{"name": s["name"], "script_text": s["text"]} for s in scripts]
        _run_stage2_core(job_id, job_dir, episodes, ref_files_global, detail_files, workers=workers)

        # Collect Stage 2 Excel outputs
        excel_files = sorted(job_dir.glob("*_breakdown.xlsx"))
        _log(f"  Stage 2 outputs: {len(excel_files)} Excel file(s)")

        # ── Stage 3 + Stage 4 in parallel ────────────────────────────────────
        jobs[job_id]["progress"]["pipeline_stage"] = "stage3+4"
        _log("\n  ══ PIPELINE STAGES 3+4: Audit + Reference Images (parallel) ══")

        s3_error = [None]
        s4_error = [None]

        def _pipeline_stage3():
            try:
                _set_job_context(job_id, log_queue)
                for ep_excel in excel_files:
                    ep_name = ep_excel.stem.replace("_breakdown", "")
                    script_match = next((s for s in scripts if re.sub(r'[^a-zA-Z0-9_\-]', '_', s["name"]) == ep_name), None)
                    if not script_match:
                        script_match = scripts[0] if scripts else None
                    if script_match:
                        script_md_path = job_dir / f"{ep_name}.md"
                        script_md_path.write_text(script_match["text"], encoding="utf-8")
                        _log(f"  [S3] Auditing {ep_excel.name}...")
                        _run_s3_one(job_id, job_dir, ep_excel, script_match["text"])
            except Exception as e:
                s3_error[0] = e
                _log(f"  [S3] ERROR: {e}")

        def _pipeline_stage4():
            try:
                _set_job_context(job_id, log_queue)
                canvas_text = ref_files_global.get("character_canvas", "")
                location_text = ref_files_global.get("location_reference", "")
                if canvas_text or location_text:
                    items = []
                    if canvas_text:
                        canvas_matches = list(show_files_dir.glob("character_canvas*.md"))
                        if canvas_matches:
                            items.extend(_s4_parse_md(canvas_matches[-1]))
                    if location_text:
                        loc_matches = list(show_files_dir.glob("location_reference*.md"))
                        if loc_matches:
                            items.extend(_s4_parse_md(loc_matches[-1]))
                    if items:
                        _log(f"  [S4] Generating {len(items)} reference images...")
                        _run_stage4_core(job_id, job_dir, items)
                    else:
                        _log("  [S4] No items parsed from reference files — skipping")
                else:
                    _log("  [S4] No character canvas or location reference — skipping Stage 4")
            except Exception as e:
                s4_error[0] = e
                _log(f"  [S4] ERROR: {e}")

        t3 = threading.Thread(target=_pipeline_stage3)
        t4 = threading.Thread(target=_pipeline_stage4)
        t3.start()
        t4.start()
        t3.join()
        t4.join()

        if s3_error[0]:
            _log(f"  !! Stage 3 failed: {s3_error[0]}")
        if s4_error[0]:
            _log(f"  !! Stage 4 failed: {s4_error[0]}")

        # ── Stage 5: Shot Image Generation ───────────────────────────────────
        jobs[job_id]["progress"]["pipeline_stage"] = "stage5"
        _log("\n  ══ PIPELINE STAGE 5: Shot Image Generation ══")

        # Find audited Excels (Stage 3 output) or fall back to Stage 2 output
        audited_excels = sorted(job_dir.glob("*_audited.xlsx"))
        s5_excels = audited_excels if audited_excels else excel_files

        # Find Stage 4 reference image dir
        s4_images_dir = job_dir / "images"
        char_dir = s4_images_dir if s4_images_dir.exists() else None
        loc_dir = char_dir  # Same dir for both in Stage 4

        for ep_excel in s5_excels:
            _log(f"  Running Stage 5 for {ep_excel.name}...")
            _run_stage5_core(job_id, job_dir, ep_excel, char_dir, loc_dir)

        # ── Pipeline complete ────────────────────────────────────────────────
        jobs[job_id]["progress"]["pipeline_stage"] = "complete"
        tok = jobs[job_id]["tokens"]
        total_cost = sum(ft.get("cost", 0) for ft in jobs[job_id].get("file_tokens", []))

        _log(f"\n{'=' * 60}")
        _log("  PIPELINE COMPLETE")
        _log(f"  Tokens: {tok['input']:,} input / {tok['output']:,} output / {tok['calls']} calls")
        _log(f"  Cost: ${total_cost:.4f}")
        _log(f"{'=' * 60}\n")

        jobs[job_id]["status"] = "done"

    except Exception as e:
        tb = traceback.format_exc()
        _log(f"\n[PIPELINE ERROR] {e}\n{tb}")
        jobs[job_id]["status"] = "failed"
    finally:
        log_queue.put(None)


# ── Flask helpers ─────────────────────────────────────────────────────────────

def _new_job(job_dir: Path) -> tuple:
    job_id    = str(uuid.uuid4())
    log_queue = queue.Queue()
    jobs[job_id] = {
        "status":      "running",
        "queue":       log_queue,
        "job_dir":     job_dir,
        "tokens":      {"input": 0, "output": 0, "calls": 0},
        "tokens_lock": threading.Lock(),
        "output_file": None,
        "file_tokens": [],
    }
    return job_id, log_queue


def _collect_stage1_files(job_dir: Path) -> list:
    result = []
    for subfolder in ["show level files", "episode details"]:
        d = job_dir / subfolder
        if d.exists():
            for f in sorted(d.glob("*.md")):
                result.append({"name": f.name, "subfolder": subfolder})
    return result


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/config")
def api_config():
    return jsonify({
        "model": ARGUS_MODEL,
        "llm_input_cost": LLM_INPUT_COST_PER_MTOK,
        "llm_output_cost": LLM_OUTPUT_COST_PER_MTOK,
    })


# Stage 1 ──

@app.route("/api/stage1/run", methods=["POST"])
def api_stage1_run():
    mode    = request.form.get("mode", "full")
    workers = int(request.form.get("workers", "3") or "3")
    episode = request.form.get("episode", "").strip()
    files   = request.files.getlist("scripts")

    if not ARGUS_API_KEY or not ARGUS_BASE_URL:
        return jsonify({"error": "ARGUS_API_KEY / ARGUS_BASE_URL not set in .env"}), 500
    if not files or all(f.filename == "" for f in files):
        return jsonify({"error": "No script files uploaded"}), 400
    if mode == "episode" and not episode:
        return jsonify({"error": "Episode filename required"}), 400

    job_id  = str(uuid.uuid4())
    job_dir = WORKSPACE / job_id
    scripts_dir = job_dir / "episodic scripts"
    scripts_dir.mkdir(parents=True, exist_ok=True)

    for f in files:
        if f.filename:
            f.save(str(scripts_dir / Path(f.filename).name))

    log_queue = queue.Queue()
    jobs[job_id] = {
        "status":       "running",
        "queue":        log_queue,
        "job_dir":      job_dir,
        "tokens":       {"input": 0, "output": 0, "calls": 0},
        "tokens_lock":  threading.Lock(),
        "output_file":  None,
        "output_files": [],
        "file_tokens":  [],
        "progress":     {"total": 0, "completed": 0, "stage": "", "started_at": time.time()},
    }

    threading.Thread(
        target=run_stage1_pipeline,
        args=(job_id, job_dir, mode, workers, episode),
        daemon=True,
    ).start()

    return jsonify({"job_id": job_id})


@app.route("/api/stage1/download/<job_id>/<path:filename>")
def api_stage1_download(job_id: str, filename: str):
    if job_id not in jobs:
        return jsonify({"error": "not found"}), 404
    job_dir = jobs[job_id]["job_dir"]
    for subfolder in ["show level files", "episode details"]:
        candidate = job_dir / subfolder / filename
        if candidate.exists():
            return send_file(str(candidate), as_attachment=True, download_name=filename)
    return jsonify({"error": "file not found"}), 404


# Stage 2 ──

@app.route("/api/stage2/run", methods=["POST"])
def api_stage2_run():
    if not ARGUS_API_KEY or not ARGUS_BASE_URL:
        return jsonify({"error": "ARGUS_API_KEY / ARGUS_BASE_URL not set in .env"}), 500

    episode_name_input = request.form.get("episode_name", "").strip()
    script_text        = request.form.get("script_text", "").strip()

    try:
        # Build episodes list — one entry per script file, or one from pasted text
        episodes = []
        if script_text:
            episodes.append({
                "name":        episode_name_input or "episode",
                "script_text": script_text,
            })
        else:
            for f in request.files.getlist("script"):
                if not f.filename:
                    continue
                stem = Path(f.filename).stem.strip()
                episodes.append({
                    "name":        stem or "episode",
                    "script_text": _read_upload_as_text(f),
                })

        if not episodes:
            return jsonify({"error": "No script provided"}), 400

        # Shared reference files (concatenated if multiple per type)
        ref_files_global = {}
        for key in ["show_tone_bible", "character_canvas", "location_reference"]:
            parts = [_read_upload_as_text(f) for f in request.files.getlist(f"ref_{key}") if f.filename]
            if parts:
                ref_files_global[key] = "\n\n---\n\n".join(parts)

        # Episode details — keep as list so we can match per episode by filename
        detail_files = [
            (f.filename, _read_upload_as_text(f))
            for f in request.files.getlist("ref_episode_detail") if f.filename
        ]
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 400

    job_id  = str(uuid.uuid4())
    job_dir = WORKSPACE / job_id
    job_dir.mkdir(parents=True, exist_ok=True)

    log_queue = queue.Queue()
    jobs[job_id] = {
        "status":       "running",
        "queue":        log_queue,
        "job_dir":      job_dir,
        "tokens":       {"input": 0, "output": 0, "calls": 0},
        "tokens_lock":  threading.Lock(),
        "output_file":  None,
        "output_files": [],
        "file_tokens":  [],
        "progress":     {"total": len(episodes), "completed": 0, "stage": "Starting", "started_at": time.time()},
    }

    threading.Thread(
        target=run_stage2_pipeline,
        args=(job_id, job_dir, episodes, ref_files_global, detail_files),
        daemon=True,
    ).start()

    return jsonify({"job_id": job_id, "episode_count": len(episodes)})


@app.route("/api/stage2/download/<job_id>")
def api_stage2_download(job_id: str):
    if job_id not in jobs:
        return jsonify({"error": "not found"}), 404
    job      = jobs[job_id]
    filename = request.args.get("filename") or job.get("output_file")
    if not filename:
        return jsonify({"error": "no output file yet"}), 404

    path = job["job_dir"] / filename
    if not path.exists():
        return jsonify({"error": "file missing"}), 404
    # Only allow files inside the job dir (prevent path escape)
    try:
        path.resolve().relative_to(job["job_dir"].resolve())
    except ValueError:
        return jsonify({"error": "invalid filename"}), 400
    return send_file(str(path), as_attachment=True, download_name=filename)



# Stage 3 ──

@app.route("/api/stage3/run", methods=["POST"])
def api_stage3_run():
    if not ARGUS_API_KEY or not ARGUS_BASE_URL:
        return jsonify({"error": "ARGUS_API_KEY / ARGUS_BASE_URL not set in .env"}), 500
    excel_files = request.files.getlist("excels")
    excel_files = [f for f in excel_files if f.filename]
    if not excel_files:
        return jsonify({"error": "No Excel file uploaded"}), 400

    script_files = request.files.getlist("scripts")
    script_files = [f for f in script_files if f.filename]
    if not script_files:
        return jsonify({"error": "No script file uploaded — needed for Line audit"}), 400

    job_id  = str(uuid.uuid4())
    job_dir = S3_WORKSPACE / job_id
    job_dir.mkdir(parents=True, exist_ok=True)

    saved_paths = []
    for ef in excel_files:
        p = job_dir / Path(ef.filename).name
        ef.save(str(p))
        saved_paths.append(p)

    script_file = script_files[0]
    script_path = job_dir / Path(script_file.filename).name
    script_file.save(str(script_path))
    try:
        script_text = _s3_read_script_file(script_path)
    except Exception as e:
        return jsonify({"error": f"Failed to read script: {e}"}), 400
    if not script_text.strip():
        return jsonify({"error": "Script file is empty"}), 400

    log_queue = queue.Queue()
    jobs[job_id] = {
        "status": "running", "queue": log_queue, "job_dir": job_dir,
        "tokens": {"input": 0, "output": 0, "calls": 0},
        "tokens_lock": threading.Lock(),
        "output_file": None, "output_files": [],
        "file_tokens": [],
        "progress": {"total": 0, "completed": 0, "stage": "Starting",
                     "started_at": time.time()},
    }
    threading.Thread(
        target=run_stage3_batch,
        args=(job_id, job_dir, saved_paths, script_text),
        daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/api/stage3/download/<job_id>")
def api_stage3_download(job_id: str):
    fname = request.args.get("filename", "")
    if not fname or job_id not in jobs:
        return "Not found", 404
    p = S3_WORKSPACE / job_id / fname
    if not p.exists():
        return "File not found", 404
    return send_file(str(p), as_attachment=True, download_name=fname)


@app.route("/api/stage3/fix/<job_id>", methods=["POST"])
def api_stage3_fix(job_id: str):
    if job_id not in jobs:
        return jsonify({"error": "Job not found"}), 404
    job = jobs[job_id]
    if job["status"] != "flagged":
        return jsonify({"error": "Job is not in flagged state"}), 400
    flags = job.get("flags")
    if not flags:
        return jsonify({"error": "No flags data"}), 400

    excel_path = Path(job.get("excel_path", ""))
    if not excel_path.exists():
        return jsonify({"error": "Source Excel not found"}), 400

    job_dir = job["job_dir"]
    all_fixes = flags.get("fixes", [])
    all_deletes = flags.get("delete_rows", [])

    output_name = f"{excel_path.stem}_audited.xlsx"
    output_path = job_dir / output_name

    stats = _s3_apply_audit_fixes(all_fixes, all_deletes, excel_path, output_path)

    job["output_files"] = [output_name]
    job["output_file"] = output_name
    job["status"] = "done"

    return jsonify({
        "status": "done",
        "output_file": output_name,
        "output_files": [output_name],
        "stats": stats,
    })


# Stage 4 ──

@app.route("/api/stage4/run", methods=["POST"])
def api_stage4_run():
    if not MADEYE_API_KEY or not MADEYE_BASE_URL:
        return jsonify({"error": "MADEYE_API_KEY / MADEYE_BASE_URL not set in .env"}), 500

    char_canvas = request.files.getlist("char_canvas")
    loc_canvas  = request.files.getlist("loc_canvas")
    char_canvas = [f for f in char_canvas if f.filename]
    loc_canvas  = [f for f in loc_canvas if f.filename]

    if not char_canvas and not loc_canvas:
        return jsonify({"error": "Upload at least one canvas file"}), 400

    job_id  = str(uuid.uuid4())
    job_dir = S4_WORKSPACE / job_id
    job_dir.mkdir(parents=True, exist_ok=True)

    items = []
    for cf in char_canvas:
        p = job_dir / Path(cf.filename).name
        cf.save(str(p))
        items.extend(_s4_parse_md(p))
    for lf in loc_canvas:
        p = job_dir / Path(lf.filename).name
        lf.save(str(p))
        items.extend(_s4_parse_md(p))

    if not items:
        return jsonify({"error": "No valid items found in canvas files"}), 400

    log_queue = queue.Queue()
    jobs[job_id] = {
        "status": "running", "queue": log_queue, "job_dir": job_dir,
        "tokens": {"input": 0, "output": 0, "calls": 0},
        "tokens_lock": threading.Lock(),
        "output_file": None, "output_files": [],
        "file_tokens": [],
        "progress": {"total": 0, "completed": 0, "stage": "Starting",
                     "started_at": time.time()},
    }
    threading.Thread(
        target=run_stage4_pipeline,
        args=(job_id, job_dir, items),
        daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/api/stage4/image/<job_id>/<filename>")
def api_stage4_image(job_id: str, filename: str):
    if job_id not in jobs:
        return "Job not found", 404
    path = S4_WORKSPACE / job_id / "images" / filename
    if not path.exists():
        return "Image not found", 404
    return send_file(str(path), mimetype="image/png")


@app.route("/api/stage4/download/<job_id>")
def api_stage4_download(job_id: str):
    fname = request.args.get("filename", "")
    if not fname or job_id not in jobs:
        return "Not found", 404
    p = S4_WORKSPACE / job_id / fname
    if not p.exists():
        return "File not found", 404
    return send_file(str(p), as_attachment=True, download_name=fname)


@app.route("/api/stage4/shots/<job_id>")
def api_stage4_shots(job_id: str):
    if job_id not in jobs:
        return jsonify([])
    items = jobs[job_id].get("s4_items", [])
    result = []
    for i, item in enumerate(items, 1):
        gen_url = item.get("generated_url", "")
        result.append({
            "num": i,
            "name": f"{item['type'].title()}: {item['name']}",
            "has_image": bool(gen_url),
            "preview_url": gen_url,
        })
    return jsonify(result)


@app.route("/api/stage4/regen", methods=["POST"])
def api_stage4_regen():
    data = request.get_json(force=True)
    orig_job_id = data.get("job_id")
    feedback    = data.get("feedback", "")
    if not orig_job_id or orig_job_id not in jobs:
        return jsonify({"error": "Invalid job_id"}), 400

    items = jobs[orig_job_id].get("s4_items", [])
    if not items:
        return jsonify({"error": "No items in original job"}), 400

    nums = set()
    for m in re.finditer(r'(?:item|character|location|#)\s*(\d+)', feedback, re.I):
        nums.add(int(m.group(1)))
    if not nums:
        for m in re.finditer(r'(\d+)', feedback):
            n = int(m.group(1))
            if 1 <= n <= len(items):
                nums.add(n)
    if not nums:
        return jsonify({"error": "Could not determine which items to regenerate from feedback"}), 400

    regen_items = []
    for n in sorted(nums):
        if 1 <= n <= len(items):
            regen_items.append((n, items[n - 1]))

    regen_job_id = str(uuid.uuid4())
    regen_dir = S4_WORKSPACE / orig_job_id
    log_queue = queue.Queue()
    jobs[regen_job_id] = {
        "status": "running", "queue": log_queue, "job_dir": regen_dir,
        "tokens": {"input": 0, "output": 0, "calls": 0},
        "tokens_lock": threading.Lock(),
        "output_file": None, "output_files": [],
        "progress": {"total": len(regen_items), "completed": 0, "stage": "Regenerating",
                     "started_at": time.time()},
    }

    def _regen():
        _set_job_context(regen_job_id, log_queue)
        try:
            for idx, (num, item) in enumerate(regen_items, 1):
                _log(f"  [{idx}/{len(regen_items)}]  Regenerating: {item['name']}")
                jobs[regen_job_id]["progress"]["completed"] = idx - 1

                itype = item.get("type", "character")
                name = item.get("name", "")
                desc = item.get("description", "")
                if itype == "character":
                    new_prompt = (
                        f"{feedback}. "
                        "Professional fashion editorial photograph, real human adult with realistic adult body proportions. "
                        "The subject's head is small relative to the body, exactly 1/7.5 of total height. "
                        "Long legs, normal-sized torso, photographed from mid-distance showing full body head to shoes. "
                        f"Subject: {name}. {desc} "
                        "Shot on 85mm lens, studio lighting, clean neutral background. "
                        "8K photorealistic. PURE IMAGE ONLY — zero text, no words, no watermark."
                    )
                else:
                    new_prompt = (
                        f"{feedback}. "
                        f"8K photorealistic wide establishing shot of {name}. {desc} "
                        "Cinematic composition, detailed environment, atmospheric lighting. "
                        "PURE IMAGE ONLY — zero text, no words, no letters, no watermark."
                    )
                url = _s4_generate_one(new_prompt, orig_job_id, num, regen_dir, regen_job_id=regen_job_id)
                if url:
                    item["generated_url"] = url
                    _log(f"  ok  Regenerated {item['name']}")
                else:
                    _log(f"  !!  Failed to regenerate {item['name']}")

                if idx < len(regen_items):
                    time.sleep(S4_RATE_LIMIT_DELAY)

            jobs[regen_job_id]["progress"]["completed"] = len(regen_items)

            output_name = jobs[orig_job_id].get("output_file")
            if output_name:
                output_path = regen_dir / output_name
                _s4_write_output_excel(items, output_path)

            jobs[regen_job_id]["status"] = "done"
            t = jobs[regen_job_id]["tokens"]
            _log(f"  ==  Regen done — {t['output']} images | ${t['output'] * S4_COST_PER_IMAGE:.2f}")
        except Exception as e:
            _log(f"[ERROR] Stage 4 regen failed: {e}")
            jobs[regen_job_id]["status"] = "failed"
        finally:
            log_queue.put(None)

    threading.Thread(target=_regen, daemon=True).start()
    return jsonify({"job_id": regen_job_id, "items": [n for n, _ in regen_items]})


# Stage 5 ──

@app.route("/api/stage5/run", methods=["POST"])
def api_stage5_run():
    if not MADEYE_API_KEY or not MADEYE_BASE_URL:
        return jsonify({"error": "MADEYE_API_KEY / MADEYE_BASE_URL not set in .env"}), 500
    excel_files = request.files.getlist("excels")
    char_files  = request.files.getlist("char_images")
    loc_files   = request.files.getlist("loc_images")
    excel_files = [f for f in excel_files if f.filename]
    char_files  = [f for f in char_files if f.filename]
    loc_files   = [f for f in loc_files if f.filename]
    if not excel_files:
        return jsonify({"error": "No Excel file uploaded"}), 400

    job_id  = str(uuid.uuid4())
    job_dir = S5_WORKSPACE / job_id
    job_dir.mkdir(parents=True, exist_ok=True)

    ep = job_dir / Path(excel_files[0].filename).name
    excel_files[0].save(str(ep))

    char_dir = job_dir / "char_refs"
    char_dir.mkdir(exist_ok=True)
    for cf in char_files:
        fname = Path(cf.filename).name
        if fname.lower().endswith(".xlsx"):
            cf.save(str(char_dir / fname))
        else:
            cf.save(str(char_dir / fname))

    loc_dir = job_dir / "loc_refs"
    loc_dir.mkdir(exist_ok=True)
    for lf in loc_files:
        fname = Path(lf.filename).name
        if fname.lower().endswith(".xlsx"):
            lf.save(str(loc_dir / fname))
        else:
            lf.save(str(loc_dir / fname))

    log_queue = queue.Queue()
    jobs[job_id] = {
        "status": "running", "queue": log_queue, "job_dir": job_dir,
        "tokens": {"input": 0, "output": 0, "calls": 0},
        "tokens_lock": threading.Lock(),
        "output_file": None, "output_files": [],
        "file_tokens": [],
        "progress": {"total": 0, "completed": 0, "stage": "Starting",
                     "started_at": time.time()},
    }
    threading.Thread(
        target=run_stage5_pipeline,
        args=(job_id, job_dir, ep, char_dir, loc_dir),
        daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/api/stage5/image/<job_id>/<filename>")
def api_stage5_image(job_id: str, filename: str):
    if job_id not in jobs:
        path = S5_WORKSPACE / job_id / "images" / filename
    else:
        path = jobs[job_id]["job_dir"] / "images" / filename
    if not path.exists():
        return jsonify({"error": "not found"}), 404
    return send_file(str(path), mimetype="image/png")


@app.route("/api/stage5/download/<job_id>")
def api_stage5_download(job_id: str):
    if job_id not in jobs:
        return jsonify({"error": "not found"}), 404
    job = jobs[job_id]
    filename = request.args.get("filename") or job.get("output_file")
    if not filename:
        return jsonify({"error": "no output file yet"}), 404
    path = job["job_dir"] / filename
    if not path.exists():
        hits = list(job["job_dir"].rglob(filename))
        path = hits[0] if hits else path
    if not path.exists():
        return jsonify({"error": "file missing"}), 404
    try:
        path.resolve().relative_to(job["job_dir"].resolve())
    except ValueError:
        return jsonify({"error": "invalid filename"}), 400
    return send_file(str(path), as_attachment=True, download_name=filename)


@app.route("/api/stage5/shots/<job_id>")
def api_stage5_shots(job_id: str):
    if job_id not in jobs:
        return jsonify({"error": "not found"}), 404
    shots = jobs[job_id].get("s5_shots", [])
    result = []
    for s in shots:
        sn = s.get("num", "")
        result.append({
            "num": sn,
            "line": (s.get("line", "")[:80] + "...") if len(s.get("line", "")) > 80 else s.get("line", ""),
            "preview_1": s.get("preview_1", ""),
            "has_image": bool(s.get("preview_1")),
        })
    return jsonify(result)


@app.route("/api/stage5/regen", methods=["POST"])
def api_stage5_regen():
    body = request.get_json(force=True) or {}
    job_id   = body.get("job_id", "")
    feedback = body.get("feedback", "").strip()
    if not job_id or job_id not in jobs:
        return jsonify({"error": "Job not found"}), 404
    if not feedback:
        return jsonify({"error": "No feedback provided"}), 400

    job = jobs[job_id]
    prompts = job.get("s5_prompts", {})
    shots   = job.get("s5_shots", [])
    if not prompts or not shots:
        return jsonify({"error": "No shot data — run Stage 5 first"}), 400

    shot_nums = set()
    for m in re.finditer(r'(?:shot\s*#?\s*|^#?\s*)(\d+)', feedback, re.IGNORECASE | re.MULTILINE):
        shot_nums.add(int(m.group(1)))

    if not shot_nums:
        return jsonify({"error": "Could not find shot numbers in feedback. Use format: Shot 12: description of issue"}), 400

    regen_queue = queue.Queue()
    regen_job_id = f"{job_id}_regen_{int(time.time())}"
    jobs[regen_job_id] = {
        "status": "running", "queue": regen_queue, "job_dir": job["job_dir"],
        "tokens": dict(job.get("tokens", {"input": 0, "output": 0, "calls": 0})),
        "tokens_lock": threading.Lock(),
        "output_file": job.get("output_file"), "output_files": job.get("output_files", []),
        "progress": {"total": len(shot_nums), "completed": 0, "stage": "Starting regen"},
        "s5_shots": shots, "s5_prompts": prompts,
    }

    def _run_regen():
        _set_job_context(regen_job_id, regen_queue)
        try:
            _log(f"\n{'=' * 60}")
            _log(f"  STAGE 3 — Regenerating {len(shot_nums)} shot(s)")
            _log(f"  Feedback: {feedback[:200]}")
            _log(f"{'=' * 60}")

            regen_count = 0
            for i, sn in enumerate(sorted(shot_nums)):
                orig_prompt = prompts.get(sn, "")
                if not orig_prompt:
                    _log(f"  !!  Shot {sn}: no original prompt found — skipping")
                    jobs[regen_job_id]["progress"]["completed"] = i + 1
                    continue

                shot_feedback = ""
                for line in feedback.split("\n"):
                    if re.search(rf'(?:shot\s*#?\s*)?{sn}\b', line, re.IGNORECASE):
                        shot_feedback += re.sub(rf'^.*?{sn}\s*[:—\-]\s*', '', line, flags=re.IGNORECASE).strip() + " "

                if not shot_feedback.strip():
                    shot_feedback = feedback

                regen_prompt = (
                    orig_prompt +
                    f"\n\nFEEDBACK — fix the following issues: {shot_feedback.strip()}"
                )

                jobs[regen_job_id]["progress"]["stage"] = f"Regenerating shot {sn} ({i+1}/{len(shot_nums)})"
                _log(f"  ... Regenerating shot {sn}: {shot_feedback.strip()[:100]}")

                url = _s5_generate_one(regen_prompt, job_id, sn, job["job_dir"], regen_job_id=regen_job_id)
                if url:
                    for s in shots:
                        if s.get("num") == sn:
                            s["preview_1"] = url
                            break
                    regen_count += 1
                    _log(f"  ok  Shot {sn} regenerated")
                else:
                    _log(f"  x   Shot {sn}: regeneration failed")

                jobs[regen_job_id]["progress"]["completed"] = i + 1

                if i < len(shot_nums) - 1:
                    time.sleep(S5_RATE_LIMIT_DELAY)

            excel_path = Path(job.get("s5_excel_path", ""))
            if excel_path.exists():
                output_name = job.get("output_file", f"{excel_path.stem}_with_images.xlsx")
                output_path = job["job_dir"] / output_name
                _s5_write_output_excel(shots, excel_path, output_path, job_dir)
                _log(f"  ok  Excel updated: {output_name}")

            tok = jobs[regen_job_id]["tokens"]
            regen_cost = regen_count * S5_COST_PER_IMAGE
            _log(f"\n{'=' * 60}")
            _log(f"  REGEN COMPLETE — {regen_count}/{len(shot_nums)} shots regenerated")
            _log(f"  API calls: {tok['calls']}  |  Images: {regen_count}  |  Est. cost: ${regen_cost:.2f}")
            _log(f"{'=' * 60}\n")

            jobs[regen_job_id]["status"] = "done"
            job["s5_shots"] = shots

        except Exception as e:
            _log(f"\n[ERROR] {e}\n{traceback.format_exc()}")
            jobs[regen_job_id]["status"] = "failed"
        finally:
            regen_queue.put(None)

    threading.Thread(target=_run_regen, daemon=True).start()
    return jsonify({"job_id": regen_job_id, "shots": sorted(shot_nums)})


@app.route("/api/stage5/resume", methods=["POST"])
def api_stage5_resume():
    """Resume a Stage 5 job that was interrupted — only generates missing images."""
    data = request.get_json(force=True) or {}
    orig_job_id = data.get("job_id", "")
    if not orig_job_id:
        return jsonify({"error": "Provide job_id"}), 400

    job_dir = S5_WORKSPACE / orig_job_id
    if not job_dir.exists():
        return jsonify({"error": f"Job directory not found: {orig_job_id}"}), 404

    img_dir = job_dir / "images"
    existing_imgs = set()
    if img_dir.exists():
        for f in img_dir.iterdir():
            if f.name.startswith("shot_") and f.suffix == ".png":
                existing_imgs.add(f.stem.replace("shot_", ""))

    excel_files = list(job_dir.glob("*.xlsx"))
    if not excel_files:
        return jsonify({"error": "No Excel file found in job directory"}), 400
    excel_path = excel_files[0]

    shots = _s5_read_excel(excel_path)
    missing = [s for s in shots if str(s["num"]) not in existing_imgs]

    if not missing:
        return jsonify({"error": "All shots already have images — nothing to resume"}), 400

    char_dir = job_dir / "char_refs"
    loc_dir  = job_dir / "loc_refs"

    resume_job_id = f"{orig_job_id}_resume_{int(time.time())}"
    log_queue = queue.Queue()
    jobs[resume_job_id] = {
        "status": "running", "queue": log_queue, "job_dir": job_dir,
        "tokens": {"input": 0, "output": 0, "calls": 0},
        "tokens_lock": threading.Lock(),
        "output_file": None, "output_files": [],
        "progress": {"total": len(missing), "completed": 0, "stage": "Resuming"},
        "s5_shots": shots, "s5_prompts": {},
    }

    def _run_resume():
        _set_job_context(resume_job_id, log_queue)
        try:
            _log("=" * 60)
            _log(f"  STAGE 5 RESUME — {len(missing)} missing shots (of {len(shots)} total)")
            _log(f"  Already generated: {len(existing_imgs)} images")
            _log("=" * 60)

            ref_map = _s5_build_ref_map(char_dir)
            if ref_map:
                _log(f"  ok  {len(ref_map)} character reference(s)")

            loc_map = {}
            if loc_dir.exists():
                loc_map = _s5_build_ref_map(loc_dir)
                if loc_map:
                    _log(f"  ok  {len(loc_map)} location reference(s)")

            generated = 0
            last_char = ""
            for i, shot in enumerate(missing):
                sn = shot["num"]
                desc = shot["shot_description"]
                if not desc:
                    _log(f"  !!  Shot {sn}: no description — skipping")
                    jobs[resume_job_id]["progress"]["completed"] = i + 1
                    continue

                ref_names, ref_instruction, last_char, ref_imgs = _s5_select_refs(shot, ref_map, last_char)

                if loc_map:
                    loc_names = _s5_match_characters(desc, shot.get("shot_detail", ""), loc_map)
                    for ln in loc_names:
                        entry = loc_map.get(ln, {})
                        lpath = entry.get("path")
                        if lpath and Path(lpath).exists():
                            ref_imgs.append({"name": f"Location: {ln}", "path": Path(lpath)})

                prompt = S5_PROMPT_PREFIX + ref_instruction + desc + " " + S5_NEGATIVE + S5_PROMPT_SUFFIX
                jobs[resume_job_id]["s5_prompts"][sn] = prompt

                jobs[resume_job_id]["progress"]["stage"] = f"Shot {sn} ({i+1}/{len(missing)})"
                ref_label = f" chars=[{', '.join(ref_names)}]" if ref_names else ""
                _log(f"  ... Shot {sn} ({i+1}/{len(missing)}){ref_label}")

                url = _s5_generate_one(prompt, orig_job_id, sn, job_dir,
                                        regen_job_id=resume_job_id, ref_images=ref_imgs)
                if url:
                    shot["preview_1"] = url
                    for s in shots:
                        if s["num"] == sn:
                            s["preview_1"] = url
                            break
                    generated += 1
                    _log(f"  ok  Shot {sn} saved")
                else:
                    _log(f"  x   Shot {sn}: generation failed")

                jobs[resume_job_id]["progress"]["completed"] = i + 1

                if i < len(missing) - 1:
                    _log(f"  ... Rate-limit pause ({S5_RATE_LIMIT_DELAY}s)...")
                    time.sleep(S5_RATE_LIMIT_DELAY)

            _log(f"\n  Writing output Excel...")
            output_name = f"{excel_path.stem}_with_images.xlsx"
            output_path = job_dir / output_name
            _s5_write_output_excel(shots, excel_path, output_path, job_dir)

            jobs[resume_job_id]["output_file"] = output_name
            jobs[resume_job_id]["output_files"] = [output_name]
            jobs[resume_job_id]["status"] = "done"

            t = jobs[resume_job_id]["tokens"]
            _log(f"  ==  Resume done — {generated}/{len(missing)} generated | ${t['output'] * 0.02:.2f}")

        except Exception as e:
            import traceback
            _log(f"\n[ERROR] {e}\n{traceback.format_exc()}")
            jobs[resume_job_id]["status"] = "failed"
        finally:
            log_queue.put(None)

    threading.Thread(target=_run_resume, daemon=True).start()
    return jsonify({
        "job_id": resume_job_id,
        "total_shots": len(shots),
        "already_done": len(existing_imgs),
        "remaining": len(missing),
    })


# Pipeline ──

@app.route("/api/pipeline/run", methods=["POST"])
def api_pipeline_run():
    if not ARGUS_API_KEY or not ARGUS_BASE_URL:
        return jsonify({"error": "ARGUS_API_KEY / ARGUS_BASE_URL not set"}), 500

    files = request.files.getlist("scripts")
    show_name = request.form.get("show_name", "").strip() or "Untitled Show"
    workers = int(request.form.get("workers", "3") or "3")

    if not files or all(f.filename == "" for f in files):
        return jsonify({"error": "No script files uploaded"}), 400

    job_id = str(uuid.uuid4())
    job_dir = WORKSPACE / f"pipeline_{job_id}"
    job_dir.mkdir(parents=True, exist_ok=True)

    scripts = []
    for f in files:
        if f.filename:
            text = _read_upload_as_text(f)
            scripts.append({
                "filename": f.filename,
                "name": Path(f.filename).stem,
                "text": text,
            })

    log_queue = queue.Queue()
    jobs[job_id] = {
        "status": "running",
        "queue": log_queue,
        "job_type": "pipeline",
        "job_dir": job_dir,
        "tokens": {"input": 0, "output": 0, "calls": 0},
        "tokens_lock": threading.Lock(),
        "output_file": None,
        "output_files": [],
        "file_tokens": [],
        "progress": {
            "pipeline_stage": "starting",
            "total": 5,
            "completed": 0,
            "stage": "Initializing pipeline",
            "started_at": time.time(),
        },
    }

    threading.Thread(
        target=run_pipeline,
        args=(job_id, job_dir, scripts, show_name, workers),
        daemon=True,
    ).start()

    return jsonify({"job_id": job_id})


# Stage 6 ──

@app.route("/api/stage6/run", methods=["POST"])
def api_stage6_run():
    excel_files   = request.files.getlist("excels")
    audio_files   = request.files.getlist("audios")
    subs          = request.form.get("subs", "yes")
    whisper_model = request.form.get("whisper_model", "base.en")
    sub_color     = request.form.get("sub_color", "yellow")
    excel_files = [f for f in excel_files if f.filename]
    audio_files = [f for f in audio_files if f.filename]
    if not excel_files:
        return jsonify({"error": "No Excel files uploaded"}), 400
    if not audio_files:
        return jsonify({"error": "No audio files uploaded"}), 400
    if len(excel_files) != len(audio_files):
        return jsonify({"error": f"Mismatch: {len(excel_files)} Excel but {len(audio_files)} audio"}), 400
    job_id  = str(uuid.uuid4())
    job_dir = S6_WORKSPACE / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    pairs = []
    for i, (ef, af) in enumerate(zip(excel_files, audio_files)):
        ep_dir = job_dir / f"pair_{i + 1}"
        ep_dir.mkdir(exist_ok=True)
        ep = ep_dir / Path(ef.filename).name
        ap = ep_dir / Path(af.filename).name
        ef.save(str(ep)); af.save(str(ap))
        pairs.append((ep, ap))
    steps_per = 6 if subs == "yes" else 4
    total = steps_per * len(pairs)
    log_queue = queue.Queue()
    jobs[job_id] = {
        "status": "running", "queue": log_queue, "job_dir": job_dir,
        "tokens": {"input": 0, "output": 0, "calls": 0},
        "tokens_lock": threading.Lock(),
        "output_file": None, "output_files": [],
        "file_tokens": [],
        "progress": {"total": total, "completed": 0, "stage": "Starting",
                     "started_at": time.time(), "pair_count": len(pairs)},
    }
    threading.Thread(
        target=run_stage6_batch,
        args=(job_id, job_dir, pairs, whisper_model, subs == "yes", sub_color),
        daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/api/stage6/download/<job_id>")
def api_stage6_download(job_id: str):
    if job_id not in jobs:
        return jsonify({"error": "not found"}), 404
    job = jobs[job_id]
    filename = request.args.get("filename") or job.get("output_file")
    if not filename:
        return jsonify({"error": "no output file yet"}), 404
    path = job["job_dir"] / filename
    if not path.exists():
        hits = list(job["job_dir"].rglob(filename))
        path = hits[0] if hits else path
    if not path.exists():
        return jsonify({"error": "file missing"}), 404
    try:
        path.resolve().relative_to(job["job_dir"].resolve())
    except ValueError:
        return jsonify({"error": "invalid filename"}), 400
    return send_file(str(path), as_attachment=True, download_name=filename)

# Database ──

@app.route("/api/db/shows")
def api_db_shows():
    try:
        shows = _supabase("GET", "/shows", params={
            "select": "id,name,created_at",
            "order":  "created_at.desc",
        })
        return jsonify(shows)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/db/shows/<show_id>/files")
def api_db_show_files(show_id: str):
    try:
        files = _supabase("GET", "/show_files", params={
            "select":  "id,file_type,filename,content,created_at",
            "show_id": f"eq.{show_id}",
            "order":   "file_type.asc,created_at.asc",
        })
        return jsonify(files)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/db/save", methods=["POST"])
def api_db_save():
    body      = request.get_json(force=True) or {}
    job_id    = body.get("job_id", "")
    show_name = body.get("show_name", "").strip()

    if not show_name:
        return jsonify({"error": "Show name required"}), 400
    if job_id not in jobs:
        return jsonify({"error": "Job not found — job may have expired"}), 404

    job_dir = jobs[job_id]["job_dir"]
    try:
        result  = _supabase("POST", "/shows", {"name": show_name})
        show_id = result[0]["id"]

        saved = 0
        for subfolder in ["show level files", "episode details"]:
            d = job_dir / subfolder
            if d.exists():
                for f in sorted(d.glob("*.md")):
                    _supabase("POST", "/show_files", {
                        "show_id":   show_id,
                        "file_type": _classify_file_type(f.name),
                        "filename":  f.name,
                        "content":   f.read_text(encoding="utf-8"),
                    })
                    saved += 1

        return jsonify({"show_id": show_id, "show_name": show_name, "files_saved": saved})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# Shared ──

@app.route("/api/stream/<job_id>")
def api_stream(job_id: str):
    if job_id not in jobs:
        return Response("data: [job not found]\n\nevent: done\ndata:\n\n", mimetype="text/event-stream")
    log_queue = jobs[job_id]["queue"]

    def generate():
        while True:
            try:
                line = log_queue.get(timeout=30)
            except queue.Empty:
                yield "data: [waiting...]\n\n"
                continue
            if line is None:
                yield "event: done\ndata:\n\n"
                break
            yield f"data: {line.replace(chr(10), ' ')}\n\n"

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/api/status/<job_id>")
def api_status(job_id: str):
    if job_id not in jobs:
        return jsonify({"error": "not found"}), 404
    job   = jobs[job_id]
    files = _collect_stage1_files(job["job_dir"])
    prog = dict(job.get("progress", {}))
    if prog.get("started_at"):
        elapsed = time.time() - prog["started_at"]
        prog["elapsed_sec"] = int(elapsed)
        done  = prog.get("completed", 0)
        total = prog.get("total", 0)
        if done > 0 and total > 0 and elapsed > 0:
            rate    = done / elapsed
            remaining = total - done
            prog["eta_sec"] = int(remaining / rate) if rate > 0 and remaining > 0 else 0
        prog.pop("started_at", None)  # don't leak raw timestamp

    resp = {
        "status":       job["status"],
        "files":        files,
        "output_file":  job.get("output_file"),
        "output_files": job.get("output_files", []),
        "sheet_url":    job.get("sheet_url", ""),
        "progress":     prog,
        "tokens":       job.get("tokens", {"input": 0, "output": 0, "calls": 0}),
        "file_tokens":  job.get("file_tokens", []),
        "total_cost":   round(sum(ft.get("cost", 0) for ft in job.get("file_tokens", [])), 4),
    }
    if job.get("flags"):
        resp["flags"] = job["flags"]
    if job.get("job_type") == "pipeline":
        resp["job_type"] = "pipeline"
        resp["pipeline_stage"] = prog.get("pipeline_stage", "")
    return jsonify(resp)


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("\n  FRAMES — Shot Generation Pipeline")
    print(f"  Argus : {ARGUS_MODEL or '(not set)'}")
    print(f"  MadEye: {len(MADEYE_API_KEYS)} key(s) — {MADEYE_BASE_URL or '(no URL)'}")
    print("  Open  : http://localhost:5000\n")
    app.run(host="127.0.0.1", port=5000, debug=False, threaded=True)
