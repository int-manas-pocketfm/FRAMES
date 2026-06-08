"""One-shot Stage 4 video generation script."""
import re, shutil, subprocess, ssl, urllib.request, time
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from difflib import SequenceMatcher

import openpyxl
from PIL import Image
from faster_whisper import WhisperModel

# ── Config ──
EXCEL = Path(r"C:\Users\manas\OneDrive\Desktop\POCKETFM PACKETS\SHOWS\disaster files\shots with images\Ep_1_-_The_Sinking_of_the_Titanic_-_Part_1_breakdown_FIXED (1).xlsx")
AUDIO = Path(r"C:\Users\manas\OneDrive\Desktop\POCKETFM PACKETS\SHOWS\disaster files\audio\Ep 1 - The Sinking of the Titanic - Part 1.mp3")
WHISPER_MODEL = "small.en"
OUTDIR = Path(r"C:\Users\manas\OneDrive\Desktop\POCKETFM PACKETS\SHOWS\disaster files\video_output")

VID_W, VID_H = 1080, 1080
VID_FPS = 24
VID_CRF = 18
VID_AUDIO_BITRATE = "192k"
VID_FONT_SIZE = 105
VID_CHARS_PER_LINE = 22

_ASS_STYLE = (
    f"Style: Default,Poppins,{VID_FONT_SIZE},"
    "&H0000C4FF,&H0000FFFF,&H00000000,&H00000000,"
    "-1,0,0,0,100,100,0,0,1,4,2,2,40,40,60,1"
)


import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

def log(msg):
    print(msg, flush=True)


# ── Step 1: Read Excel ──
def read_excel(path):
    wb = openpyxl.load_workbook(str(path))
    ws = wb["Shot Breakdown"] if "Shot Breakdown" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    log(f"[1] Sheet '{ws.title}' | {ws.max_row - 1} rows | Headers: {[h for h in headers if h]}")

    def _find(names):
        for name in names:
            for i, h in enumerate(headers):
                if h and name.lower() in str(h).lower():
                    return i + 1
        return None

    num_col = _find(["#", "num", "shot"])
    line_col = _find(["line", "narration", "dialogue", "text"])
    url_col = _find(["generated_url", "url", "image_url", "image", "preview"])

    if not line_col:
        raise RuntimeError(f"No line column. Headers: {headers}")
    if not url_col:
        raise RuntimeError(f"No image URL column. Headers: {headers}")

    shots = []
    for r in range(2, ws.max_row + 1):
        line = ws.cell(r, line_col).value
        url = ws.cell(r, url_col).value
        if not line or not url:
            continue
        raw_num = ws.cell(r, num_col).value if num_col else len(shots) + 1
        try:
            seq = int(float(str(raw_num)))
        except (ValueError, TypeError):
            seq = len(shots) + 1
        shots.append({
            "num": seq, "idx": len(shots) + 1,
            "line": str(line).strip(), "url": str(url).strip(),
        })
    log(f"[1] {len(shots)} shots parsed")
    return shots


# ── Step 2: Download images ──
def download_images(shots, img_dir):
    img_dir.mkdir(parents=True, exist_ok=True)
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    def _dl(shot):
        dest = img_dir / f"{shot['idx']:04d}.png"
        shot["img_path"] = str(dest)
        if dest.exists() and dest.stat().st_size > 1000:
            return "cached"
        try:
            req = urllib.request.Request(shot["url"], headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, context=ctx, timeout=30) as resp:
                dest.write_bytes(resp.read())
            return "ok"
        except Exception as e:
            log(f"  x  Shot {shot['idx']}: {e}")
            shot["img_path"] = None
            return "fail"

    ok = fail = cached = 0
    with ThreadPoolExecutor(max_workers=10) as pool:
        futs = {pool.submit(_dl, s): s for s in shots}
        for fut in as_completed(futs):
            r = fut.result()
            if r == "ok": ok += 1
            elif r == "fail": fail += 1
            else: cached += 1
            done = ok + fail + cached
            if done % 50 == 0 or done == len(shots):
                log(f"  ... {done}/{len(shots)}")
    log(f"[2] {ok} downloaded, {cached} cached, {fail} failed")


# ── Step 3: Resize to 1080x1080 ──
def resize_images(shots, img_dir):
    for shot in shots:
        src = shot.get("img_path")
        if not src or not Path(src).exists():
            shot["sq_path"] = None
            continue
        out = img_dir / f"{shot['idx']:04d}_r.png"
        shot["sq_path"] = str(out)
        if out.exists():
            continue
        try:
            img = Image.open(src).convert("RGB")
            ow, oh = img.size
            scale = max(VID_W / ow, VID_H / oh)
            nw, nh = int(ow * scale), int(oh * scale)
            img = img.resize((nw, nh), Image.LANCZOS)
            left, top = (nw - VID_W) // 2, (nh - VID_H) // 2
            img.crop((left, top, left + VID_W, top + VID_H)).save(str(out), quality=95)
        except Exception as e:
            log(f"  x  Resize shot {shot['idx']}: {e}")
            shot["sq_path"] = None
    valid = sum(1 for s in shots if s.get("sq_path") and Path(s["sq_path"]).exists())
    log(f"[3] {valid}/{len(shots)} resized to {VID_W}x{VID_H}")


# ── Step 4: Align audio with Whisper ──
def align_audio(shots, audio_path, whisper_model):
    log(f"[4] Loading Whisper '{whisper_model}'...")
    model = WhisperModel(whisper_model, device="cpu", compute_type="int8")
    segs, info = model.transcribe(str(audio_path), word_timestamps=True, language="en")

    words = []
    for seg in segs:
        if seg.words:
            for w in seg.words:
                words.append({"word": w.word.strip(), "start": w.start, "end": w.end})
    duration = info.duration
    log(f"[4] {len(words)} words transcribed, {duration:.1f}s total")

    def _clean(t):
        return re.sub(r'[^\w\s]', '', t.lower()).split()

    wc = [re.sub(r'[^\w]', '', w["word"].lower()) for w in words]

    if not words:
        log("[4] !! Whisper returned 0 words — equal durations fallback")
        per = duration / max(len(shots), 1)
        for i, shot in enumerate(shots):
            shot["w_start"] = shot["w_end"] = 0
            shot["start"] = i * per
            shot["end"] = (i + 1) * per
        shots[-1]["end"] = duration
        return duration

    ptr = 0
    for shot in shots:
        lw = _clean(shot["line"])
        nl = len(lw)
        if nl == 0 or ptr >= len(words):
            shot["w_start"] = shot["w_end"] = max(0, min(ptr, len(words) - 1))
            continue
        best_s, best_sc = ptr, -1
        for off in range(min(30, len(words) - ptr)):
            ts = ptr + off
            cl = min(nl, 5, len(wc) - ts)
            if cl <= 0: break
            sc = SequenceMatcher(None, " ".join(lw[:cl]), " ".join(wc[ts:ts + cl])).ratio()
            if sc > best_sc:
                best_sc = sc; best_s = ts
        exp_end = best_s + nl
        best_e = min(exp_end, len(words) - 1)
        if best_e + 6 < len(words):
            best_es = -1
            for eo in range(-3, 7):
                te = exp_end + eo
                if te < best_s or te >= len(words): continue
                cl = min(4, te - best_s + 1)
                if te - cl + 1 < 0: continue
                sc = SequenceMatcher(None, " ".join(lw[-cl:]), " ".join(wc[te - cl + 1:te + 1])).ratio()
                if sc > best_es:
                    best_es = sc; best_e = te
        shot["w_start"] = max(0, min(best_s, len(words) - 1))
        shot["w_end"] = max(0, min(best_e, len(words) - 1))
        ptr = shot["w_end"] + 1

    for shot in shots:
        ws_idx = max(0, min(shot["w_start"], len(words) - 1))
        we_idx = max(0, min(shot["w_end"], len(words) - 1))
        shot["start"] = words[ws_idx]["start"]
        shot["end"] = words[we_idx]["end"]
    for i in range(len(shots) - 1):
        shots[i]["end"] = shots[i + 1]["start"]
    shots[0]["start"] = 0.0
    shots[-1]["end"] = duration
    for i, shot in enumerate(shots):
        if shot["end"] - shot["start"] <= 0:
            if i > 0:
                shot["start"] = shots[i - 1]["end"] - 0.2
                shots[i - 1]["end"] = shot["start"]
            shot["end"] = shot["start"] + 0.5
    log(f"[4] Aligned {len(shots)} shots | 0.00s - {shots[-1]['end']:.2f}s")
    return duration


# ── Step 5: Generate ASS subtitles ──
def ass_time(sec):
    h = int(sec // 3600)
    m = int((sec % 3600) // 60)
    s = sec % 60
    return f"{h}:{m:02d}:{s:05.2f}"


def wrap_text(text):
    mx = VID_CHARS_PER_LINE
    if len(text) <= mx:
        return [text]
    words, chunks = text.split(), []
    l1, l2, in2, i = "", "", False, 0
    while i < len(words):
        w = words[i]
        if not in2:
            t = (l1 + " " + w).strip() if l1 else w
            if len(t) <= mx: l1 = t; i += 1
            else: in2 = True
        else:
            t = (l2 + " " + w).strip() if l2 else w
            if len(t) <= mx: l2 = t; i += 1
            else:
                chunks.append(l1 + ("\\N" + l2 if l2 else ""))
                l1 = l2 = ""; in2 = False
    if l1:
        chunks.append(l1 + ("\\N" + l2 if l2 else ""))
    return chunks or [text[:mx * 2]]


def generate_subtitles(shots, ass_path):
    events = []
    for shot in shots:
        chunks = wrap_text(shot["line"])
        dur = shot["end"] - shot["start"]
        if len(chunks) == 1:
            events.append((shot["start"], shot["end"], chunks[0]))
        else:
            total_c = sum(len(c.replace("\\N", " ")) for c in chunks)
            t = shot["start"]
            for ci, ch in enumerate(chunks):
                cc = len(ch.replace("\\N", " "))
                cd = dur * (cc / total_c) if total_c > 0 else dur / len(chunks)
                ce = t + cd if ci < len(chunks) - 1 else shot["end"]
                events.append((t, ce, ch))
                t = ce
    with open(ass_path, "w", encoding="utf-8-sig") as f:
        f.write("[Script Info]\nTitle: Subtitles\nScriptType: v4.00+\n")
        f.write(f"PlayResX: {VID_W}\nPlayResY: {VID_H}\n")
        f.write("WrapStyle: 2\nScaledBorderAndShadow: yes\n\n")
        f.write("[V4+ Styles]\n")
        f.write("Format: Name, Fontname, Fontsize, PrimaryColour, SecondaryColour, "
                "OutlineColour, BackColour, Bold, Italic, Underline, StrikeOut, "
                "ScaleX, ScaleY, Spacing, Angle, BorderStyle, Outline, Shadow, "
                "Alignment, MarginL, MarginR, MarginV, Encoding\n")
        f.write(f"{_ASS_STYLE}\n\n")
        f.write("[Events]\n")
        f.write("Format: Layer, Start, End, Style, Name, MarginL, MarginR, MarginV, Effect, Text\n")
        for st, en, tx in events:
            f.write(f"Dialogue: 0,{ass_time(st)},{ass_time(en)},Default,,0,0,0,,{tx}\n")
    log(f"[5] {len(events)} subtitle events written")


# ── Step 6: Ensure Poppins-Bold font ──
def ensure_font(font_dir):
    font_dir.mkdir(parents=True, exist_ok=True)
    fp = font_dir / "Poppins-Bold.ttf"
    if fp.exists():
        return str(font_dir)
    log("  Downloading Poppins-Bold.ttf...")
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    req = urllib.request.Request(
        "https://github.com/google/fonts/raw/main/ofl/poppins/Poppins-Bold.ttf",
        headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, context=ctx, timeout=30) as resp:
        fp.write_bytes(resp.read())
    log("  Font downloaded")
    return str(font_dir)


# ── Step 7: Render video with FFmpeg ──
def render_video(shots, concat_path, audio_path, ass_path, font_dir, output_path):
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(concat_path, "w") as f:
        last = None
        for shot in shots:
            sq = shot.get("sq_path")
            if not sq or not Path(sq).exists(): continue
            dur = max(shot["end"] - shot["start"], 0.1)
            sq_fwd = sq.replace("\\", "/")
            f.write(f"file '{sq_fwd}'\nduration {dur:.6f}\n")
            last = sq_fwd
        if last:
            f.write(f"file '{last}'\n")

    base_tmp = str(output_path) + ".tmp.mp4"
    log("[6] Building base video...")
    cmd1 = ["ffmpeg", "-y", "-f", "concat", "-safe", "0", "-i", str(concat_path),
            "-i", str(audio_path), "-r", str(VID_FPS),
            "-c:v", "libx264", "-pix_fmt", "yuv420p",
            "-preset", "ultrafast", "-crf", str(VID_CRF),
            "-c:a", "aac", "-b:a", VID_AUDIO_BITRATE,
            "-shortest", "-movflags", "+faststart", base_tmp]
    r = subprocess.run(cmd1, capture_output=True, text=True)
    if r.returncode != 0:
        raise RuntimeError(f"ffmpeg base pass failed:\n{r.stderr[-800:]}")
    log("[6] Base video done, burning subtitles...")

    ass_esc = str(ass_path).replace("\\", "/").replace(":", "\\:")
    fd_esc = font_dir.replace("\\", "/").replace(":", "\\:")
    vf_attempts = [
        f"ass='{ass_esc}':fontsdir='{fd_esc}'",
        f"ass={ass_esc}:fontsdir={fd_esc}",
        f"subtitles={ass_esc}:fontsdir={fd_esc}",
    ]
    burned = False
    for i, vf in enumerate(vf_attempts):
        log(f"  Subtitle burn attempt {i+1}/{len(vf_attempts)}...")
        cmd2 = ["ffmpeg", "-y", "-i", base_tmp, "-vf", vf,
                "-c:v", "libx264", "-pix_fmt", "yuv420p",
                "-preset", "medium", "-crf", str(VID_CRF),
                "-c:a", "copy", "-movflags", "+faststart", str(output_path)]
        r = subprocess.run(cmd2, capture_output=True, text=True)
        if r.returncode == 0:
            log("[6] Subtitles burned successfully"); burned = True; break
        log(f"  x  Attempt {i+1} failed: {r.stderr[-200:].strip()}")
    if not burned:
        log("[6] !! Subtitle burn failed — delivering without subtitles")
        shutil.copy2(base_tmp, str(output_path))
    Path(base_tmp).unlink(missing_ok=True)

    probe = subprocess.run(
        ["ffprobe", "-v", "quiet", "-show_entries", "format=duration",
         "-of", "csv=p=0", str(output_path)], capture_output=True, text=True)
    try:
        dur = float(probe.stdout.strip())
    except ValueError:
        dur = 0.0
    size_mb = output_path.stat().st_size / (1024 * 1024)
    log(f"[6] Done! {output_path.name} | {dur:.1f}s | {size_mb:.1f} MB")
    return dur


# ── Main ──
if __name__ == "__main__":
    t0 = time.time()
    OUTDIR.mkdir(parents=True, exist_ok=True)
    pair_dir = OUTDIR / "work"
    pair_dir.mkdir(parents=True, exist_ok=True)
    img_dir = pair_dir / "images"
    font_dir = pair_dir / "fonts"
    ass_path = pair_dir / "subtitles.ass"
    concat_path = pair_dir / "concat.txt"

    safe = re.sub(r'[^a-zA-Z0-9_\-]', '_', EXCEL.stem)
    out_path = OUTDIR / f"{safe}_1x1.mp4"

    log("=" * 60)
    log("STAGE 4 — VIDEO GENERATION")
    log("=" * 60)

    log("\n── Step 1: Reading Excel ──")
    shots = read_excel(EXCEL)

    log("\n── Step 2: Downloading images ──")
    download_images(shots, img_dir)

    log("\n── Step 3: Resizing to 1080x1080 ──")
    resize_images(shots, img_dir)

    log("\n── Step 4: Aligning audio (Whisper) ──")
    align_audio(shots, AUDIO, WHISPER_MODEL)

    log("\n── Step 5: Generating subtitles ──")
    font_dir_str = ensure_font(font_dir)
    generate_subtitles(shots, ass_path)

    log("\n── Step 6: Rendering video ──")
    render_video(shots, concat_path, AUDIO, ass_path, font_dir_str, out_path)

    elapsed = time.time() - t0
    log(f"\n{'=' * 60}")
    log(f"COMPLETE in {elapsed:.0f}s")
    log(f"Output: {out_path}")
    log(f"{'=' * 60}")
